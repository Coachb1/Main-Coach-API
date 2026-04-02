"""
excel_exporter.py
─────────────────────────────────────────────────────────────────────────────
Universal multi-tab Excel (and zipped-CSV) exporter.

Drop this file anywhere in your project — zero framework dependencies.
Works standalone or inside Django / Flask / FastAPI.

Quick start
───────────
    from excel_exporter import ExcelExporter, Sheet, Theme

    exporter = ExcelExporter(title="Monthly Report")

    exporter.add_sheet(Sheet(
        name="Sales",
        columns=["Region", "Product", "Revenue", "Units"],
        rows=[
            ["North", "Widget A", 12000, 340],
            ["South", "Widget B",  8500, 210],
        ],
        summary={"Total Revenue": "=SUM(C4:C1000)", "Total Units": "=SUM(D4:D1000)"},
    ))

    # Django:  return exporter.to_django_response("report.xlsx")
    # Flask:   return exporter.to_flask_response("report.xlsx")
    # File:    exporter.save("report.xlsx")
    # Bytes:   buf = exporter.to_bytes()
"""

from __future__ import annotations

import csv
import io
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any, Callable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ─── Theme ────────────────────────────────────────────────────────────────────

@dataclass
class Theme:
    """Color palette + font for the workbook.  Override any field you like."""

    # Header / title bar
    header_bg:      str = "1F3864"   # deep navy
    header_fg:      str = "FFFFFF"
    subheader_bg:   str = "2E75B6"   # mid blue
    subheader_fg:   str = "FFFFFF"

    # Column labels row
    col_label_bg:   str = "2E75B6"
    col_label_fg:   str = "FFFFFF"

    # Body rows
    row_even_bg:    str = "D6E4F0"   # light blue tint
    row_odd_bg:     str = "FFFFFF"
    row_fg:         str = "1A1A2E"

    # Summary row (bottom)
    summary_bg:     str = "1F3864"
    summary_fg:     str = "F4B942"   # gold accent

    # Grid lines
    border_color:   str = "C0C0C0"

    font_name:      str = "Arial"

    # Pre-built theme presets
    @classmethod
    def navy(cls) -> "Theme":
        return cls()  # default

    @classmethod
    def teal(cls) -> "Theme":
        return cls(
            header_bg="0D4F4F", subheader_bg="1A7A7A",
            col_label_bg="1A7A7A", row_even_bg="D0EEEE",
            summary_bg="0D4F4F",
        )

    @classmethod
    def charcoal(cls) -> "Theme":
        return cls(
            header_bg="2C2C2C", subheader_bg="4A4A4A",
            col_label_bg="4A4A4A", row_even_bg="F0F0F0",
            summary_bg="2C2C2C",
        )

    @classmethod
    def purple(cls) -> "Theme":
        return cls(
            header_bg="3B1F6B", subheader_bg="6A3CA8",
            col_label_bg="6A3CA8", row_even_bg="EAE0F5",
            summary_bg="3B1F6B",
        )


# ─── Column descriptor ────────────────────────────────────────────────────────

@dataclass
class Column:
    """Optional per-column config.  Use plain strings for zero-config columns."""

    label:      str
    width:      int | None   = None    # fixed width; None = auto
    fmt:        str | None   = None    # Excel number format string
    align:      str          = "left"  # "left" | "center" | "right"
    bold_col:   bool         = False   # bold every value in this column


# ─── Sheet descriptor ─────────────────────────────────────────────────────────

@dataclass
class Sheet:
    """One tab in the workbook.

    Parameters
    ──────────
    name        Tab label (emoji OK, e.g. "📊 Sales")
    columns     List of column names (str) or Column objects
    rows        Iterable of row sequences — any Python values
    subtitle    Short text shown under the title banner
    summary     Dict of {label: value/formula} appended as a highlighted footer row
    row_filter  Optional callable(row) → bool; falsy rows are skipped
    """

    name:       str
    columns:    Sequence[str | Column]
    rows:       Sequence[Sequence[Any]]
    subtitle:   str                    = ""
    summary:    dict[str, Any]         = field(default_factory=dict)
    row_filter: Callable | None        = None


# ─── Internal style helpers ───────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(name: str, size: int = 10, bold: bool = False,
          italic: bool = False, color: str = "000000") -> Font:
    return Font(name=name, size=size, bold=bold, italic=italic, color=color)

def _align(h: str = "left") -> Alignment:
    return Alignment(horizontal=h, vertical="center", wrap_text=True)

def _border(color: str) -> Border:
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _col_objects(columns: Sequence[str | Column]) -> list[Column]:
    return [c if isinstance(c, Column) else Column(label=c) for c in columns]


# ─── Core sheet writer ────────────────────────────────────────────────────────

def _write_sheet(wb: Workbook, sheet: Sheet, theme: Theme, title: str) -> None:
    ws = wb.create_sheet(sheet.name)
    cols = _col_objects(sheet.columns)
    n = len(cols)

    brd = _border(theme.border_color)

    # ── Row 1: title banner ──────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    c = ws.cell(row=1, column=1, value=title)
    c.font      = _font(theme.font_name, 14, bold=True, color=theme.header_fg)
    c.fill      = _fill(theme.header_bg)
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 28

    # ── Row 2: subtitle ──────────────────────────────────────────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n)
    sub_text = sheet.subtitle or f"{sheet.name} data"
    s = ws.cell(row=2, column=1, value=sub_text)
    s.font      = _font(theme.font_name, 9, italic=True, color=theme.subheader_fg)
    s.fill      = _fill(theme.subheader_bg)
    s.alignment = _align("center")
    ws.row_dimensions[2].height = 16

    # ── Row 3: column headers ────────────────────────────────────────────────
    for ci, col in enumerate(cols, start=1):
        c = ws.cell(row=3, column=ci, value=col.label)
        c.font      = _font(theme.font_name, 10, bold=True, color=theme.col_label_fg)
        c.fill      = _fill(theme.col_label_bg)
        c.alignment = _align("center")
        c.border    = brd
    ws.row_dimensions[3].height = 20

    # ── Rows 4+: data ────────────────────────────────────────────────────────
    data_start = 4
    row_idx    = data_start
    visible    = 0

    for raw_row in sheet.rows:
        if sheet.row_filter and not sheet.row_filter(raw_row):
            continue

        bg = theme.row_even_bg if visible % 2 == 0 else theme.row_odd_bg
        for ci, (col, val) in enumerate(zip(cols, raw_row), start=1):
            c = ws.cell(row=row_idx, column=ci, value=_coerce(val))
            c.font      = _font(theme.font_name, 10,
                                bold=col.bold_col, color=theme.row_fg)
            c.fill      = _fill(bg)
            c.alignment = _align(col.align)
            c.border    = brd
            if col.fmt:
                c.number_format = col.fmt
        row_idx += 1
        visible += 1

    # ── Summary footer ───────────────────────────────────────────────────────
    if sheet.summary:
        label_col, *val_cols = sheet.summary.items()
        # First cell: merged label
        first_key = next(iter(sheet.summary))
        first_val = sheet.summary[first_key]

        label_cell = ws.cell(row=row_idx, column=1, value=first_key)
        label_cell.font   = _font(theme.font_name, 10, bold=True, color=theme.summary_fg)
        label_cell.fill   = _fill(theme.summary_bg)
        label_cell.border = brd
        label_cell.alignment = _align("right")

        # Map remaining keys to columns by position (2nd col → col 2, etc.)
        items = list(sheet.summary.items())
        for ci, (key, val) in enumerate(items, start=1):
            c = ws.cell(row=row_idx, column=ci, value=_coerce(val))
            c.font      = _font(theme.font_name, 10, bold=True, color=theme.summary_fg)
            c.fill      = _fill(theme.summary_bg)
            c.border    = brd
            c.alignment = _align("center")
        ws.row_dimensions[row_idx].height = 18

    # ── Auto-filter on header row ────────────────────────────────────────────
    ws.auto_filter.ref = f"A3:{get_column_letter(n)}3"

    # ── Freeze title + header ────────────────────────────────────────────────
    ws.freeze_panes = ws.cell(row=data_start, column=1)

    # ── Column widths ────────────────────────────────────────────────────────
    for ci, col in enumerate(cols, start=1):
        letter = get_column_letter(ci)
        if col.width:
            ws.column_dimensions[letter].width = col.width
        else:
            max_len = len(col.label)
            for r in ws.iter_rows(min_row=data_start, max_row=row_idx,
                                   min_col=ci, max_col=ci):
                for cell in r:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[letter].width = min(max(max_len + 3, 10), 45)


# ─── Summary sheet ────────────────────────────────────────────────────────────

def _write_summary_sheet(wb: Workbook, sheets: list[Sheet],
                         title: str, meta: dict, theme: Theme) -> None:
    ws = wb.create_sheet("📋 Summary")

    n = 3
    brd = _border(theme.border_color)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    c = ws.cell(row=1, column=1, value=title)
    c.font      = _font(theme.font_name, 16, bold=True, color=theme.header_fg)
    c.fill      = _fill(theme.header_bg)
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n)
    ts = ws.cell(row=2, column=1,
                 value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ts.font      = _font(theme.font_name, 9, italic=True, color=theme.subheader_fg)
    ts.fill      = _fill(theme.subheader_bg)
    ts.alignment = _align("center")
    ws.row_dimensions[2].height = 16

    # Column headers
    for ci, label in enumerate(["Section", "Details", "Rows"], start=1):
        c = ws.cell(row=3, column=ci, value=label)
        c.font      = _font(theme.font_name, 10, bold=True, color=theme.col_label_fg)
        c.fill      = _fill(theme.col_label_bg)
        c.alignment = _align("center")
        c.border    = brd
    ws.row_dimensions[3].height = 20

    # Meta rows
    row = 4
    for i, (k, v) in enumerate(meta.items()):
        bg = theme.row_even_bg if i % 2 == 0 else theme.row_odd_bg
        for ci, val in enumerate([k, str(v), ""], start=1):
            c = ws.cell(row=row, column=ci, value=val)
            c.font      = _font(theme.font_name, 10, color=theme.row_fg)
            c.fill      = _fill(bg)
            c.border    = brd
            c.alignment = _align("left" if ci == 1 else "center")
        row += 1

    # Spacer
    row += 1

    # Sheets index
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n)
    h = ws.cell(row=row, column=1, value="Sheets in this workbook")
    h.font      = _font(theme.font_name, 10, bold=True, color=theme.header_fg)
    h.fill      = _fill(theme.header_bg)
    h.alignment = _align("center")
    row += 1

    for ci, label in enumerate(["Tab Name", "Subtitle", "Row Count"], start=1):
        c = ws.cell(row=row, column=ci, value=label)
        c.font      = _font(theme.font_name, 10, bold=True, color=theme.col_label_fg)
        c.fill      = _fill(theme.col_label_bg)
        c.alignment = _align("center")
        c.border    = brd
    row += 1

    for i, s in enumerate(sheets):
        bg = theme.row_even_bg if i % 2 == 0 else theme.row_odd_bg
        row_count = sum(1 for r in s.rows
                        if not s.row_filter or s.row_filter(r))
        for ci, val in enumerate([s.name, s.subtitle or "—", row_count], start=1):
            c = ws.cell(row=row, column=ci, value=val)
            c.font      = _font(theme.font_name, 10, color=theme.row_fg)
            c.fill      = _fill(bg)
            c.border    = brd
            c.alignment = _align("left" if ci == 1 else "center")
        row += 1

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 12
    ws.freeze_panes = "A4"


# ─── Value coercion ───────────────────────────────────────────────────────────

def _coerce(val: Any) -> Any:
    """Convert Python types to Excel-safe values."""
    if isinstance(val, (date, datetime)):
        return val
    if val is None:
        return ""
    return val


# ─── Main exporter class ──────────────────────────────────────────────────────

class ExcelExporter:
    """
    Universal multi-tab Excel exporter.

    Usage
    ─────
        exp = ExcelExporter(title="Q3 Analytics", theme=Theme.teal())
        exp.add_sheet(Sheet(name="Sales", columns=[...], rows=[...]))
        exp.add_sheet(Sheet(name="Users", columns=[...], rows=[...]))

        # Save to file
        exp.save("report.xlsx")

        # Get raw bytes (for any HTTP framework)
        data = exp.to_bytes()

        # Django
        return exp.to_django_response("report.xlsx")

        # Flask
        return exp.to_flask_response("report.xlsx")

        # Zipped CSV (one file per sheet)
        return exp.to_zip_bytes()
    """

    def __init__(
        self,
        title:    str   = "Export",
        theme:    Theme = None,
        meta:     dict  = None,          # extra key/values shown on Summary tab
        summary_sheet: bool = True,      # add a 📋 Summary tab
    ):
        self.title          = title
        self.theme          = theme or Theme.teal()
        self.meta           = meta or {}
        self.summary_sheet  = summary_sheet
        self._sheets: list[Sheet] = []

    def add_sheet(self, sheet: Sheet) -> "ExcelExporter":
        """Add a sheet. Returns self so you can chain calls."""
        self._sheets.append(sheet)
        return self

    # ── Build workbook ───────────────────────────────────────────────────────

    def _build(self) -> Workbook:
        wb = Workbook()
        wb.remove(wb.active)  # remove default empty sheet

        if self.summary_sheet:
            _write_summary_sheet(wb, self._sheets, self.title, self.meta, self.theme)

        for sheet in self._sheets:
            _write_sheet(wb, sheet, self.theme, self.title)

        return wb

    # ── Output methods ───────────────────────────────────────────────────────

    def to_bytes(self) -> bytes:
        """Return the .xlsx file as raw bytes."""
        buf = io.BytesIO()
        self._build().save(buf)
        buf.seek(0)
        return buf.read()

    def save(self, path: str) -> None:
        """Write the .xlsx to a file path."""
        self._build().save(path)

    def to_zip_bytes(self) -> bytes:
        """Return a ZIP archive containing one CSV per sheet (no styling)."""
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for sheet in self._sheets:
                csv_buf = io.StringIO()
                writer  = csv.writer(csv_buf)
                cols    = _col_objects(sheet.columns)
                writer.writerow([c.label for c in cols])
                for row in sheet.rows:
                    if sheet.row_filter and not sheet.row_filter(row):
                        continue
                    writer.writerow([_coerce(v) for v in row])
                safe_name = "".join(
                    ch for ch in sheet.name if ch.isalnum() or ch in " _-"
                ).strip() or "sheet"
                zf.writestr(f"{safe_name}.csv", csv_buf.getvalue())
        buf.seek(0)
        return buf.read()

    def to_django_response(self, filename: str = "export.xlsx"):
        """Return an HttpResponse for Django views."""
        try:
            from django.http import HttpResponse
        except ImportError:
            raise RuntimeError("Django is not installed.")

        data = self.to_bytes()
        resp = HttpResponse(
            data,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp


    def to_zip_django_response(self, filename: str = "export.zip"):
        """Return a zipped CSV archive as a Django HttpResponse."""
        try:
            from django.http import HttpResponse
        except ImportError:
            raise RuntimeError("Django is not installed.")

        resp = HttpResponse(self.to_zip_bytes(), content_type="application/zip")
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp