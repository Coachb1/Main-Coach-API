from apis.accounts.serializers import BotAttributeSerializer, CoachCoacheeMentorMenteeProfileSerializer, LLMMappingSerializer, SignatureBotSerializer, clientUserInfoSerializer
from commons.anthropic import anthropic_completion
from commons.langchain import download_and_transcribe_audio
from commons.youtube_utils import get_youtube_transcript
from documents.utils import get_document_summary
from skills.models import SkillsRating
from users.db import get_user_by_id, get_user_display_name
from users.helpers import get_client_info_from_user_detail, sync_user_low_high_skills
from users.models import SignatureBot, BotAttribute, ClientUserInfo, CoachCoacheeRating,CoachCoacheeMentorMenteeProfile, User, UserAttribute, CoachCoacheeConnection, get_default_signature_bot_page_information
from utilities.helpers import generate_email, get_llm_order
from utilities.models import BotQnA, DirectoryPageInfo, GlobalSystemInstructions, LLMMappingTable, SessionNotesRecommendations, UserActionInfo
from tests.models import TestAttemptSession, Test
from coaching_conversations.helpers import add_or_remove_emails_from_client, get_client_user_info
from identities.models import Identity
from django.db import transaction
from commons.notifications import send_error_notification, send_generic_email
from email_sender.helpers import send_email_with_html_template
import logging
import traceback
import datetime
from django.db.models import Q
from users.choices import BotTypeChoice, ProfileTypeChoice, StatusChoice
from commons.cache_utils import generate_cache_key, get_cache, reset_cache_with_prefix, set_cache
import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
import os
from collections.abc import MutableMapping
import json

from utilities.prompts import get_intake_summary_prompt

logger = logging.getLogger(__name__)


def delete_user_resources(user_uid, remove_from_client=False,
                          delete_profile=True,
                          delete_bot =True,
                          delete_user_connections=True,
                          delete_session =True,
                          delete_session_notes=True,
                          delete_user_action = True,
                          soft_delete=False,
                          bot_types = [BotTypeChoice.avatar_bot, BotTypeChoice.subject_specific_bot, BotTypeChoice.user_bot, BotTypeChoice.feedback_bot],
                          bot_ids = None
                          ):
    logger.info("Deleting user resources for user %s", user_uid)
    deleted_list = []
    deleted_msg = ""
    try:
        user = User.objects.get(uid=user_uid)
        tenant_id = user.tenant_id
        connection_uids = []
        profiles = CoachCoacheeMentorMenteeProfile.objects.filter(tenant_id=tenant_id,user_id=user_uid)
        profile_ids = []
        for profile in profiles:
            profile_ids.append(profile.uid)
            

            if delete_user_connections or delete_profile:
                logger.info(f'====================deleting user connections=========================')
                # delete connections if user has coachee profile
                connections = CoachCoacheeConnection.objects.filter(tenant_id=tenant_id,coachee_id=profile.uid)
                for connection in connections:
                    deleted_list.append(f"connection coachee: {connection.uid}")
                    connection_uids.append(connection.uid)
                    connection.delete()
                
                # delete connections if user has coach profile
                connections = CoachCoacheeConnection.objects.filter(tenant_id=tenant_id,coach_id=profile.uid)
                for connection in connections:
                    deleted_list.append(f"connection coach : {connection.uid}")
                    connection_uids.append(connection.uid)
                    connection.delete()

                logger.info(f'====================deleted user connections=========================')
            

            if delete_profile:
                logger.info(f'====================deleting user profile=========================')

                # delete directorypage for this profile
                dir_to_be_deleted = [profile.uid]
                if 'user_bot' in bot_types:
                    dir_to_be_deleted.append(user_uid)

                dir_infos = DirectoryPageInfo.objects.filter(profile_id__in=dir_to_be_deleted)
                for dir_info in dir_infos:
                    deleted_list.append(f"directorypage : {dir_info.name}")
                    dir_info.delete()
                    
                deleted_list.append(f"profile : {profile.uid}")
                if soft_delete:
                    profile.deleted = True
                    profile.save(update_fields=['deleted'])
                else:
                    profile.delete()
        

        if delete_profile:
            reset_cache_with_prefix('profiles_by_user_id')
            reset_cache_with_prefix('profile_by_id')
            reset_cache_with_prefix('all_profiles')

        # delete bots if user has any
        if delete_bot:
            logger.info(f'====================deleting user bot=========================')
            bots = None
            if bot_ids:
                bots = SignatureBot.objects.filter(tenant_id=tenant_id,uid__in =bot_ids)
            else:
                bots = SignatureBot.objects.filter(tenant_id=tenant_id,user_id=user_uid, bot_type__in=bot_types)
            for bot in bots:
                # delete bot related resources
                bot_attributes = BotAttribute.objects.filter(bot_id=bot.bot_id)
                for bot_attribute in bot_attributes:
                    deleted_list.append(f"bot_attribute : {bot_attribute.uid}")
                    if soft_delete:
                        bot_attribute.deleted = True
                        bot_attribute.save(update_fields=['deleted'])
                    else:
                        bot_attribute.delete()
                    
                bot_qnas = BotQnA.objects.filter(bot_id=bot.bot_id)
                for bot_qna in bot_qnas:
                    deleted_list.append(f"bot_qna_{bot_qna.bot_id} : {bot_qna.uid}")
                    if soft_delete:
                        bot_qna.deleted = True
                        bot_qna.save(update_fields=['deleted'])
                    else:
                        bot_qna.delete()
                    
                deleted_list.append(f"bot : {bot.bot_id}")
                if soft_delete:
                    bot.deleted = True
                    bot.save(update_fields=['deleted'])
                else:
                    bot.delete()
            
            reset_cache_with_prefix('get_bots')

        if user:
            with transaction.atomic():
                if delete_user_action:
                    logger.info(f'====================deleting user action=========================')
                    UserActionInfo.objects.filter(tenant_id=tenant_id, user_id=user.uid).delete()
                if delete_session:
                    logger.info(f'====================deleting user sessions=========================')
                    TestAttemptSession.objects.filter(tenant_id=tenant_id, participant_id=user.uid).update(deleted = True)
                    SkillsRating.objects.filter(tenant_id=tenant_id, participant_id=user.uid).update(deleted = True)
                if delete_session_notes:
                    logger.info(f'====================deleting user session notes=========================')
                    SessionNotesRecommendations.objects.filter(
                        tenant_id=tenant_id
                    ).filter(Q(mentee_id=user.uid) | Q(mentor_id=user.uid)).delete()
        
        if remove_from_client:
            logger.info(f'====================removing from client=========================')

            try:
                identity = Identity.objects.get(user_id=user_uid)
                user_email = identity.value
                clients = ClientUserInfo.objects.filter(tenant_id=tenant_id, member_emails__contains=user_email)
                for client in clients:
                    add_or_remove_emails_from_client(client,'member_emails',user_email,True)
                    add_or_remove_emails_from_client(client,'demo_ids',user_email,True)
                    deleted_list.append(f'removed from client : {client.client_name}')
            except Exception as e:
                logger.exception(f"failed to delete client for the user {user_uid}: {e}")


        # checking if all data is deleted:
        profile_ids.append(user_uid)
        deleted_msg += f"""
        
        profile: {CoachCoacheeMentorMenteeProfile.objects.filter(tenant_id=tenant_id,user_id=user_uid).count()}
        connections coach: {CoachCoacheeConnection.objects.filter(tenant_id=tenant_id,coach_id__in=profile_ids).count()}
        connections coachee: {CoachCoacheeConnection.objects.filter(tenant_id=tenant_id,coachee_id__in=profile_ids).count()}
        directorypage: {DirectoryPageInfo.objects.filter(profile_id__in=profile_ids).count()}
        bot: {SignatureBot.objects.filter(tenant_id=tenant_id,user_id=user_uid).count()}
        user_action_info: {UserActionInfo.objects.filter(tenant_id=tenant_id, user_id=user.uid).count()}
        test_attempt_session: {TestAttemptSession.objects.filter(tenant_id=tenant_id, deleted=False, participant_id=user.uid).count()}
        session_notes_recommendations_mentee: {SessionNotesRecommendations.objects.filter(tenant_id=tenant_id, mentee_id=user.uid).count()}
        session_notes_recommendations_mentor: {SessionNotesRecommendations.objects.filter(tenant_id=tenant_id, mentor_id=user.uid).count()}
        """
        if remove_from_client:
            deleted_msg += f"clients: {ClientUserInfo.objects.filter(tenant_id=tenant_id, member_emails__contains=user_email).count()}"


        logger.info(f"delete_list : {deleted_list}, deleted_msg: {deleted_msg}")
        content = f"User: {user.name}- ({user.uid}) \n at => " + datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "\n" + "*"*20 + "User resources deleted." + "*"*20 + "<br/><br/>"
        content += "deleted check: " + deleted_msg + "\n" 
        send_email_with_html_template("User Resource Deletion",content)
    except Exception as e:
        logger.error({"Error":e}, exc_info=True)
        error = f"Got Error: {e}"
        error += "\n traceback: " + traceback.format_exc()
        send_error_notification("delete user resources after client change" if not remove_from_client else "delete user resources from collab", error,{'user_uid':user_uid,'deleted_list': deleted_list,"deleted_confirmation": deleted_msg})
        raise e


def flatten_dict(d, parent_key='', sep='_'):
    items = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, MutableMapping):
            items.extend(flatten_dict(v, new_key, sep=sep).items())
        else:
            items.append((new_key, v))
    return dict(items)


def generate_user_details_excel(data):
    # Client ID, Date created, user Name, Email ID, Profile Type, Connections, Intake form data, Report links.
    # Define the specific fields to keep first in order
    specific_fields = ["client_id", "intake_date", "intake_data_name", "user_email", "user_type", "connections_count", "details_user_email", "details_user_type"]

    # Flatten each dictionary in the list
    flattened_data = [flatten_dict(item) for item in data]

    # Collect all unique keys from the flattened dictionaries
    all_fieldnames = set()
    for item in flattened_data:
        all_fieldnames.update(item.keys())

    # Ensure the specific fields are first, followed by the rest of the fields
    fieldnames = specific_fields + [field for field in all_fieldnames if field not in specific_fields]

    # Create a workbook and a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "User Details"

    # Write the header row
    header_font = Font(bold=True)
    for col_num, fieldname in enumerate(fieldnames, 1):
        cell = ws.cell(row=1, column=col_num, value=fieldname)
        cell.font = header_font
        ws.column_dimensions[cell.column_letter].width = 15  # Set column width

    # Write the data rows
    for row_num, item in enumerate(flattened_data, start=2):
        for col_num, fieldname in enumerate(fieldnames, start=1):
            ws.cell(row=row_num, column=col_num, value=item.get(fieldname))

    # Save the workbook to a file
    filename = "formatted_output.xlsx"
    wb.save(filename)

    with open('formatted_output.xlsx', 'rb') as fh:
        file_response = HttpResponse(
            fh.read(), content_type="text/csv", status=200)
        file_response['Content-Disposition'] = 'inline; filename=' + os.path.basename('formatted_output.xlsx')

    return file_response


def get_bot_details_helper(signature_bot: SignatureBot):
    data = {}
    data['faqs'] = signature_bot.faqs
    data['attributes'] = signature_bot.attributes
    data['bot_details'] = signature_bot.bot_details
    data['recommended_codes'] = signature_bot.recommended_codes
    data['bot_type'] = signature_bot.bot_type
    data['user_id'] = signature_bot.user_id
    data['is_fitment_analysis'] = signature_bot.is_fitment_analysis
    data['is_strict_fitment'] = signature_bot.is_strict_fitment
    data['is_sample_bot'] = signature_bot.is_sample_bot
    data['is_system_bot'] = signature_bot.is_system_bot
    data['additional_data'] = signature_bot.data.get('additional_data',None)
    data['scenario_case'] = signature_bot.bot_scenario_case
    data['bot_expires_at'] = signature_bot.bot_expires_at
    data['access_code'] = signature_bot.access_code
    data['tag'] = signature_bot.tag
    data['page_information'] = signature_bot.page_informations or get_default_signature_bot_page_information()
    data['is_private'] = signature_bot.is_private
    data['allow_public_access'] = signature_bot.allow_public_access
    
    if signature_bot.system_instructions:
        data['system_instructions'] = signature_bot.system_instructions
    else:
        system_instruction = GlobalSystemInstructions.objects.filter(deleted=False,tenant_id=signature_bot.tenant_id, resourse_type=signature_bot.bot_type).first()
        logger.info(f"system_instruction: {system_instruction.instruction if system_instruction else None}")
        data['system_instructions'] = system_instruction.instruction if system_instruction else None

    
    
    client = get_client_info_from_user_detail(tenant_id=signature_bot.tenant_id, user_uid=signature_bot.user_id)

    llms = LLMMappingTable.objects.filter(deleted=False, bot_type=signature_bot.bot_type, tenant_id=signature_bot.tenant_id).first()
    if llms:
        data['selected_llms'] = LLMMappingSerializer(llms).data
    data['llm_order'] = get_llm_order(bot_type=signature_bot.bot_type, tenant_id=signature_bot.tenant_id)
    

    if client:
        data["allowed_ips"] = client.allowed_ips

    if signature_bot.bot_type == 'deep_dive':
        data['deep_dive_data'] = signature_bot.data
        data['deepdive_prompt'] = signature_bot.custom_prompt

    try:
        bot_att = BotAttribute.objects.get(bot_id=signature_bot.uid)
        data['is_audio_response'] = bot_att.is_audio_response
        data['ui_information'] = bot_att.ui_information
        data['extracted_data'] = bot_att.extracted_documents

        if bot_att.fitment_data:
            data['fitment_qna'] = bot_att.fitment_data['mentee_que']
        if bot_att.fitment_data:
            data['fitment_options'] = bot_att.fitment_data['options']
        
        if bot_att.feedback_questions:
            data['feedback_qna'] = bot_att.feedback_questions
        if bot_att.initial_qnas:
            data['initial_qna'] = bot_att.initial_qnas
        if bot_att.bot_name:
            data['bot_name'] = bot_att.bot_name
        if bot_att.about:
            data['description'] = bot_att.about

        coach_profile = CoachCoacheeMentorMenteeProfile.objects.filter(deleted=False,user_id=signature_bot.user_id).first()

        if coach_profile:
            data["coaching_for_fitment"] = coach_profile.coaching_for_fitment.lower() if coach_profile.coaching_for_fitment else None
            data['profile_details'] = CoachCoacheeMentorMenteeProfileSerializer(coach_profile).data

        
        feedback_bot = SignatureBot.objects.filter(tenant_id=signature_bot.tenant_id,user_id=signature_bot.user_id,bot_type=BotTypeChoice.feedback_bot).first()
        if feedback_bot:
            data['feedback_id'] = feedback_bot.bot_id
        else:
            data['feedback_id'] = None

        if not signature_bot.is_system_bot and not signature_bot.is_sample_bot:
            if coach_profile:
                data['owner_profile_image'] = coach_profile.profile_image_url
        
    except Exception as e:
        logger.exception(e)
        error_msg = f"Failed to get bot details: {e}\n\n"
        error_msg += traceback.format_exc()
        # send_slack_message({"module": "########### get_bot_details ###########", "error": str(e)})
        send_error_notification("get_bot_details",error_msg,{"bot_id":signature_bot.uid})


    return data


def resolve_client_information(
    *,
    mode,
    client_info,
    user_id=None,
    email=None,
    mob_number=None,
    request=None,
):
    """
    Resolve client information payload based on requested mode.
    """

    data = {}

    if mode == "my_lib":
        client_and_emails_map = []

        for client in client_info:
            client_and_emails_map.append({
                "group": client.client_name,
                "emails": [e for e in client.member_emails.split(",")] if client.member_emails else []
            })

        data["my_lib"] = client_and_emails_map

    elif mode == "user_info":
        user = ""

        if user_id:
            user = client_info.filter(member_user_ids__contains=user_id)
        if email:
            user = client_info.filter(member_emails__contains=email)
        if mob_number:
            user = client_info.filter(member_mob_numbers__contains=mob_number)

        user_info = []

        for u in user:
            client_user_data = get_client_user_info(u, email)
            user_info.append(client_user_data)

        if len(user_info) == 0:
            user_info.append({
                "msg": "user not found",
                "is_restricted": False,
                "is_demo_user": True
            })

        data["user_info"] = user_info

    elif mode == "only_client_data":
        client = None
        client_name = request.query_params.get("client_name")
        client_id = request.query_params.get("client_id")

        if client_name:
            client = client_info.filter(client_name=client_name)
        elif client_id:
            client = client_info.filter(uid=client_id)
        else:
            if user_id:
                client = client_info.filter(member_user_ids__contains=user_id)
            if email:
                client = client_info.filter(member_emails__contains=email)
            if mob_number:
                client = client_info.filter(member_mob_numbers__contains=mob_number)

        data["only_client_data"] = (
            clientUserInfoSerializer(client.first()).data if client else {}
        )

    return data




def get_profiles_by_user_id(user_id):
    profile = CoachCoacheeMentorMenteeProfile.objects.filter(deleted=False, user_id=user_id)
    return CoachCoacheeMentorMenteeProfileSerializer(profile, many=True).data

def get_profile_by_id(profile_id):
    profile = CoachCoacheeMentorMenteeProfile.objects.get(deleted=False, uid=profile_id)
    return CoachCoacheeMentorMenteeProfileSerializer(profile).data

def get_all_profiles(profile_type=None):
    profiles = CoachCoacheeMentorMenteeProfile.objects.filter(deleted=False, is_approved=True)
    if profile_type:
        profiles = profiles.filter(profile_type=profile_type)
    return CoachCoacheeMentorMenteeProfileSerializer(profiles, many=True).data


def get_coach_profiles_helper(user_id=None, profile_id=None, profile_type=None):
    if user_id:
        try:
            cache_key = generate_cache_key('profiles_by_user_id', user_id=user_id)
            data = get_cache(cache_key)
            if data is None:
                data = get_profiles_by_user_id(user_id)
                set_cache(cache_key, data)
            return data
        except Exception as e:
            logger.exception(e)
            raise e
    if profile_id:
        try:
            cache_key = generate_cache_key('profile_by_id', profile_id=profile_id)
            data = get_cache(cache_key)
            if data is None:
                data = get_profile_by_id(profile_id)
                set_cache(cache_key, data)
            return data
        except Exception as e:
            logger.exception(e)
            raise e
    
    else:
        cache_key = generate_cache_key('all_profiles', profile_type=profile_type)
        data = get_cache(cache_key)
        if data is None:
            data = get_all_profiles(profile_type)
            set_cache(cache_key, data)
        return data



def update_coach_profile_helper(tenant_id, profile_id, request_data, for_reapproval=False):
    print("*"*100)
    logger.info(f"data : {request_data}")
    data = {"tenant_id" : tenant_id}
    data.update(request_data)
    profile = CoachCoacheeMentorMenteeProfile.objects.get(deleted=False,uid=profile_id)
    print(data, profile.name)
    serializer = CoachCoacheeMentorMenteeProfileSerializer(profile,data=data,partial=True)
    serializer.is_valid(raise_exception=True)

    # saving bot_data if any before profile sync
    if data.get('bot_data'):

        bot_data = json.loads(data.get('bot_data')) if isinstance(data.get('bot_data'),str) else data.get('bot_data')
        bot_description = bot_data.get('bot_description')
        bot_name = bot_data.get('bot_name')

        bot_id = bot_data.get('bot_id')
        bot = SignatureBot.objects.filter(deleted=False,uid=bot_id).last()
        if bot:
            bot_att = BotAttribute.objects.get(bot_id=bot.uid)
            bot_att.bot_name = bot_name
            bot_att.about = bot_description
            bot_att.save(update_fields=['bot_name','about'])

            add_data = bot.data['additional_data']
            additional_data = {
                "bot_area_of_coaching": bot_data.get('bot_area_of_coaching'),
                "bot_description": bot_data.get('bot_description')
            }
            if add_data:
                for key, value in additional_data.items():
                    add_data[key] = value
                bot.data['additional_data'] = add_data
                bot.save(update_fields=['data'])

                
    serializer.save()

    # sending for reapproval to directory page info
    logger.info(f"sending for reapproval:  {for_reapproval}")

    directory = DirectoryPageInfo.objects.filter(profile_id=profile_id).first()
    if directory and for_reapproval:

        avata_bot_id = directory.avatar_bot_id
        bot = SignatureBot.objects.filter(deleted=False,tenant_id=tenant_id,bot_id=avata_bot_id).first()
        if bot:
            bot.is_approved = False
            bot.save()

        # to send reapproval msg
        # profile.is_approved_email_sent = False
        # profile.save()


        DirectoryPageInfo.objects.create(
            name = directory.name,
            profile_id = directory.profile_id,
            department = directory.department,
            bot_type = directory.bot_type,
            profile_pic_url = directory.profile_pic_url,
            profile_type = directory.profile_type,
            description = directory.description,
            experience = directory.experience,
            expertise = directory.expertise,
            status = directory.status,
            avatar_bot_id = directory.avatar_bot_id,
            feedback_wall = directory.feedback_wall,
            skills = directory.skills,
            is_visible = directory.is_visible,
            is_approved = False,
            avatar_snippit = directory.avatar_snippit,
            avatar_bot_url = directory.avatar_bot_url,
            custom_user_bot_url = directory.custom_user_bot_url,
            custom_user_bot_id = directory.custom_user_bot_id,
            subject_specific_bot_url = directory.subject_specific_bot_url,
            subject_specific_bot_id = directory.subject_specific_bot_id,
            subject_specific_bot_snippit= directory.subject_specific_bot_snippit,
            timer_enabled = directory.timer_enabled,
            time_value_in_days = directory.time_value_in_days,
            timer_reset = directory.timer_reset,
            visual_tag = directory.visual_tag,
            ai_email = directory.ai_email,
            integratable_snippet = directory.integratable_snippet,
        )


        # directory.save()
        try:
            subject = "AI Copilot Updation"
            html = f"""
                <p style="font-family: sans-serif; font-size: 14px; font-weight: normal; margin: 0; margin-bottom: 15px;">Thank you for updating your AI Copilot/Profile. It is under processing pipeline and you will soon receive a confirmation when it's live. You can always edit the same via the profile section.</p>
                """

            send_email_with_html_template(subject=subject,html_content=html,to_email=profile.email,title=f'Hey {profile.name}!')
            html = f"""
                <p style="font-family: sans-serif; font-size: 14px; font-weight: normal; margin: 0; margin-bottom: 15px;">{profile.name} - {profile.email} Updated a bot/profile. Please check it out and re-approve it from Django Admin Panel.</p>
                """
            send_email_with_html_template(subject=subject,html_content=html)

        except Exception as e:
            logger.error(f"Got error in sending email for reapproval : {e}")
            send_error_notification("coach_coachee_mentor_mentee_profile",f"Got error in sending email for reapproval : {e}",{"data":data})
            
        directory.delete()
    reset_cache_with_prefix('profiles_by_user_id')
    reset_cache_with_prefix('profile_by_id')
    reset_cache_with_prefix('all_profiles')
    return serializer.data


def create_coach_profile_helper(data, profile_approved):
    
    serializer = CoachCoacheeMentorMenteeProfileSerializer(data=data)
    serializer.is_valid(raise_exception=True)
    logger.info(f"serializer data: {serializer.validated_data}")
    created_profile = serializer.save()

    tenant_id = data.get('tenant_id')
    
    low_skill = serializer.validated_data.get("low_rating_characteristics")
    high_skill = serializer.validated_data.get("high_rating_characteristics")
    
    # if None in [low_skill, high_skill]:
    #     return Response({"error": "low_rating_characteristics and high_rating_characteristics is required"},status=status.HTTP_400_BAD_REQUEST)
    
    sync_user_low_high_skills(tenant_id, data['user_id'], low_skill, high_skill)
    
    if (created_profile.profile_type) in ('coachee','mentee'):
        created_profile.is_approved = True
        created_profile.save(update_fields=["is_approved"])
        
    send_generic_email(f"{created_profile.name} just created {created_profile.profile_type}  Account",
                    f"{created_profile.name} just created {created_profile.profile_type}  Account. check it out on admin panel(https://coach-api-ovh.coachbots.com/custom-admin/) and approve it, to make it display on Directory page")
    # send_generic_email(f"{created_profile.name} just created {created_profile.profile_type}  Account",
    #                    f"{created_profile.name} just created {created_profile.profile_type}  Account. check it out on admin panel(https://coach-api-ovh.coachbots.com/custom-admin/) and approve it, to make it display on Directory page",
    #                    'aadil611ofc@gmail.com')
    profile_type = created_profile.profile_type
    if created_profile.is_mentor:
        profile_type = ProfileTypeChoice.coach_mentor

    DirectoryPageInfo.objects.create(
            name=created_profile.name,
            profile_id=created_profile.uid,
            department=created_profile.department,
            profile_pic_url=created_profile.profile_image_url or "None",
            profile_type=profile_type,
            description=created_profile.about,
            experience=created_profile.experience,
            expertise=created_profile.area_domain,
            status=StatusChoice.available,
            skills=created_profile.high_rating_characteristics,
            is_visible= True,
            is_approved =  True if (created_profile.profile_type) in ('coachee','mentee') else profile_approved,
            ai_email = generate_email(created_profile.name,created_profile.id)
            )
    
    reset_cache_with_prefix('profiles_by_user_id')
    reset_cache_with_prefix('profile_by_id')
    reset_cache_with_prefix('all_profiles')
                    
    return  serializer.data


def get_bots_helper(
    *,
    tenant_id,
    user_id=None,
    bot_type=None,
    client=None,
    approved_only=False,
):
    bots_qs = SignatureBot.objects.filter(
        deleted=False,
        tenant_id=tenant_id,
    )

    if approved_only:
        bots_qs = bots_qs.filter(is_approved=True)

    if user_id:
        bots_qs = bots_qs.filter(user_id=user_id)

    if bot_type:
        bots_qs = bots_qs.filter(bot_type=bot_type)

    deepdive_bot_access = None

    # ---------- CLIENT FILTER (optimized) ----------
    if client:
        
            bot_user_ids = list(bots_qs.values_list("user_id", flat=True))

            user_attrs = UserAttribute.objects.filter(
                deleted=False,
                tenant_id=tenant_id,
                user_id__in=bot_user_ids,
            ).values("user_id", "attributes")

            email_by_user = {
                ua["user_id"]: ua["attributes"].get("email")
                for ua in user_attrs
                if ua["attributes"]
            }

            

            allowed_emails = set()
            if client.member_emails:
                allowed_emails.update(
                    e.strip() for e in client.member_emails.split(",")
                )

            if bot_type == BotTypeChoice.deep_dive and client.deepdive_accessed_emails:
                deepdive_bot_access = [
                    e.strip()
                    for e in client.deepdive_accessed_emails.split(",")
                ]

            allowed_user_ids = [
                uid for uid, email in email_by_user.items() if email in allowed_emails
            ]

            bots_qs = bots_qs.filter(user_id__in=allowed_user_ids)

    # ---------- BATCH LOAD ATTRIBUTES ----------
    bot_attrs = {
        ba.bot_id: ba
        for ba in BotAttribute.objects.filter(
            bot_id__in=bots_qs.values_list("uid", flat=True)
        )
    }

    data = []
    for bot in bots_qs:
        bot_serializer = SignatureBotSerializer(bot)
        bot_att = bot_attrs.get(bot.uid)

        bot_data = {
            "creator_name": bot_serializer.data.get("creator_name"),
            "signature_bot": bot_serializer.data,
            "bot_attributes": BotAttributeSerializer(bot_att).data if bot_att else None,
        }

        if deepdive_bot_access:
            bot_data["deepdive_access"] = deepdive_bot_access

        data.append(bot_data)

    return data


#************* utility methods ***************
def process_and_store_youtube_transcript(youtube_links,signature_bot,overwrite=False, deleted_data = {}):
        """
        Processes YouTube links to extract transcripts and stores them.

        For each link, it attempts to fetch the transcript. If successful,
        the transcript is stored in the `media_data` of the `SignatureBot`
        and `extracted_documents` of the `BotAttribute`.

        Args:
            youtube_links (list or str): A list of YouTube video URLs.
            signature_bot (SignatureBot): The bot instance to associate the transcripts with.
            overwrite (bool): If True, existing YouTube transcripts will be replaced.
            deleted_data (dict): A dictionary specifying any data to be deleted.
        
        Returns:
            str: The transcript of the last processed YouTube link, or None.
        """
        extracted_from_youtube = {}
        extracted_media_data = {}
        
        if not isinstance(youtube_links, list):
            youtube_links = [youtube_links]

        logger.info(f"*************** youtube_links in process_and_store : {youtube_links}")

        transcript = None
        for link in youtube_links:
            if link != '':
                logger.info(f"Gettinggs transcript for youtube link: {link}")
                try:
                    for i in range(2):
                        transcript = get_youtube_transcript(link)
                        if transcript is not None:
                            logger.info(f"transcript: {transcript}")
                            break
                    if transcript is None:
                        logger.info(f"Could not get transcript for youtube link: {link} from package so trying to download and transcribe")
                        transcript = download_and_transcribe_audio(link)
                        logger.info(f"Transcript after download and transcribe : {transcript}")
                    if signature_bot.bot_type in [BotTypeChoice.avatar_bot, BotTypeChoice.subject_specific_bot]:
                        extracted_media_data[link] = transcript
                        transcript = get_document_summary(transcript)
                    extracted_from_youtube[link] = transcript
                except Exception as e:
                    logger.exception(e)
                    extracted_from_youtube[link] = {"error": "Restricted video. error in extracting transcript. Please try another."}
            
        # extracted_media_data['extracted_from_youtube'] = extracted_from_youtube
        logger.info(f"extratedz youtube: {extracted_from_youtube}")
        signature_bot.refresh_from_db()
        bot_media_data = signature_bot.data['media_data']
        if overwrite and extracted_from_youtube:
            bot_media_data['extracted_from_youtube'] = extracted_from_youtube
        else:
            prev_extracted_from_youtube = bot_media_data.get('extracted_from_youtube',{})
            # if "youtube_links" in deleted_data:
            #     for link in deleted_data["youtube_links"].strip().split(","):
            #         prev_extracted_from_youtube.pop(link.strip(),None)
            bot_media_data['extracted_from_youtube'] = {**prev_extracted_from_youtube,**extracted_from_youtube}

        signature_bot.data['media_data'] = bot_media_data
        signature_bot.save(update_fields=["data"])

        bot_att = BotAttribute.objects.get(bot_id=signature_bot.uid)
        bot_att.refresh_from_db()
        bot_media_data = bot_att.extracted_documents if bot_att.extracted_documents else {}
        bot_media_data['extracted_from_youtube'] = {**bot_media_data.get('extracted_from_youtube',{}),**extracted_media_data}
        bot_att.extracted_documents = bot_media_data
        bot_att.save(update_fields=["extracted_documents"])
        
        return transcript


def get_or_create_feedback_helper(
        method:str,
        qna: str,
        bot_id:str,
        feedback_type:str,
        participant_id:str,
        qna_type:str,
        tenant_id:str,
        is_positive:bool,
        is_anonymous:bool,
        signature_bot: SignatureBot
        ):
    
    data = {}
    cache_key = ""

    if method.lower() == 'get':
        cache_key = generate_cache_key('get-user-feedback-data', bot_id=bot_id, feedback_type=feedback_type, participant_id=participant_id, qna_type=qna_type, tenant_id=tenant_id)
        
        # Try to get data from cache
        cached_data = get_cache(cache_key)
        if cached_data:
            return cached_data
        
        # to get latest botqna for a user using participant_id
        if participant_id and signature_bot:
            recent_intake_data = BotQnA.objects.filter(tenant_id = tenant_id,bot_id=signature_bot.uid,participant_id=participant_id,qna_type=qna_type).order_by('-created').first()
            if recent_intake_data:
                set_cache(cache_key, data)
                return {"intake_summary": recent_intake_data.intake_summary,"intake_id":recent_intake_data.uid}
            else:
                return {"error": "No Intake found for user."}
        

        # to get feedback bot's feedback msg using bot_id
        feedback_data = BotQnA.objects.filter(tenant_id = tenant_id,bot_id=signature_bot.uid,qna_type='feedback')
        msg_data = []
        for feed in feedback_data:
            try:
                participant_name = get_user_display_name(
                    get_user_by_id(feed.participant_id))

                bot = SignatureBot.objects.filter(deleted=False, uid=feed.bot_id).first()
                coach_name = "Unknown"
                if bot:
                    coach_name = get_user_display_name(
                        get_user_by_id(feed.participant_id))
                else:
                    logger.info(f"Bot not found: {feed.bot_id}")
                    continue

                
            except Exception as e:
                logger.exception(f'{e}')
                logger.info(f"User not found: {feed.participant_id}")
                continue
            
            if feedback_type == "negative":
                if not feed.is_positive:
                    msg_data.append({
                        "participant_name": participant_name,
                        "date": feed.created,
                        "msg": feed.participant_qna,
                        "participant_id": feed.participant_id,
                        "is_anonymous": feed.is_anonymous,
                        "coach_name": coach_name,
                        "bot_uid": feed.bot_id,
                        "bot_id": bot.bot_id

                    })
            elif feedback_type == 'positive':
                if feed.is_positive:
                    msg_data.append({
                        "participant_name": participant_name,
                        "date": feed.created,
                        "msg": feed.participant_qna,
                        "participant_id": feed.participant_id,
                        "is_anonymous": feed.is_anonymous,
                        "coach_name": coach_name,
                        "bot_uid": feed.bot_id,
                        "bot_id": bot.bot_id

                    })
            else:
                msg_data.append({
                    "participant_name": participant_name,
                    "date": feed.created,
                    "msg": feed.participant_qna,
                    "participant_id": feed.participant_id,
                    "is_anonymous": feed.is_anonymous,
                    "coach_name": coach_name,
                    "bot_uid": feed.bot_id,
                    "bot_id": bot.bot_id
                })
        if feedback_type == "negative":
            data['critical_msgs'] = msg_data
        elif feedback_type == 'positive':
            data['positive_msgs'] = msg_data
        else:
            data['message'] = msg_data

        set_cache(cache_key, data)
        

    elif method.lower() == 'post':
        
        logger.info(f"qna : {qna}, ispositive: {is_positive} , is_anonymous: {is_anonymous}")

        intake_summary_prompt = get_intake_summary_prompt(qna)
        intake_summary = anthropic_completion(intake_summary_prompt,50000)

        BotQnA.objects.create(
            tenant_id = tenant_id,
            participant_id = participant_id,
            participant_qna = json.loads(qna),
            is_positive = is_positive,
            bot_id = signature_bot.uid if signature_bot else None,
            qna_type = qna_type,
            intake_summary = intake_summary,
            is_anonymous = is_anonymous
        )
        data['message'] = "created"

        reset_cache_with_prefix('get-user-feedback-data')

    return data


def normalize_bool(value, default=False):
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return value == 1
    if isinstance(value, str):
        return value.strip().lower() in ("true", "1", "yes")
    return default


def validate_user_feedback_input(request):
    """
    Validate and normalize input for user feedback API.

    GET supports exactly two modes:
        1) Intake Summary Mode   -> user_id present
        2) Feedback Listing Mode -> user_id absent

    POST supports creating a new QnA entry.
    """

    source = (
        request.query_params.dict()
        if request.method == "GET"
        else dict(request.data)
    )

    method = request.method.lower()

    bot_id = source.get("bot_id")
    qna_type = source.get("qna_type")
    participant_id = source.get("user_id")
    feedback_type = source.get("feedback_type")
    qna = source.get("qna")

    is_positive = normalize_bool(source.get("is_positive"))
    is_anonymous = normalize_bool(source.get("is_anonymous"))

    # -------------------------
    # COMMON VALIDATION
    # -------------------------
    

    # if qna_type not in ("feedback", "fitment", "initial_qna"):
    #     raise ValueError(
    #         "Invalid qna_type. Allowed values: feedback, fitment, initial_qna"
    #     )

    if qna_type != "fitment" and not bot_id:
        raise ValueError("bot_id is required for this qna_type")

    # -------------------------
    # GET VALIDATION
    # -------------------------
    if method == "get":


        # CASE 1: Intake summary
        if participant_id :
            # Valid: user-specific intake fetch
            pass


        # CASE 2: Feedback listing
        else:
            if qna_type != "feedback":
                raise ValueError(
                    "Feedback listing is only supported for qna_type=feedback"
                )

            if feedback_type and feedback_type not in ("positive", "negative"):
                raise ValueError(
                    "feedback_type must be 'positive' or 'negative'"
                )

    # -------------------------
    # POST VALIDATION
    # -------------------------
    elif method == "post":
        if not qna_type:
            raise ValueError("qna_type is required")

        if not participant_id:
            raise ValueError("user_id is required for POST")

        if not qna:
            raise ValueError("qna is required for POST")

        if qna_type == "feedback" and source.get("is_positive") is None:
            raise ValueError(
                "is_positive is required when submitting feedback"
            )

    else:
        raise ValueError("Unsupported HTTP method")

    return {
        "method": method,
        "bot_id": bot_id,
        "qna_type": qna_type,
        "user_id": participant_id,
        "feedback_type": feedback_type,
        "qna": qna,
        "is_positive": is_positive,
        "is_anonymous": is_anonymous,
    }
