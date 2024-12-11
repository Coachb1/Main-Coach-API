import datetime
from rest_framework import mixins
from rest_framework import status
from rest_framework.decorators import action
from rest_framework.response import Response
import logging
from django.db.models import Subquery
from django.utils import timezone
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
import threading
import csv
from io import TextIOWrapper
from external_apis.slack_alert_api import send_slack_message
from commons.notifications import send_error_notification


from apis.accounts.aggregator import create_user_account
from apis.accounts.dtos import UserCreateContextDto, IdentityCreateContextDto
from apis.accounts.serializers import AccountSerializer, UserAttributesUserContextSerializer, CoachCoacheeConnectionSerializer
from apis.accounts.serializers import (SetupAccountSerializer, CoachCoacheeMentorMenteeProfileSerializer,
                                        SignatureBotSerializer, BotAttributeSerializer,DirectoryInfoSErializer,
                                        CoachCoacheeJoiningPreviledgeSerializer, CoachCoacheeRatingSerializer, LLMMappingSerializer)
from clients.permissions import IsAuthenticatedClient
from tests.models import TestAttemptSession, Test
from users.permissions import IsAuthenticatedUser
from commons.viewset import ApiViewSet
from identities.helpers import get_user_via_identity
from pdf_generator.helpers import get_participant_report
from users.helpers import upsert_user_attributes, get_client_info_from_user_detail, update_user_account, sync_user_low_high_skills
from users.models import CoachCoacheeMentorMenteeProfile, User, UserAttribute, CoachCoacheeConnection
from users.choices import BotTypeChoice, UserRoleChoice
from tenants.models import Tenant
from tests.choices import TestAttemptSessionStatusChoices
from users.models import SignatureBot, BotAttribute, ClientUserInfo, CoachCoacheeRating
from users.choices import StatusChoice, ProfileTypeChoice, CoachCoacheeConnectionStatusChoice
from tests.helpers import scrape_article_data, get_unique_deep_dive_access_code


from identities.models import Identity
from skills.models import SkillsRating
from utilities.models import BotQnA, DirectoryPageInfo
from users.db import get_user_by_id,get_user_display_name, get_user_attribute
import json
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser

from email_sender.helpers import send_generic_email, send_email_with_html_template
from utilities.helpers import extract_fields
from commons.langchain import download_and_transcribe_audio, extract_text_from_pdf, extract_text_from_doc
from coaching_conversations.helpers import signature_bot_default_prompt, get_client_user_data, update_member_client_id, create_or_assign_client_id, disable_or_enable_client, get_client_user_info
from utilities.helpers import process_idp, regenerate_idp_or_scenarios, generate_email
from utilities.models import UserActionInfo, CoachCoacheeJoiningPreviledge, LLMMappingTable, GlobalSystemInstructions
from commons.utils import extract_file_and_text, get_list_from_string
                    
from itertools import groupby
from operator import attrgetter
from django.db.models import Q
from commons.youtube_utils import get_youtube_transcript
from utilities.prompts import get_intake_summary_prompt
from commons.anthropic import anthropic_completion
from utilities.helpers import custom_sort_reverse
from coaching_conversations.choices import BotScenarioCaseChoice
from coaching_conversations.helpers import generate_title_and_objective_for_deep_dive
import traceback
from documents.utils import get_document_summary
import random
from coaching_conversations.helpers import update_or_create_client_id
from apis.accounts.serializers import clientUserInfoSerializer
from apis.accounts.utils import delete_user_resources
from utilities.models import SessionNotesRecommendations
from django.db import transaction
from coaching_conversations.helpers import add_or_remove_emails_from_client
from users.models import get_default_signature_bot_page_information

from django.utils.decorators import method_decorator
from django.views.decorators.cache import cache_page
from django.views.decorators.vary import vary_on_cookie, vary_on_headers
from django.core.cache import cache
import time
from commons.cache_utils import get_cache, set_cache, delete_cache, generate_cache_key, reset_cache_with_prefix

import openpyxl
import os
from openpyxl.styles import Font
from collections.abc import MutableMapping
from django.http import HttpResponse
from users.helpers import generate_bot_id
import copy

logger = logging.getLogger(__name__)

class AccountsViewSet(ApiViewSet,
                      mixins.ListModelMixin):
    queryset = User.objects.filter(deleted=0)
    serializer_class = AccountSerializer
    permission_classes = (IsAuthenticatedClient, IsAuthenticatedUser)
    lookup_field = "uid"
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def get_queryset(self):
        return super().get_queryset().filter(tenant_id=self.request.tenant.uid)

    def create(self, request, *args, **kwargs):
        """
        Create a user account.

        Args:
            request (object): The HTTP request object containing the user and identity data.

        Returns:
            object: The HTTP response object containing the serialized user account data.
        """
        serializer = SetupAccountSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        user_context = serializer.validated_data["user_context"]
        identity_context = serializer.validated_data["identity_context"]


        try:
            i_context=IdentityCreateContextDto(**identity_context)

            identity = Identity.objects.get(
                tenant_id=request.tenant.uid,
                identity_type=i_context.identity_type,
                value=i_context.value,
                deleted=0
                )

            user = User.objects.get(
                tenant_id=request.tenant.uid,
                uid=identity.user_id,
                deleted=0
            )
            logger.info("got user")
        except Exception as e:
            logger.info("creating user")
            user = create_user_account(tenant=request.tenant,
                                user_context=UserCreateContextDto(
                                    **user_context),
                                identity_context=IdentityCreateContextDto(**identity_context))

        return Response(AccountSerializer(instance=user).data, status=status.HTTP_201_CREATED)

    @action(methods=["GET"],
            detail=False,
            url_path=r"identities/(?P<identity_type>[^\s]+)/(?P<identity_value>[^\s]+)")
    def get_account_via_identity(self, request, identity_type, identity_value, *args, **kwargs):
        """
        Retrieves a user account based on the provided identity type and identity value.

        :param request: The HTTP request object.
        :param identity_type: The type of identity to search for.
        :param identity_value: The value of the identity to search for.
        :return: The HTTP response object containing the serialized user account data.
        """
        user = get_user_via_identity(
            tenant=request.tenant,
            identity_type=identity_type,
            identity_value=identity_value
        )
        return Response(AccountSerializer(instance=user).data, status=status.HTTP_200_OK)

    @action(methods=["POST"], detail=True, url_path="upsert-attributes")
    def upsert_user_attributes_view(self, request, *args, **kwargs):
        """
        Update or insert user attributes in the database.

        :param request: The HTTP request object.
        :type request: HttpRequest
        :param args: Additional positional arguments.
        :type args: tuple
        :param kwargs: Additional keyword arguments.
        :type kwargs: dict
        :return: The HTTP response object containing the serialized user account data.
        :rtype: HttpResponse
        """
        serializer = UserAttributesUserContextSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        tag = serializer.validated_data["tag"]
        attributes = serializer.validated_data["attributes"]

        user = self.get_object()

        user_attribute = upsert_user_attributes(user=user,
                                                tag=tag,
                                                attributes=attributes)

        return Response(AccountSerializer(instance=user).data, status=status.HTTP_200_OK)

    @action(methods=["GET"], detail=True, url_path="participant-report")
    def get_participant_report_pdf_view(self, request, *args, **kwargs):
        user = self.get_object()

        report_url = get_participant_report(user)

        return Response({"report_url": report_url})

    @action(methods=["GET"], detail=True, url_path="participant-report-data")
    def get_participant_report_frontend(self, request, *args, **kwargs):
        user = self.get_object()

        data = get_participant_report(user, only_data=True)

        return Response({"data": data, "status": "completed"})


    @action(methods=["GET"], detail=False, url_path="get-workspace-users")
    def get_workspace_users(self, request, *args, **kwargs):
        """
        Retrieves a list of workspace users who have attempted tests.

        Args:
            request (object): The HTTP request object.

        Returns:
            dict: A dictionary containing the names and IDs of workspace users who have attempted tests.

        """
        try:
            included_users = self.get_queryset().filter(is_excluded=0).values('uid')
            users = UserAttribute.objects.filter(user_id__in=Subquery(included_users))

            user_data = {}
            for user in users:
                try:
                    skills_rating = SkillsRating.objects.get(participant_id=user.user_id)
                    if skills_rating.total_tests_attempted > 0:
                        user_data[f"{user.attributes['real_name']} - {user.attributes['name']}"] = user.attributes['id']
                except:
                    pass

            return Response(user_data, status=status.HTTP_200_OK)
        except Exception as e:
            logger.error({"!!!! Error":e},exc_info=True)


    @action(methods=['GET'], detail=False, url_path="get_is_repeat_status")
    def get_is_repeat_status(self, request, *args, **kwargs):
        """
        Retrieves the repeat status of a participant for a given tenant.

        This method calculates the number of tests the participant has attempted in the current month and compares it to the maximum number of tests allowed per month.
        It returns the repeat status and the remaining number of tests for the month.

        :param request: The HTTP request object.
        :param participant_id: The ID of the participant for whom the repeat status is being checked.
        :return: A dictionary containing the tenant ID, repeat status, and the remaining number of tests for the month.
        """

        tenant = self.request.tenant
        participant_id = request.query_params.get("participant_id")

        try:
            test_per_month = tenant.test_per_month
            current_month = timezone.now().month
            # date_month_ago = timezone.make_aware(date_month_ago, timezone.get_current_timezone())
            sessions = TestAttemptSession.objects.filter(participant_id = participant_id,tenant_id=tenant.uid, status=TestAttemptSessionStatusChoices.completed)
            this_month_sessions = []
            for session in sessions:
                if session.created.month == current_month:
                    this_month_sessions.append(session)
            total_test_attempted = len(this_month_sessions)
        except Exception as e:
            logger.error({"!!!!!!!!!! Error": e}, exc_info=True)
            total_test_attempted = 0

        data = {"tenant_id": tenant.uid, "is_repeat": tenant.is_repeat, "monthly_remaining_tests": test_per_month - total_test_attempted}
        return Response(data, status=status.HTTP_200_OK)

    @action(methods=['GET'], detail=False, url_path="get-user-type")
    def get_user_type(self,request,*args, **kwargs):
        """
            Retrieves the role of a user based on their user ID if user Id given else it retrives a list of all roles in our system.
        """
        user_id = request.query_params.get('user_id')

        if user_id:
            user = User.objects.get(uid=user_id)

            user_role = user.role

            return Response({"user_role":user_role}, status=status.HTTP_200_OK)
        else:
            # get all userRole in system
            user_roles = list(UserRoleChoice.values.keys())
            return Response({"user_roles":user_roles}, status=status.HTTP_200_OK)

    @action(methods=['GET'], detail=False, url_path="get-mobile-number-restriction-list-whatsapp")
    def get_mobile_number_res_list_whatsapp(self,request,*args, **kwargs):
        """
        Only for Whatsapp use.
        Retrive restricted mobile numbers from db
        """
        tenant = self.request.tenant
        number_list = []

        if tenant.mobile_number_restriction_whatsapp:
            number_list = tenant.mobile_number_list.split(',')
            number_list = [number.strip() for number in number_list]

        return Response({"mobile_numbers": number_list,'active': tenant.mobile_number_restriction_whatsapp},status=status.HTTP_200_OK)


    @action(methods=['GET'], detail=False, url_path="get-test-codes-for-web")
    def get_test_codes_for_web(self,request,*args, **kwargs):
        '''
            Retrives all testcode json for web environment(deepchat)
        '''

        tenant = self.request.tenant
        logger.info({'tenant': tenant,"test_code_json":tenant.web_test_code_json})


        test_code_json = tenant.web_test_code_json

        return Response({"data":test_code_json},status=status.HTTP_200_OK)
    
    @action(methods=['GET','POST'],detail=False, url_path="get-my-lib-data")
    def get_my_lib_data(self,request,*args, **kwargs):
        """
        Retrieves library data based on a group name.

        Args:
            request (object): The HTTP request object.
            group (string): The name of the group to filter tests.

        Returns:
            Dictionary: A dictionary containing the library data grouped by domain. Each domain key maps to a list of test items, where each item contains the title, description, domain, test code, and interaction mode.
        """
        # test_codes = request.query_params.get('test_codes').split(',')
        group_name = request.query_params.get('group',None)
        candidate_type = request.query_params.get('candidate_type',None)
        data = []
        if group_name:
            tests = Test.objects.filter(deleted=0,client_name=group_name)
            for item in tests:
                title_parts = item.title.split(':')
                key = title_parts[0].strip().capitalize()
            
                data.append({"title": item.title,"description": item.description, "domain": key,
                                "test_code": item.test_code, "interaction_mode": item.interaction_mode, "is_micro": item.is_micro, "is_recommended": item.is_recommended, "scenario_case": item.scenario_case})

        if candidate_type:
            tests = Test.objects.filter(deleted=0,tenant_id=self.request.tenant.uid,candidate_type=candidate_type.strip().lower())
            for test in tests:
                if test.area_domain:
                    key = test.area_domain
                    data.append({"title": item.title,"description": item.description, "domain": key,
                                    "test_code": item.test_code, "interaction_mode": item.interaction_mode, "is_micro": item.is_micro, "is_recommended": item.is_recommended, "scenario_case": item.scenario_case})

        
        group_data = {}
        for item in data:
            domain = item['domain']
            if domain in group_data:
                group_data[domain].append(item)
            else:
                group_data[domain] = [item]

        return Response({"data":group_data},status=status.HTTP_200_OK)


    @action(methods=['GET'],detail=False, url_path="get-bot-details")
    def get_bot_details(self,request,*args, **kwargs):
        """
        Retrieves details of a bot based on the provided bot ID.

        :param request: The HTTP request object.
        :param bot_id: The ID of the bot to retrieve details for.
        :return: A dictionary containing the bot details with keys 'faqs', 'attributes', 'bot_details', and 'recommended_codes'.
        :rtype: dict
        """
        bot_id = request.query_params.get('bot_id')

        logger.info(f"****************** Bot ID: {bot_id} **********************")

        try:
            signature_bot = SignatureBot.objects.get(deleted=False,tenant_id=self.request.tenant.uid,bot_id=bot_id)
        except Exception as e:
            logger.exception({"!!!!!!!!!! Error": e}, exc_info=True)
            return Response({"error": "Bot not found"},status=status.HTTP_404_NOT_FOUND)

        data = {}
        data['faqs'] = signature_bot.faqs
        data['attributes'] = signature_bot.attributes
        data['bot_details'] = signature_bot.bot_details
        data['recommended_codes'] = signature_bot.recommended_codes
        data['bot_type'] = signature_bot.bot_type
        data['user_id'] = signature_bot.user_id
        data['is_fitment_analysis'] = signature_bot.is_fitment_analysis
        data['is_strict_fitment'] = signature_bot.is_strict_fitment
        data['is_sample_bot'] = signature_bot.is_sample_bot
        data['is_system_bot'] = signature_bot.is_system_bot
        data['additional_data'] = signature_bot.data.get('additional_data',None)
        data['scenario_case'] = signature_bot.bot_scenario_case
        data['bot_expires_at'] = signature_bot.bot_expires_at
        data['access_code'] = signature_bot.access_code
        data['tag'] = signature_bot.tag
        data['page_information'] = signature_bot.page_informations or get_default_signature_bot_page_information()
        data['is_private'] = signature_bot.is_private
        data['allow_public_access'] = signature_bot.allow_public_access
        
        if signature_bot.system_instructions:
            data['system_instructions'] = signature_bot.system_instructions
        else:
            system_instruction = GlobalSystemInstructions.objects.filter(deleted=False,tenant_id=signature_bot.tenant_id, resourse_type=signature_bot.bot_type).first()
            logger.info(f"system_instruction: {system_instruction.instruction if system_instruction else None}")
            data['system_instructions'] = system_instruction.instruction if system_instruction else None

        
        
        client = get_client_info_from_user_detail(tenant_id=signature_bot.tenant_id, user_uid=signature_bot.user_id)

        llms = LLMMappingTable.objects.filter(deleted=False, bot_type=signature_bot.bot_type, tenant_id=signature_bot.tenant_id).first()
        if llms:
            data['selected_llms'] = LLMMappingSerializer(llms).data

        if client:
            data["allowed_ips"] = client.allowed_ips

        if signature_bot.bot_type == 'deep_dive':
            data['deep_dive_data'] = signature_bot.data
            data['deepdive_prompt'] = signature_bot.custom_prompt

        try:
            bot_att = BotAttribute.objects.get(bot_id=signature_bot.uid)
            data['is_audio_response'] = bot_att.is_audio_response
            data['ui_information'] = bot_att.ui_information
            data['extracted_data'] = bot_att.extracted_documents

            if bot_att.fitment_data:
                data['fitment_qna'] = bot_att.fitment_data['mentee_que']
            if bot_att.fitment_data:
                data['fitment_options'] = bot_att.fitment_data['options']
            
            if bot_att.feedback_questions:
                data['feedback_qna'] = bot_att.feedback_questions
            if bot_att.initial_qnas:
                data['initial_qna'] = bot_att.initial_qnas
            if bot_att.bot_name:
                data['bot_name'] = bot_att.bot_name
            if bot_att.about:
                data['description'] = bot_att.about

            coach_profile = CoachCoacheeMentorMenteeProfile.objects.filter(deleted=False,user_id=signature_bot.user_id).first()

            if coach_profile:
                data["coaching_for_fitment"] = coach_profile.coaching_for_fitment.lower() if coach_profile.coaching_for_fitment else None
                data['profile_details'] = CoachCoacheeMentorMenteeProfileSerializer(coach_profile).data

            
            feedback_bot = SignatureBot.objects.filter(tenant_id=self.request.tenant.uid,user_id=signature_bot.user_id,bot_type=BotTypeChoice.feedback_bot).first()
            if feedback_bot:
                data['feedback_id'] = feedback_bot.bot_id
            else:
                data['feedback_id'] = None

            if not signature_bot.is_system_bot and not signature_bot.is_sample_bot:
                if coach_profile:
                    data['owner_profile_image'] = coach_profile.profile_image_url
            
        except Exception as e:
            logger.exception(e)
            error_msg = f"Failed to get bot details: {e}\n\n"
            error_msg += traceback.format_exc()
            # send_slack_message({"module": "########### get_bot_details ###########", "error": str(e)})
            send_error_notification("get_bot_details",error_msg,{"bot_id":bot_id})


        return Response({"data":data},status=status.HTTP_200_OK)


    @action(methods=['GET'], detail=False, url_path="get-client-information")
    def get_client_informations(self,request,*args, **kwargs):
        """
        Retrieves client information based on the provided parameters.
        Returns:
            dict: A dictionary containing the retrieved client information. The structure of the dictionary depends on the `mode` parameter.
        """
        try:
            logger.info(f"(((((((((((((((((((((((((((((((( REQUEST ORIGIN : {request.META.get('HTTP_REFERER')} |   {request.headers.get('Origin')}  |  {request.META.get('HTTP_ORIGIN')}   ))))))))))))))))))))))))))))))))")
            mode = request.query_params.get('for',None)  # can be my_lib, user_info
            user_id = request.query_params.get('user_id',None)
            email = request.query_params.get('email',None)
            mob_number = request.query_params.get('mob_number',None)
            tenant = self.request.tenant

            logger.info(f"Received request with parameters - mode: {mode}, user_id: {user_id}, email: {email}, mob_number: {mob_number}")
            

            cache_key = generate_cache_key("client_info",tenant.uid,mode,user_id,email,mob_number)
            cached_data = get_cache(cache_key)
            # if cached_data:
            #     return Response({"data":cached_data},status=status.HTTP_200_OK)
            
            client_info = ClientUserInfo.objects.filter(tenant_id=tenant.uid, deleted=False)
            data = {}

            if mode == 'my_lib':
                client_and_emails_map = []

                for client in client_info:
                    client_and_emails_map.append({"group": client.client_name,
                                                "emails": [email for email in client.member_emails.split(',')] if client.member_emails else []
                                                })
                
                data['my_lib'] = client_and_emails_map
            
            elif mode == 'user_info':
                user = ''
                if user_id:
                    user = client_info.filter(member_user_ids__contains = user_id)
                if email:
                    user = client_info.filter(member_emails__contains = email)
                if mob_number:
                    user = client_info.filter(member_mob_numbers__contains = mob_number)

                user_info = []

                for u in user:
                    client_user_data = get_client_user_info(u,email)
                    user_info.append(client_user_data)

                if len(user_info) == 0:
                    user_info.append({"msg": "user not found",
                                      "is_restricted": False,
                                      "is_demo_user": True},
                                      )

                data['user_info'] = user_info

            elif mode == 'only_client_data':
                client = None
                client_name = self.request.query_params.get('client_name')
                if client_name:
                    client = client_info.filter(client_name=client_name)
                else:
                    if user_id:
                        client = client_info.filter(member_user_ids__contains = user_id)
                    if email:
                        client = client_info.filter(member_emails__contains = email)
                    if mob_number:
                        client = client_info.filter(member_mob_numbers__contains = mob_number)

                data['only_client_data'] = clientUserInfoSerializer(client.first()).data if client else {}
                
            set_cache(cache_key, data)
            logger.info("Client information retrieval successful")

            return Response({"data":data },status=status.HTTP_200_OK)

        except Exception as e:
            logger.exception(f"got error: {e}")
            error_msg = f"failed to get client information: {e}\n\n"
            error_msg += traceback.format_exc()
            # send_slack_message({"module": "########### get_client_informations ###########", "error": str(e)})
            # send_error_notification("get_client_informations",error_msg,{"mode":mode,"user_id":user_id,"email":email,"mob_number":mob_number})
            return Response({"error":e},status=status.HTTP_400_BAD_REQUEST)
        

    @action(methods=['GET','POST'], detail=False, url_path="get-user-feedback-data")
    def get_user_feedback_data(self,request,*args, **kwargs):
        """
        Retrieves or creates user feedback data for a specific bot.

        Args:
            request (object): The HTTP request object.
            method (string): The method to perform, either "get" or "post".
            feedback_type(string):  nagetive then fetches only critical msg.
            bot_id (string): The ID of the bot for which to retrieve or create feedback data.
            user_id (string): The ID of the user for whom to create feedback data (only required for "post" method).
            qna (string): The question and answer data for the feedback (only required for "post" method).
            is_positive (boolean): Indicates whether the feedback is positive or not (only required for "post" method).
            qna_type (string): The type of the feedback (only required for "post" method).

        Returns:
            If the method is "get":
                A dictionary containing the positive messages data.
            If the method is "post":
                A dictionary with a "message" key indicating the success of the creation.
        """
        try:
            if request.method == 'GET':
                method = request.query_params.get('method',None)
                bot_id = request.query_params.get('bot_id',None)
                feedback_type = request.query_params.get("feedback_type",None)
                participant_id = request.query_params.get('user_id',None)
                qna_type = request.query_params.get('qna_type',None)
                qna = request.query_params.get('qna',None)
                is_positive = request.query_params.get('is_positive',"False")
                is_anonymous = request.query_params.get('is_anonymous',"False")
                
            elif request.method == 'POST':
                method = request.data.get('method',None)
                bot_id = request.data.get('bot_id',None)
                feedback_type = request.data.get("feedback_type",None)
                participant_id = request.data.get('user_id',None)
                qna_type = request.data.get('qna_type',None)
                qna = request.data.get('qna',None)
                is_positive = request.data.get('is_positive',"False")
                is_anonymous = request.data.get('is_anonymous',"False")

            cache_key = ''

            data = {}
            signature_bot = None
            if qna_type != 'fitment':
                try:
                    signature_bot = SignatureBot.objects.get(deleted=False,tenant_id = self.request.tenant.uid,bot_id=bot_id)
                except Exception as e:
                    logger.exception(e)
                    return Response({"error":"bot not found"},status=status.HTTP_400_BAD_REQUEST)
            if method.lower() == 'get':
                # try:
                #     feedback_bot_id = SignatureBot.objects.get(tenant_id = self.request.tenant.uid,user_id=signature_bot.user_id,bot_type='feedback_bot').uid
                #     print(feedback_bot_id)
                # except Exception as e:
                #     logger.exception(f"Feedback bot not found {e}")
                #     data['positive_msgs'] = []
                #     return Response(data,status=status.HTTP_200_OK)
                cache_key = generate_cache_key('get-user-feedback-data', bot_id=bot_id, feedback_type=feedback_type, participant_id=participant_id, qna_type=qna_type, tenant_id=request.tenant.uid)
                
                # Try to get data from cache
                cached_data = get_cache(cache_key)
                if cached_data:
                    return Response(cached_data, status=status.HTTP_200_OK)
                
                # to get latest botqna for a user using participant_id
                if participant_id:
                    recent_intake_data = BotQnA.objects.filter(tenant_id = self.request.tenant.uid,bot_id=signature_bot.uid,participant_id=participant_id,qna_type=qna_type).order_by('-created').first()
                    if recent_intake_data:
                        set_cache(cache_key, data)
                        return Response({"intake_summary": recent_intake_data.intake_summary,"intake_id":recent_intake_data.uid},status=status.HTTP_200_OK)
                    else:
                        return Response({"error": "No Intake found for user."},status=status.HTTP_400_BAD_REQUEST)
                

                # to get feedback bot's feedback msg using bot_id
                feedback_data = BotQnA.objects.filter(tenant_id = self.request.tenant.uid,bot_id=signature_bot.uid,qna_type='feedback')
                msg_data = []
                for feed in feedback_data:
                    try:
                        participant_name = get_user_display_name(
                            get_user_by_id(feed.participant_id))

                        bot = SignatureBot.objects.filter(delted=False, uid=feed.bot_id).first()
                        coach_name = "Unknown"
                        if bot:
                            coach_name = get_user_display_name(
                                get_user_by_id(feed.bot_id))
                        else:
                            logger.info(f"Bot not found: {feed.bot_id}")
                            continue

                        
                    except Exception as e:
                        logger.info(f"User not found: {feed.participant_id}")
                        continue
                    
                    if feedback_type == "negative":
                        if not feed.is_positive:
                            msg_data.append({
                                "participant_name": participant_name,
                                "date": feed.created,
                                "msg": feed.participant_qna,
                                "participant_id": feed.participant_id,
                                "is_anonymous": feed.is_anonymous,
                                "coach_name": coach_name,
                                "bot_uid": feed.bot_id,
                                "bot_id": bot.bot_id

                            })
                    elif feedback_type == 'positive':
                        if feed.is_positive:
                            msg_data.append({
                                "participant_name": participant_name,
                                "date": feed.created,
                                "msg": feed.participant_qna,
                                "participant_id": feed.participant_id,
                                "is_anonymous": feed.is_anonymous,
                                "coach_name": coach_name,
                                "bot_uid": feed.bot_id,
                                "bot_id": bot.bot_id

                            })
                    else:
                        msg_data.append({
                            "participant_name": participant_name,
                            "date": feed.created,
                            "msg": feed.participant_qna,
                            "participant_id": feed.participant_id,
                            "is_anonymous": feed.is_anonymous,
                            "coach_name": coach_name,
                            "bot_uid": feed.bot_id,
                            "bot_id": bot.bot_id
                        })
                if feedback_type == "negative":
                    data['critical_msgs'] = msg_data
                elif feedback_type == 'positive':
                    data['positive_msgs'] = msg_data
                else:
                    data['message'] = msg_data

            elif method.lower() == 'post':
                
                logger.info(f"qna : {qna}, ispositive: {is_positive} , is_anonymous: {is_anonymous}")

                intake_summary_prompt = get_intake_summary_prompt(qna)
                intake_summary = anthropic_completion(intake_summary_prompt,50000)

                BotQnA.objects.create(
                    tenant_id = self.request.tenant.uid,
                    participant_id = participant_id,
                    participant_qna = json.loads(qna),
                    is_positive = True if is_positive.lower() == 'true' else False,
                    bot_id = signature_bot.uid if qna_type != 'fitment' else None,
                    qna_type = qna_type,
                    intake_summary = intake_summary,
                    is_anonymous = True if is_anonymous.lower() == 'true' else False
                )
                data['message'] = "created"

            set_cache(cache_key, data)
            return Response(data,status=status.HTTP_200_OK)
        
        except Exception as e:
            logger.exception(f"got error: {e}")
            send_slack_message({"module": "########### get_user_feedback_data ###########", "error": str(e)})
            return Response({"error":e},status=status.HTTP_400_BAD_REQUEST)
        

    @action(methods=['GET','POST','PATCH'], detail=False, url_path="coach-coachee-mentor-mentee-profile")
    def coach_coachee_mentor_mentee_profile(self,request,*args, **kwargs):
        """
        Retrieves or creates a coach-coachee-mentor-mentee profile for a user.

        Args:
            request (object): The HTTP request object.

        Returns:
            dict: A dictionary containing the coach-coachee-mentor-mentee profile data.
        """
        def get_profiles_by_user_id(user_id):
            profile = CoachCoacheeMentorMenteeProfile.objects.filter(deleted=False, user_id=user_id)
            return CoachCoacheeMentorMenteeProfileSerializer(profile, many=True).data

        def get_profile_by_id(profile_id):
            profile = CoachCoacheeMentorMenteeProfile.objects.get(deleted=False, uid=profile_id)
            return CoachCoacheeMentorMenteeProfileSerializer(profile).data

        def get_all_profiles(profile_type=None):
            profiles = CoachCoacheeMentorMenteeProfile.objects.filter(deleted=False, is_approved=True)
            if profile_type:
                profiles = profiles.filter(profile_type=profile_type)
            return CoachCoacheeMentorMenteeProfileSerializer(profiles, many=True).data


        # return Response({"data":data},status=status.HTTP_200_OK)
        if request.method == 'GET':
            profile_id = request.query_params.get('profile_id',None)
            user_id = request.query_params.get('user_id',None)
            profile_type = request.query_params.get('profile_type',None)
            
            if user_id:
                try:
                    cache_key = generate_cache_key('profiles_by_user_id', user_id=user_id)
                    data = get_cache(cache_key)
                    if data is None:
                        data = get_profiles_by_user_id(user_id)
                        set_cache(cache_key, data)
                    return Response({"data": data}, status=status.HTTP_200_OK)
                except Exception as e:
                    logger.exception(e)
                    return Response({"error":"profile not found"},status=status.HTTP_404_NOT_FOUND)
            if profile_id:
                try:
                    cache_key = generate_cache_key('profile_by_id', profile_id=profile_id)
                    data = get_cache(cache_key)
                    if data is None:
                        data = get_profile_by_id(profile_id)
                        set_cache(cache_key, data)
                    return Response({"data": data}, status=status.HTTP_200_OK)
                except Exception as e:
                    logger.exception(e)
                    return Response({"error":"profile not found"},status=status.HTTP_404_NOT_FOUND)
            else:
                cache_key = generate_cache_key('all_profiles', profile_type=profile_type)
                data = get_cache(cache_key)
                if data is None:
                    data = get_all_profiles(profile_type)
                    set_cache(cache_key, data)
                return Response({"data": data}, status=status.HTTP_200_OK)

        if request.method == 'PATCH':
            try:
                profile_id = request.query_params.get('profile_id',None)
                print("*"*100)
                logger.info(f"data : {request.data}")
                data = {"tenant_id" : self.request.tenant.uid}
                data.update(request.data)
                print(data)
                profile = CoachCoacheeMentorMenteeProfile.objects.get(deleted=False,uid=profile_id)
                serializer = CoachCoacheeMentorMenteeProfileSerializer(profile,data=data,partial=True)
                serializer.is_valid(raise_exception=True)

                # saving bot_data if any before profile sync
                if data.get('bot_data'):

                    bot_data = json.loads(data.get('bot_data')) if isinstance(data.get('bot_data'),str) else data.get('bot_data')
                    bot_description = bot_data.get('bot_description')
                    bot_name = bot_data.get('bot_name')

                    bot_id = bot_data.get('bot_id')
                    bot = SignatureBot.objects.filter(deleted=False,uid=bot_id).last()
                    if bot:
                        bot_att = BotAttribute.objects.get(bot_id=bot.uid)
                        bot_att.bot_name = bot_name
                        bot_att.about = bot_description
                        bot_att.save(update_fields=['bot_name','about'])

                        add_data = bot.data['additional_data']
                        additional_data = {
                            "bot_area_of_coaching": bot_data.get('bot_area_of_coaching'),
                            "bot_description": bot_data.get('bot_description')
                        }
                        if add_data:
                            for key, value in additional_data.items():
                                add_data[key] = value
                            bot.data['additional_data'] = add_data
                            bot.save(update_fields=['data'])

                            
                serializer.save()

                for_reapproval = request.query_params.get('for_reapproval',None).lower().strip() == 'true' if request.query_params.get('for_reapproval',None) else False
                

                # sending for reapproval to directory page info
                logger.info(f"sending for reapproval: {request.query_params.get('for_reapproval',None)}, {for_reapproval}")

                directory = DirectoryPageInfo.objects.filter(profile_id=profile_id).first()
                if directory and for_reapproval:

                    avata_bot_id = directory.avatar_bot_id
                    bot = SignatureBot.objects.filter(deleted=False,tenant_id=self.request.tenant.uid,bot_id=avata_bot_id).first()
                    if bot:
                        bot.is_approved = False
                        bot.save()

                    # to send reapproval msg
                    # profile.is_approved_email_sent = False
                    # profile.save()


                    DirectoryPageInfo.objects.create(
                        name = directory.name,
                        profile_id = directory.profile_id,
                        department = directory.department,
                        bot_type = directory.bot_type,
                        profile_pic_url = directory.profile_pic_url,
                        profile_type = directory.profile_type,
                        description = directory.description,
                        experience = directory.experience,
                        expertise = directory.expertise,
                        status = directory.status,
                        avatar_bot_id = directory.avatar_bot_id,
                        feedback_wall = directory.feedback_wall,
                        skills = directory.skills,
                        is_visible = directory.is_visible,
                        is_approved = False,
                        avatar_snippit = directory.avatar_snippit,
                        avatar_bot_url = directory.avatar_bot_url,
                        custom_user_bot_url = directory.custom_user_bot_url,
                        custom_user_bot_id = directory.custom_user_bot_id,
                        subject_specific_bot_url = directory.subject_specific_bot_url,
                        subject_specific_bot_id = directory.subject_specific_bot_id,
                        subject_specific_bot_snippit= directory.subject_specific_bot_snippit,
                        timer_enabled = directory.timer_enabled,
                        time_value_in_days = directory.time_value_in_days,
                        timer_reset = directory.timer_reset,
                        visual_tag = directory.visual_tag,
                        ai_email = directory.ai_email,
                        integratable_snippet = directory.integratable_snippet,
                    )


                    # directory.save()
                    try:
                        subject = "AI Frame Updation"
                        html = f"""
                            <p style="font-family: sans-serif; font-size: 14px; font-weight: normal; margin: 0; margin-bottom: 15px;">Thank you for updating your AI frame/Profile. It is under processing pipeline and you will soon receive a confirmation when it's live. You can always edit the same via the profile section.</p>
                            """

                        send_email_with_html_template(subject=subject,html_content=html,to_email=profile.email,title=f'Hey {profile.name}!')
                        html = f"""
                            <p style="font-family: sans-serif; font-size: 14px; font-weight: normal; margin: 0; margin-bottom: 15px;">{profile.name} - {profile.email} Updated a bot/profile. Please check it out and re-approve it from Django Admin Panel.</p>
                            """
                        send_email_with_html_template(subject=subject,html_content=html)

                    except Exception as e:
                        logger.error(f"Got error in sending email for reapproval : {e}")
                        send_error_notification("coach_coachee_mentor_mentee_profile",f"Got error in sending email for reapproval : {e}",{"data":data})
                        
                    directory.delete()
                    reset_cache_with_prefix('profiles_by_user_id')
                    reset_cache_with_prefix('profile_by_id')
                    reset_cache_with_prefix('all_profiles')

                return Response({"data": CoachCoacheeMentorMenteeProfileSerializer(profile).data },status=status.HTTP_200_OK)
            except Exception as e:
                logger.exception(e)
                send_slack_message({"module": "########### coach_coachee_mentor_mentee_profile ###########", "error": str(e)})
                return Response({"error":"got error"},status=status.HTTP_400_BAD_REQUEST)

        if request.method == 'POST':
            logger.info(f"************************************** request files : {request}**************************************************************************** request data: {request.data}")
            try:
                data = request.data.copy()
                data['tenant_id'] = self.request.tenant.uid
                profile_approved = data.get('is_approved',False)
                

                if profile_approved:
                    data['is_approved'] = profile_approved
                else:
                    data['is_approved'] = False

                profile = CoachCoacheeMentorMenteeProfile.objects.filter(deleted=False,tenant_id=self.request.tenant.uid,user_id=data['user_id'])
                if profile.count() > 0:
                    return Response({"msg": "Entry already Exist","data": CoachCoacheeMentorMenteeProfileSerializer(profile,many=True).data },status=status.HTTP_200_OK)
                
                # _data = data.copy()
                
                serializer = CoachCoacheeMentorMenteeProfileSerializer(data=data)
                serializer.is_valid(raise_exception=True)
                logger.info(f"serializer data: {serializer.validated_data}")
                created_profile = serializer.save()
                
                low_skill = serializer.validated_data.get("low_rating_characteristics")
                high_skill = serializer.validated_data.get("high_rating_characteristics")
                
                # if None in [low_skill, high_skill]:
                #     return Response({"error": "low_rating_characteristics and high_rating_characteristics is required"},status=status.HTTP_400_BAD_REQUEST)
                
                sync_user_low_high_skills(self.request.tenant.uid, data['user_id'], low_skill, high_skill)
                
                if (created_profile.profile_type) in ('coachee','mentee'):
                    created_profile.is_approved = True
                    created_profile.save(update_fields=["is_approved"])
                    
                send_generic_email(f"{created_profile.name} just created {created_profile.profile_type}  Account",
                                f"{created_profile.name} just created {created_profile.profile_type}  Account. check it out on admin panel(https://coach-api-ovh.coachbots.com/custom-admin/) and approve it, to make it display on Directory page")
                # send_generic_email(f"{created_profile.name} just created {created_profile.profile_type}  Account",
                #                    f"{created_profile.name} just created {created_profile.profile_type}  Account. check it out on admin panel(https://coach-api-ovh.coachbots.com/custom-admin/) and approve it, to make it display on Directory page",
                #                    'aadil611ofc@gmail.com')
                profile_type = created_profile.profile_type
                if created_profile.is_mentor:
                    profile_type = ProfileTypeChoice.coach_mentor

                DirectoryPageInfo.objects.create(
                        name=created_profile.name,
                        profile_id=created_profile.uid,
                        department=created_profile.department,
                        profile_pic_url=created_profile.profile_image_url or "None",
                        profile_type=profile_type,
                        description=created_profile.about,
                        experience=created_profile.experience,
                        expertise=created_profile.area_domain,
                        status=StatusChoice.available,
                        skills=created_profile.high_rating_characteristics,
                        is_visible= True,
                        is_approved =  True if (created_profile.profile_type) in ('coachee','mentee') else profile_approved,
                        ai_email = generate_email(created_profile.name,created_profile.id)
                        )
                
                reset_cache_with_prefix('profiles_by_user_id')
                reset_cache_with_prefix('profile_by_id')
                reset_cache_with_prefix('all_profiles')
                
                return Response({"data": CoachCoacheeMentorMenteeProfileSerializer(created_profile).data },status=status.HTTP_200_OK)
            except Exception as e:
                logger.exception(e)
                error_msg = f"failed to create profile: {e}\n\n"
                error_msg += traceback.format_exc()
                # send_slack_message({"module": "########### coach_coachee_mentor_mentee_profile ###########", "error": str(e)})
                send_error_notification("coach_coachee_mentor_mentee_profile",error_msg,{"data":data})
                return Response({"error":"got error"},status=status.HTTP_400_BAD_REQUEST)
            
    @action(methods=['GET','POST'], detail=False, url_path="update-coach-mentor-meeting-availability")
    def update_coach_mentor_meeting_availability(self,request,*args, **kwargs):
        if request.method == 'GET':
            user_id = request.query_params.get('user_id',None)
            profile_id = request.query_params.get('profile_id',None)
            if not user_id and not profile_id:
                return Response({"error": "user_id or profile_id is required"},status=status.HTTP_400_BAD_REQUEST)
            
            try:
                if user_id:
                    profile = CoachCoacheeMentorMenteeProfile.objects.get(deleted=False,tenant_id=self.request.tenant.uid,user_id=user_id)
                elif profile_id:
                    profile = CoachCoacheeMentorMenteeProfile.objects.get(deleted=False,tenant_id=self.request.tenant.uid,uid=profile_id)
                return Response({"data": profile.meeting_availability },status=status.HTTP_200_OK)
            except Exception as e:
                logger.exception(e)
                return Response({"error": f"{e.args}"},status=status.HTTP_400_BAD_REQUEST)
            
        elif request.method == 'POST':
            user_id = request.data.get('user_id',None)
            profile_id = request.data.get('profile_id',None)
            availability = request.data.get('availability',None)
            
            logger.info(f"*************** user_id: {user_id}, profile_id: {profile_id}, availability: {availability} ***************")
            
            if not user_id and not profile_id:
                return Response({"error": "user_id or profile_id is required"},status=status.HTTP_400_BAD_REQUEST)
            
            if not availability:
                return Response({"error": "availability is required"},status=status.HTTP_400_BAD_REQUEST)
            try:
                if user_id:
                    profile = CoachCoacheeMentorMenteeProfile.objects.get(deleted=False,tenant_id=self.request.tenant.uid,user_id=user_id)
                elif profile_id:
                    profile = CoachCoacheeMentorMenteeProfile.objects.get(deleted=False,tenant_id=self.request.tenant.uid,uid=profile_id)
                
                if profile is None:
                    return Response({"error": "profile not found"},status=status.HTTP_400_BAD_REQUEST)
                profile.meeting_availability = availability
                profile.save(update_fields=["meeting_availability"])
                return Response({"data": CoachCoacheeMentorMenteeProfileSerializer(profile).data },status=status.HTTP_200_OK)
            except Exception as e:
                logger.exception(e)
                return Response({"error": f"{e.args}"},status=status.HTTP_404_NOT_FOUND)
                    

    @action(methods=['GET'], detail=False, url_path="get-bots")
    def get_bots(self,request,*args, **kwargs):
        user_id = request.query_params.get('user_id',None)
        bot_type = request.query_params.get('bot_type',None)
        client_name = request.query_params.get('client_name',None)
        approved_only = request.query_params.get('approved_only',None)
        approved_only = True if approved_only is not None and approved_only in ["true","True"] else False
        tenant_id = self.request.tenant.uid

        logger.info(f"################### user_id: {user_id}, bot_type: {bot_type}, client_name: {client_name} , approved_only: {approved_only} ###################")
        
        cache_key = generate_cache_key(
            'get_bots', tenant_id=tenant_id, user_id=user_id, bot_type=bot_type, client_name=client_name, approved_only=approved_only
        )

        # Attempt to retrieve data from cache
        cached_data = get_cache(cache_key)
        if cached_data is not None:
            return Response({"data": cached_data}, status=status.HTTP_200_OK)
        
        all_bots = SignatureBot.objects.filter(deleted=False,tenant_id=tenant_id)
        if approved_only:
            all_bots = all_bots.filter(is_approved=True)
        data = []

        if user_id:
            all_bots = all_bots.filter(user_id=user_id)

        if bot_type:
            all_bots = all_bots.filter(bot_type=bot_type)
        

        deepdive_bot_access = None
        if client_name:
            user_ids = []
            bot_user_ids = list(all_bots.values_list('user_id',flat=True))
            for u_id in bot_user_ids:
                user_email = UserAttribute.objects.get(deleted=False,tenant_id=tenant_id,user_id=u_id).attributes.get('email',None)
                client = ClientUserInfo.objects.filter(deleted=False,tenant_id=tenant_id,member_emails__contains=user_email).first()
                if client:
                    # data.append({"allowed_ips": client.allowed_ips})
                    if bot_type == BotTypeChoice.deep_dive:
                        deepdive_bot_access = client.deepdive_accessed_emails.spllit(',') if client.deepdive_accessed_emails else []
                    if client.client_name == client_name:
                        user_ids.append(u_id)

            logger.info(f"client members ids: {user_ids}")
            all_bots = all_bots.filter(user_id__in = user_ids)

        for bot in all_bots:
            serializer = SignatureBotSerializer(bot)
            bot_att = BotAttribute.objects.get(bot_id=bot.uid)
            botser = BotAttributeSerializer(bot_att)
            all_bots_data = {"creator_name":serializer.data.get('creator_name') ,"signature_bot": serializer.data,
                            "bot_attributes": botser.data}
            if deepdive_bot_access:
                all_bots_data['deepdive_access'] = deepdive_bot_access
            data.append(all_bots_data)
        set_cache(cache_key, data)
        return Response({"data": data},status=status.HTTP_200_OK)
        

    #************* utility methods ***************
    def process_and_store_youtube_transcript(self,youtube_links,signature_bot,overwrite=False, deleted_data = {}):
        extracted_from_youtube = {}
        extracted_media_data = {}
        
        if not isinstance(youtube_links, list):
            youtube_links = [youtube_links]

        logger.info(f"*************** youtube_links in process_and_store : {youtube_links}")

        transcript = None
        for link in youtube_links:
            if link != '':
                logger.info(f"Gettinggs transcript for youtube link: {link}")
                try:
                    for i in range(2):
                        transcript = get_youtube_transcript(link)
                        if transcript is not None:
                            logger.info(f"transcript: {transcript}")
                            break
                    if transcript is None:
                        logger.info(f"Could not get transcript for youtube link: {link} from package so trying to download and transcribe")
                        transcript = download_and_transcribe_audio(link)
                        logger.info(f"Transcript after download and transcribe : {transcript}")
                    if signature_bot.bot_type in [BotTypeChoice.avatar_bot, BotTypeChoice.subject_specific_bot]:
                        extracted_media_data[link] = transcript
                        transcript = get_document_summary(transcript)
                    extracted_from_youtube[link] = transcript
                except Exception as e:
                    logger.exception(e)
                    extracted_from_youtube[link] = {"error": "Restricted video. error in extracting transcript. Please try another."}
            
        # extracted_media_data['extracted_from_youtube'] = extracted_from_youtube
        logger.info(f"extratedz youtube: {extracted_from_youtube}")
        signature_bot.refresh_from_db()
        bot_media_data = signature_bot.data['media_data']
        if overwrite and extracted_from_youtube:
            bot_media_data['extracted_from_youtube'] = extracted_from_youtube
        else:
            prev_extracted_from_youtube = bot_media_data.get('extracted_from_youtube',{})
            # if "youtube_links" in deleted_data:
            #     for link in deleted_data["youtube_links"].strip().split(","):
            #         prev_extracted_from_youtube.pop(link.strip(),None)
            bot_media_data['extracted_from_youtube'] = {**prev_extracted_from_youtube,**extracted_from_youtube}

        signature_bot.data['media_data'] = bot_media_data
        signature_bot.save(update_fields=["data"])

        bot_att = BotAttribute.objects.get(bot_id=signature_bot.uid)
        bot_att.refresh_from_db()
        bot_media_data = bot_att.extracted_documents if bot_att.extracted_documents else {}
        bot_media_data['extracted_from_youtube'] = {**bot_media_data.get('extracted_from_youtube',{}),**extracted_media_data}
        bot_att.extracted_documents = bot_media_data
        bot_att.save(update_fields=["extracted_documents"])
        
        return transcript

    
    @action(methods=['POST','PATCH'],detail=False, url_path="create-bot-by-details")
    def create_bot_by_details(self, request, *args, **kwargs):
        """
    Creates or updates a bot based on the provided details in the request. This method handles both POST and PATCH requests.

    For a POST request, it creates a new bot with the specified attributes and media data. It requires details such as bot type, participant ID, bot name, and attributes. Optionally, media data like YouTube links, article links, PDFs, and documents can be included for content extraction and storage.

    For a PATCH request, it updates an existing bot, primarily used for re-approving or updating bot details and media content. It can also handle media data processing similarly to POST.

    Parameters:
        request (HttpRequest): The request object containing all necessary data.
        *args: Variable length argument list.
        **kwargs: Arbitrary keyword arguments.

    Returns:
        HttpResponse: JSON response containing the bot ID and status or error message.

    Raises:
        HTTP_400_BAD_REQUEST: If required fields are missing or validation fails.
        HTTP_404_NOT_FOUND: If the bot to be updated does not exist.

    Example:
        POST /create-bot-by-details
        Request Body:
            {
                "bot_type": "avatar_bot",
                "participant_id": "12345",
                "bot_name": "Health Bot",
                "attributes": {"description": "A bot that provides health advice."},
                "media_data": {"youtube_links": "http://youtube.com/example"}
            }
        Response:
            {
                "bot_id": "avatar_bot-12345-health-bot",
                "bot_uid": "abcd1234"
            }

        PATCH /create-bot-by-details
        Request Body:
            {
                "bot_id": "avatar_bot-12345-health-bot",
                "updated_data": {"bot_name": "Updated Health Bot"}
            }
        Response:
            {
                "msg": "updated"
            }
    """


        data = request.data

        # tenant = self.request.tenant
        # print("Tenant: ",tenant, "T"*100)

        try:
            if request.method == 'POST':
                profile_id = data.get('profile_id',None)
                try:
            
                    bot_type = data.get('bot_type')
                    if bot_type is None or bot_type == '' or bot_type not in [choice[0] for choice in BotTypeChoice.choices]:
                        return Response({"error": "bot_type is required"},status=status.HTTP_400_BAD_REQUEST)
                    
                    if (profile_id is None or profile_id == '' ) and bot_type not in [BotTypeChoice.feedback_bot ,BotTypeChoice.user_bot, BotTypeChoice.deep_dive]:
                        return Response({"error": "profile_id is required"},status=status.HTTP_400_BAD_REQUEST)

                    context = data.get('context')
                    bot_title = data.get('bot_title')
                    bot_objective = data.get('bot_objective')
                    if (context is None or context == '' ) and bot_type in [BotTypeChoice.deep_dive]:
                        return Response({"error": "context is required"},status=status.HTTP_400_BAD_REQUEST)
                    if (data.get('bot_title') is None or data.get('bot_title')  == '' ) and bot_type in [BotTypeChoice.deep_dive]:
                        return Response({"error": "bot_title is required"},status=status.HTTP_400_BAD_REQUEST)
                    
                    if (bot_objective is None or bot_objective  == '' ) and bot_type in [BotTypeChoice.deep_dive]:
                        return Response({"error": "bot_objective is required"},status=status.HTTP_400_BAD_REQUEST)
                    

                    participant_id = data.get('participant_id')
                    if participant_id is None or participant_id == '':
                        return Response({"error": "participant_id is required"},status=status.HTTP_400_BAD_REQUEST)

                    bot_name = data.get('bot_name')
                    if bot_name is None or bot_name == '':
                        return Response({"error": "bot_name is required"},status=status.HTTP_400_BAD_REQUEST)

                    bot_id = generate_bot_id(
                        bot_type=bot_type,
                        participant_id=participant_id,
                        bot_name=bot_name
                        )
                    existing_bots = SignatureBot.objects.filter(bot_id=bot_id,tenant_id=self.request.tenant.uid,deleted=False)
                    if existing_bots.count() > 0:
                        return Response({"error": "Bot already exists"},status=status.HTTP_400_BAD_REQUEST)

                    try:
                        user = User.objects.get(uid=participant_id)
                    except:
                        return Response({"error": "User not found"},status=status.HTTP_404_NOT_FOUND)
                    

                    bot_attributes = data.get('attributes')
                    if bot_attributes is None or bot_attributes == '':
                        return Response({"error": "attributes is required"},status=status.HTTP_400_BAD_REQUEST)
                    
                    feedback_questions = data.get("feedback_questions")
                    if bot_type == BotTypeChoice.feedback_bot:
                        if feedback_questions is None or feedback_questions == '':
                            return Response({"error": "feedback_questions is required"},status=status.HTTP_400_BAD_REQUEST)

                    fitment_answer = data.get('fitment_answer',None)
                    if bot_type == BotTypeChoice.avatar_bot:
                        if fitment_answer is None or fitment_answer == "":
                            return Response({"error": "fitment_answer is required"},status=status.HTTP_400_BAD_REQUEST)
                        
                    bot_base_url = data.get('bot_base_url',None)
                    if not bot_base_url:
                        return Response({"error": "bot_base_url is required"},status=status.HTTP_400_BAD_REQUEST)
                    
                    expiry_date = data.get('expiry_date',datetime.datetime.today() + datetime.timedelta(weeks=2))
                    
                    bot_approved = data.get('is_approved',False)
                    bot_scenario_case = data.get('bot_scenario_case',None)
                    

                    faqs = data.get('faqs')
                    fitment_details = data.get('fitment_data',None)
                    initial_questions = data.get('initial_questions',None)
                    media_data = data.get('media_data')
                    bot_details = data.get('bot_details',{})
                    additional_data = data.get('additional_data')
                    coach_data = data.get('coach_data')

                    bot_details["is_login_required"] = False
                    bot_details["is_strict_login_required"] = False
                    

                    all_data = {}
                    deep_dive_data = {}
                    if bot_type == BotTypeChoice.deep_dive:
                        additional_prompt_for_deepdive = data.get('additional_prompt_for_deep',None)
                        # deep_dive_data = generate_title_and_objective_for_deep_dive(context,additional_prompt=additional_prompt_for_deepdive)
                        all_data['bot_title'] = bot_title
                        all_data['bot_objective'] = bot_objective
                        all_data['bot_context'] = bot_objective + '\n' + context
                        deep_dive_data = all_data

                    all_data['coach_data'] = coach_data

                    extracted_media_data = {}

                    if additional_data:
                        all_data['additional_data'] = additional_data

                    print("################# media_data: ",media_data)

                    if 'attatched_pdfs' in request.data:
                        if media_data is None:
                            media_data = {}
                        media_data = json.loads(media_data)
                        media_data['attatched_pdfs'] = request.data.getlist('attatched_pdfs')
                        logger.info(f"*************** attached_pdfs files in request: {media_data['attatched_pdfs']}, <<tag>>:{data.get('tag')}")
                    extracted_media_data = {}

                    logger.info(f"*************** attached_pdfs files in request: {request.data}, $$$$$$$$ {'attatched_pdfs' in request.data}")

                    client = None
                    try:
                        user_identity = Identity.objects.get(user_id=participant_id)
                        user_email = user_identity.value
                        client = ClientUserInfo.objects.get(deleted=0,tenant_id=request.tenant.uid,member_emails__icontains=user_email)
                    except Exception as e:
                        logger.exception(f"Client not found for user: {e}")
                        

                    widget_snippet = f"""
                        <script src="https://playground.coachbots.com/widget/coachbots-stt-widget.js"></script>
                        <div
                        data-client-id="{client.client_name if client else ''}"
                        data-allow-audio-interaction="true"
                        data-is-demo=""false"
                        class="coachbots-coachscribe"
                        data-bot-id="{bot_id}"
                        ></div>
                    """
                    signature_bot = SignatureBot.objects.create(
                        bot_id=bot_id,
                        tenant_id=self.request.tenant.uid,
                        user_id=participant_id,
                        bot_type=bot_type,
                        tag=data.get('tag'),
                        data = {
                            "media_data": {

                            }
                        },
                        is_approved= True if bot_type == BotTypeChoice.deep_dive else bot_approved ,
                        integratable_widget_snippet = widget_snippet,
                    )

                    bot_att = BotAttribute.objects.create(tenant_id=self.request.tenant.uid,
                                                        bot_id=signature_bot.uid,
                                                        bot_name=bot_name,
                                                        coach_name=get_user_display_name(get_user_by_id(signature_bot.user_id))
                                                        )
                    
                    # fitment_data= {"options": {"1": ["Weekly commitments", "Bi-weekly sessions", "Monthly sessions", "Ad Hoc / On demand", "Customized Structure"], "2": ["Career advancement", "Skill development", "Introspection & Reflection", "Industry insights", "Networking & Leadership"], "3": ["Technology", "Business Operations", "Industrial Operations", "Sales & Marketing", "HR & People Management"]}, "mentee_que": {"1": "How much time and commitment are you prepared to invest in a mentoring/coaching journey? Are there specific timelines or availability constraints you'd like your mentor/coach to consider?", "2": "What are your expectations from this mentoring/coaching session - in terms of key outcome areas?", "3": "What are the hard skill areas, if any, you want to improve upon?"}, "mentor_que": {"1": "How much time are you willing to commit to mentoring/coaching? Are there specific times or days that work best for you, or any constraints you'd like a potential mentee to be aware of?", "2": "What are your expectations from this mentoring/coaching session - in terms of key outcome areas?", "3": "What are the hard skill areas, if any, you want to contribute in?"}, "fitment_measures": {"mid": "The score reflects a promising yet moderate fit in coaching dynamics. Acknowledge existing areas for improvement and work collaboratively to address specific concerns. Proactively work on refining coaching dynamics to elevate the overall experience for both the coach and coachee. Continuous effort and attention to areas of improvement can lead to a more effective coaching partnership.", "top": "The score refelcts a robust alignment between the coach and coachee, laying the foundation for an optimal coaching relationship. The coaching relationship is optimal, providing a strong foundation for success. Maintain and nurture open communication and collaboration, as these are key elements in sustaining the excellence of the coaching dynamic.", "bottom": "The score signals a notable discord in coaching dynamics, this suggests a reconsideration of the coaching relationship. Misalignment may impede progress, so exploring alternative matches could unveil better synergies and enhance overall effectiveness. Re-evaluate if the coaching partnership aligns with the coachee's goals and needs."}}
                    fitment_data= {
                        "options":{
                        "1": ["Anyone above me","Same or up to two level below","Any level"],
                        "2": ["Yes","No"],
                        "3": ["Career advancement","Skill development", "Introspection & reflection", "Networking & leadership"]
                        },
                        "mentee_que": {"1": "What level of coach & mentor do you want?", "2": "I want a coach & mentor someone from the same department.", "3": "What kind of outcome do you want from these sessions the most?"},
                        "mentor_que": {"1": "What level of participant do you want to coach & mentor?", "2": "I want to coach & mentor someone in the same department.", "3": "What kind of outcome can you support in these sessions the most?"}
                        }
                    if fitment_details:
                        fitment_data = fitment_details

                    fitment_data["fitment_measures"] = {
                                                        "top": "The score refelcts a robust alignment between the coach and coachee, laying the foundation for an optimal coaching relationship. The coaching relationship is optimal, providing a strong foundation for success. Maintain and nurture open communication and collaboration, as these are key elements in sustaining the excellence of the coaching dynamic.",
                                                        "bottom": "The score signals a notable discord in coaching dynamics, this suggests a reconsideration of the coaching relationship. Misalignment may impede progress, so exploring alternative matches could unveil better synergies and enhance overall effectiveness. Re-evaluate if the coaching partnership aligns with the coachee's goals and needs.",
                                                        "mid": "The score reflects a promising yet moderate fit in coaching dynamics. Acknowledge existing areas for improvement and work collaboratively to address specific concerns. Proactively work on refining coaching dynamics to elevate the overall experience for both the coach and coachee. Continuous effort and attention to areas of improvement can lead to a more effective coaching partnership."
                                                    }
                    
                    initial_qna = {
                            "1": "Thank you for considering a virtual session. Please let me know more about you as a person that you think might be relevant to our session today.",
                            "2": "What do you want to achieve with your session with me today - let me know the goals you have in mind.",
                            "3": "What specific problems you are facing currently that are a priority for you? What have you tried so far in terms of finding your solutions?",
                            "4": "Do you believe your solutions have worked so far? Why or why not?",
                            "5": {"options": ["Yes", "No"], "question": "Is this discussion related to your goal in a way to consider your IDP (individual development plan)? "}
                        }
                    if initial_questions:
                        initial_qna = initial_questions
                    updated_fields = []
                    if bot_details:
                        signature_bot.bot_details = bot_details
                        updated_fields.append("bot_details")
                    else:
                        signature_bot.bot_details = {"info":bot_attributes['heading']}
                        updated_fields.append("bot_details")

                    if bot_attributes:
                        signature_bot.attributes = bot_attributes
                        updated_fields.append("attributes")

                    if bot_scenario_case:
                        signature_bot.bot_scenario_case = bot_scenario_case
                        updated_fields.append('bot_scenario_case')

                    if data.get('is_private',None) != None:
                        signature_bot.is_private = True if data.get('is_private') in ['True','true',1,True] else False
                        updated_fields.append('is_private')


                    if faqs:
                        try:
                            new_faqs = json.loads(faqs)
                        except:
                            new_faqs = faqs
                            
                        signature_bot.faqs = new_faqs
                        updated_fields.append("faqs")

                    if bot_type in [BotTypeChoice.avatar_bot, BotTypeChoice.subject_specific_bot]:
                        signature_bot.custom_prompt = signature_bot_default_prompt(bot_type=bot_type)
                        updated_fields.append("custom_prompt")

                    if bot_type == BotTypeChoice.deep_dive:
                        prompt = signature_bot_default_prompt(bot_type)
                        signature_bot.custom_prompt = prompt
                        updated_fields.append("custom_prompt")
                        
                        
                    if bot_type == BotTypeChoice.feedback_bot:
                        low_skill = data.get("low_rating_characteristics")
                        high_skill = data.get("high_rating_characteristics")
                        
                        if None in [low_skill, high_skill]:
                            return Response({"error": "low_rating_characteristics and high_rating_characteristics is required"},status=status.HTTP_400_BAD_REQUEST)
                        
                        sync_user_low_high_skills(self.request.tenant.uid, participant_id, low_skill, high_skill)
                        

                    if all_data:
                        bot_data = {**signature_bot.data,**all_data}
                        signature_bot.data = bot_data
                        updated_fields.append("data")
                        
                    if bot_type == BotTypeChoice.feedback_bot:
                        signature_bot.is_approved = True
                        updated_fields.append('is_approved')

                    if bot_type == BotTypeChoice.deep_dive:
                        signature_bot.bot_expires_at = expiry_date
                        updated_fields.append('bot_expires_at')

                        access_code = get_unique_deep_dive_access_code(tenant=self.request.tenant)
                        signature_bot.access_code = access_code
                        updated_fields.append('access_code')
                        deep_dive_data['bot_expires_at'] = expiry_date
                        deep_dive_data['access_code'] = access_code
                        
                    

                    if updated_fields:
                        signature_bot.save(update_fields=updated_fields)

                    updated_fields = []
                    if fitment_answer and bot_type == BotTypeChoice.avatar_bot:
                        bot_att.fitment_answers = {"mentor_answer": fitment_answer.split(",")}
                        bot_att.fitment_data = fitment_data
                        updated_fields.extend(["fitment_answers","fitment_data"])

                    if feedback_questions:
                        bot_att.feedback_questions = feedback_questions
                        updated_fields.append("feedback_questions")

                    email = data.get('email',None)
                    if email:
                        bot_att.coach_email = email
                        updated_fields.append("coach_email")
                    
                    if additional_data:
                        if bot_type == BotTypeChoice.subject_specific_bot and additional_data.get("bot_description",None):
                            bot_att.about = additional_data.get("bot_description",None)
                            updated_fields.append('about')

                        else:
                            bot_att.about = additional_data.get("profile_description",None)
                            updated_fields.append('about')

                    if initial_qna and bot_type not in [BotTypeChoice.feedback_bot, BotTypeChoice.user_bot, BotTypeChoice.deep_dive]:
                        bot_att.initial_qnas = initial_qna
                        updated_fields.append("initial_qnas")

                    
                    if updated_fields:
                        bot_att.save(update_fields=updated_fields)

                    # SAVING BOTURL AND bot_snippets
                    try: 
                        bot_url =''
                        if bot_type in [BotTypeChoice.avatar_bot, BotTypeChoice.subject_specific_bot]:
                            bot_url = f"{bot_base_url}/coach/{bot_id}"
                        elif bot_type == BotTypeChoice.feedback_bot:
                            bot_url = f"{bot_base_url}/feedback/{bot_id}"
                        elif bot_type == BotTypeChoice.subject_matter_bot:
                            bot_url = f"{bot_base_url}/subject-expert/{bot_id}"
                        elif bot_type == BotTypeChoice.helper_bot:
                            bot_url = f"{bot_base_url}/subject-expert/{bot_id}"
                        elif bot_type == BotTypeChoice.user_bot:
                            bot_url = f"{bot_base_url}/knowledge-bot/{bot_id}"
                        elif bot_type == BotTypeChoice.deep_dive:
                            bot_url = f"{bot_base_url}/engagement-survey/{bot_id}"

                        bot_snippet = f"""
                                    <div class="deep-chat-poc2" data-bot-id="{bot_id}"></div>
                                    <script src="{bot_base_url}/widget/coachbots-stt-widget.js" defer></script>
                                        """
                        coach_profile = CoachCoacheeMentorMenteeProfile.objects.filter(deleted=0,uid=profile_id).first()
                        if coach_profile:
                            coach_profile.bot_urls = (coach_profile.bot_urls + f", {bot_url}") if coach_profile.bot_urls else bot_url
                            coach_profile.bot_ids = (coach_profile.bot_ids + f", {bot_id}") if coach_profile.bot_ids else bot_id
                            snippet = coach_profile.bot_snippets
                            if snippet:
                                snippet[bot_type] = bot_snippet
                            else:
                                snippet = {f"{bot_type}": bot_snippet}
                            
                            coach_profile.bot_snippets = snippet

                                

                            coach_profile.save(update_fields=["bot_urls","bot_ids","bot_snippets"])

                        directory = DirectoryPageInfo.objects.filter(profile_id = profile_id).first()
                        if bot_type in [BotTypeChoice.avatar_bot, BotTypeChoice.subject_specific_bot]:
                            if directory:
                                if bot_type == BotTypeChoice.avatar_bot:
                                    directory.avatar_bot_id = bot_id
                                    directory.avatar_snippit = bot_snippet
                                    directory.avatar_bot_url = bot_url
                                    directory.save(update_fields=["avatar_bot_id","avatar_snippit","avatar_bot_url"])

                                elif bot_type == BotTypeChoice.subject_specific_bot:
                                    directory.subject_specific_bot_id = bot_id
                                    directory.subject_specific_bot_url = bot_url
                                    directory.subject_specific_bot_snippit = bot_snippet
                                    directory.save(update_fields=["subject_specific_bot_id","subject_specific_bot_url","subject_specific_bot_snippit"])


                                try:
                                    # subject = ""
                                    # html = ""
                                    # if bot_type == BotTypeChoice.avatar_bot:
                                    #     subject = "Your Coachbots AI Frame is in the Pipeline"
                                    #     html = f"""
                                    #         <p style="font-family: sans-serif; font-size: 14px; font-weight: normal; margin: 0; margin-bottom: 15px;">
                                    #             <div style="margin: 15px;">
                                    #                 <p>Thank you for joining the Coachbots network as a coach/mentor. Your AI Frame is currently in the processing pipeline, and we will send you a confirmation when it's ready.</p>
                                    #                 <p>Once your AI Frame is approved, you'll have full access to the platform and can begin leveraging its features to support your coaching engagements.</p>
                                    #                 <p>We're excited to have you on board and look forward to empowering you to make a meaningful impact on your coachees' journeys.</p>
                                    #                 <p>If you have any questions or need assistance, please don't hesitate to reach out to our support team.</p>
                                    #             </div>
                                    #         </p>
                                    #         """
                                    # elif bot_type == BotTypeChoice.subject_specific_bot:
                                    subject = "Your Coachbot Co-pilot is in the Pipeline"
                                    html = f"""
                                        <p style="font-family: sans-serif; font-size: 14px; font-weight: normal; margin: 0; margin-bottom: 15px;">
                                            <div style="margin: 15px;">
                                                <p>Thank you for joining the Coachbot network. Your Copilot is currently in the processing pipeline, and we will send you a confirmation when it's ready.</p>
                                                <p>Once your Copilot is approved, you'll have full access to the platform and can begin leveraging its features to support your coaching engagements.</p>
                                                <p>We're excited to have you on board and look forward to empowering you to make a meaningful impact on your coachees' journeys.</p>
                                                <p>If you have any questions or need assistance, please don't hesitate to reach out to our support team.</p>
                                            </div>
                                        </p>
                                        """


                                    send_email_with_html_template(subject=subject,html_content=html,to_email=email,title=f'Hey {coach_profile.name}!')
                                    html = f"""
                                        <p style="font-family: sans-serif; font-size: 14px; font-weight: normal; margin: 0; margin-bottom: 15px;">{coach_profile.name} - {coach_profile.email} created a bot/profile. Please check it out and approve it from Django Admin Panel.</p>
                                        """
                                    send_email_with_html_template(subject=subject,html_content=html)
                            
                                except Exception as e:
                                    logger.exception(f"Ai frame creation email is failed reason: {e}")
                                    error_msg = f"Ai frame creation email is failed reason: {e}\n\n"
                                    error_msg += traceback.format_exc()
                                    send_error_notification("create_bot_by_details",error_msg,{"bot_id":bot_id,"profile_id":profile_id,'email': email, 'profile': coach_profile})



                            
                        if bot_type == BotTypeChoice.feedback_bot:
                            if directory:
                                directory.feedback_wall = bot_url
                                directory.save(update_fields=['feedback_wall'])

                        if bot_type in [BotTypeChoice.user_bot] :
                            new_dir = DirectoryPageInfo.objects.create(
                            name=bot_name,
                            department=coach_profile.department if coach_profile else "",
                            profile_id= user.uid, # in case of user_bot or deep_dive storing user id instead of profile id
                            profile_pic_url=coach_profile.profile_image_url if coach_profile else "https://res.cloudinary.com/dtbl4jg02/image/upload/v1710139318/mdzmknenvvv4llgevykz.png",
                            profile_type= ProfileTypeChoice.knowledge_bot,
                            description=coach_profile.about if coach_profile else "No Description",
                            experience=coach_profile.experience if coach_profile else "",
                            expertise=coach_profile.area_domain if coach_profile else "",
                            status=StatusChoice.available,
                            skills=coach_profile.high_rating_characteristics if coach_profile else "",
                            is_visible= False,
                            ai_email = generate_email(coach_profile.name,coach_profile.id) if coach_profile else None
                            )
                            
                            if bot_type == BotTypeChoice.user_bot:
                                new_dir.custom_user_bot_url = bot_url
                                new_dir.custom_user_bot_id = bot_id
                                new_dir.is_approved = False

                            # elif bot_type == BotTypeChoice.deep_dive:
                            #     new_dir.deep_dive_bot_url = bot_url
                            #     new_dir.deep_dive_bot_id = bot_id
                            #     new_dir.is_approved = True


                            new_dir.save()

                            try:
                                subject = "Knowledge Bots"
                                html = f"""
                                    <p style="font-family: sans-serif; font-size: 14px; font-weight: normal; margin: 0; margin-bottom: 15px;">Thank you for creating your knowledge bot- <b>{bot_name}</b>. It is under processing pipeline and you will soon receive a confirmation when it's live. You can always edit the same via the profile section.</p>
                                    """

                                send_email_with_html_template(subject=subject,html_content=html,to_email=email,title=f'Hey!')
                                html = f"""
                                    <p style="font-family: sans-serif; font-size: 14px; font-weight: normal; margin: 0; margin-bottom: 15px;">{user.name} - {email} created a knowledge bot - {bot_name}. Please check it out and approve it from Django Admin Panel.</p>
                                    """
                                send_email_with_html_template(subject=subject,html_content=html)
                        
                            except Exception as e:
                                logger.exception(f"Knowledge bot creation email is failed reason: {e}")
                                error_msg = f"Knowledge bot creation email is failed reason: {e}\n\n"
                                error_msg += traceback.format_exc()
                                send_error_notification("create_bot_by_details",error_msg,{"bot_id":bot_id,"profile_id":profile_id,'email': email, 'profile': coach_profile})

                    except Exception as e:
                        logger.exception(f"couldn't save bot_url in CoachCoacheeMentorMenteeProfile")
                        error_msg = f"couldn't save bot_url in CoachCoacheeMentorMenteeProfile: {e}\n\n"
                        error_msg += traceback.format_exc()
                        send_error_notification("create_bot_by_details",error_msg,{"bot_id":bot_id,"profile_id":profile_id})
                        
                    reset_cache_with_prefix("get_bots")
                    return Response({"bot_id":signature_bot.bot_id,"bot_uid": signature_bot.uid, 'deep_dive_data': deep_dive_data },status=status.HTTP_200_OK)
                
                except Exception as e:
                    logger.exception("Got error while creating bot: {e}")
                    bot_type = data.get('bot_type')
                    if bot_type in [BotTypeChoice.avatar_bot, BotTypeChoice.subject_specific_bot]:
                        try:
                            coach_profile = CoachCoacheeMentorMenteeProfile.objects.get(deleted=0,uid=profile_id)
                            coach_profile.deleted = True
                            coach_profile.save(update_fields=["deleted"])
                            DirectoryPageInfo.objects.filter(profile_id=profile_id).delete()
                        except Exception as e:
                            logger.error(f"Got error in crating bot : profile_id is missing")
                            send_error_notification("create_bot_by_details",f"Got error in crating bot : profile_id is missing: {e}",{"data":data})

                    return Response({"msg":f"Got error : {e}" },status=status.HTTP_400_BAD_REQUEST)
            
            
            
            
            
            
            elif request.method == "PATCH":
                bot_id = data.get("bot_id",None)
                profile_id = data.get("profile_id",None)
                for_reapproval = data.get('for_reapproval',None).lower().strip() == 'true' if data.get('for_reapproval',None) else False
                deleted_data = data.get('deleted_data',None)
                deleted_data = json.loads(deleted_data) if deleted_data is not None else deleted_data
                
                logger.info(f"{'$'*100} deleted_data : {deleted_data}")
                
                try:
                    signature_bot = SignatureBot.objects.get(deleted=False,tenant_id=self.request.tenant.uid,uid=bot_id)
                    bot_att = BotAttribute.objects.get(bot_id=signature_bot.uid)
                except SignatureBot.DoesNotExist:
                    return Response({"error": "SignatureBot not found"}, status=status.HTTP_404_NOT_FOUND)
                
                
                # delete files data specified in deleted_data
                
                
                if deleted_data is not None:
                    bot_media_data = signature_bot.data['media_data']
                    #**** delete articles data ***
                    if  "article_links" in deleted_data:
                        prev_extracted_from_article = bot_media_data.get('extracted_from_article',{})
                        logger.info(f"<<<<<<<<<<<<<<<<<<>>>>>>>>>>>>>>>> prev_extracted_from_article before: {prev_extracted_from_article}")
                        
                        for link in deleted_data["article_links"].split(","):
                            prev_extracted_from_article.pop(link,None)
                            
                        logger.info(f"<<<<<<<<<<<<<<<<<<>>>>>>>>>>>>>>>> prev_extracted_from_article after: {prev_extracted_from_article}")
                        bot_media_data['extracted_from_article'] = prev_extracted_from_article
                        
                    
                    #*** delete pdf's data ***

                    if "pdf_files" in deleted_data:
                        prev_extracted_from_pdf = bot_media_data.get('extracted_from_pdf',{})
                        logger.info(f"<<<<<<<<<<<<<<<<<<>>>>>>>>>>>>>>>> prev_extracted_from_pdf before: {prev_extracted_from_pdf}")
                        
                        for link in deleted_data["pdf_files"].split(","):
                            prev_extracted_from_pdf.pop(link,None)
                            
                        logger.info(f"<<<<<<<<<<<<<<<<<<>>>>>>>>>>>>>>>> prev_extracted_from_pdf after: {prev_extracted_from_pdf}")

                        bot_media_data['extracted_from_pdf'] = prev_extracted_from_pdf
                        
                        
                    #*** delete docs data ***
                    
                    if "doc_files" in deleted_data:
                        prev_extracted_from_doc = bot_media_data.get('extracted_from_doc',{})
                        logger.info(f"<<<<<<<<<<<<<<<<<<>>>>>>>>>>>>>>>> prev_extracted_from_doc before: {prev_extracted_from_doc}")
                        
                        for link in deleted_data["doc_files"].split(","):
                            prev_extracted_from_doc.pop(link,None)
                            
                        logger.info(f"<<<<<<<<<<<<<<<<<<>>>>>>>>>>>>>>>> prev_extracted_from_doc after: {prev_extracted_from_doc}")
                        bot_media_data['extracted_from_doc'] = prev_extracted_from_doc
                    
                    
                    #*** delete youtube links ***
                    
                    if "youtube_links" in deleted_data:
                        prev_extracted_from_youtube = bot_media_data.get('extracted_from_youtube',{})
                        logger.info(f"<<<<<<<<<<<<<<<<<<>>>>>>>>>>>>>>>> prev_extracted_from_youtube before: {prev_extracted_from_youtube}")
                        
                        for link in deleted_data["youtube_links"].split(","):
                            prev_extracted_from_youtube.pop(link,None)
                            
                        logger.info(f"<<<<<<<<<<<<<<<<<<>>>>>>>>>>>>>>>> prev_extracted_from_youtube after: {prev_extracted_from_youtube}")
                        bot_media_data['extracted_from_youtube'] = prev_extracted_from_youtube

                        
                    signature_bot.data['media_data'] = bot_media_data
                    signature_bot.save(update_fields=["data"])

                    #*** delete optional file data ***
                     
                    if "optional_files" in deleted_data:
                        extracted_optional_file_data = bot_att.extracted_documents if bot_att.extracted_documents else {}

                        logger.info(f"<<<<<<<<<<<<<<<<<<>>>>>>>>>>>>>>>> extracted_optional_file_data before: {extracted_optional_file_data}")
                        if extracted_optional_file_data.get('extracted_from_optional_file'):
                            deleted_optional_files = deleted_data["optional_files"].split(",")
                            optional_file_extracted = extracted_optional_file_data['extracted_from_optional_file']
                            for file_name in deleted_optional_files:
                                optional_file_extracted.pop(file_name,None)
                            extracted_optional_file_data['extracted_from_optional_file'] = optional_file_extracted
                            logger.info(f"<<<<<<<<<<<<<<<<<<>>>>>>>>>>>>>>>> extracted_optional_file_data after: {extracted_optional_file_data}")
                            bot_att.extracted_documents = extracted_optional_file_data
                            bot_att.save(update_fields=["extracted_documents"])



                user = get_user_by_id(signature_bot.user_id)
                user_att = UserAttribute.objects.get(deleted=False,user_id=user.uid)


                # sending for reapproval to directory page info
                if signature_bot.bot_type == BotTypeChoice.user_bot:

                    # for user bot or knwoledge bot we are storing user_id instead of profile_id so can multiple row or bots in same user_id
                    directory = DirectoryPageInfo.objects.filter(profile_id=signature_bot.user_id,custom_user_bot_id=signature_bot.bot_id).first() 
                    if directory and for_reapproval:
                            signature_bot.is_approval_email_sent = False
                            signature_bot.is_approved = False
                            signature_bot.save()

                            
                            DirectoryPageInfo.objects.create(
                                name = directory.name,
                                profile_id = directory.profile_id,
                                department = directory.department,
                                bot_type = directory.bot_type,
                                profile_pic_url = directory.profile_pic_url,
                                profile_type = directory.profile_type,
                                description = directory.description,
                                experience = directory.experience,
                                expertise = directory.expertise,
                                status = directory.status,
                                avatar_bot_id = directory.avatar_bot_id,
                                feedback_wall = directory.feedback_wall,
                                skills = directory.skills,
                                is_visible = directory.is_visible,
                                is_approved = False,
                                avatar_snippit = directory.avatar_snippit,
                                avatar_bot_url = directory.avatar_bot_url,
                                custom_user_bot_url = directory.custom_user_bot_url,
                                custom_user_bot_id = directory.custom_user_bot_id,
                                subject_specific_bot_url = directory.subject_specific_bot_url,
                                subject_specific_bot_id = directory.subject_specific_bot_id,
                                subject_specific_bot_snippit= directory.subject_specific_bot_snippit,
                                timer_enabled = directory.timer_enabled,
                                time_value_in_days = directory.time_value_in_days,
                                timer_reset = directory.timer_reset,
                                visual_tag = directory.visual_tag,
                                ai_email = directory.ai_email
                            )


                            # directory.save()
                            try:
                                subject = "Knowledge Bots"
                                html = f"""
                                    <p style="font-family: sans-serif; font-size: 14px; font-weight: normal; margin: 0; margin-bottom: 15px;">Thank you for updating your knowledge bot- <b>{bot_att.bot_name}</b>. It is under processing pipeline and you will soon receive a confirmation when it's live. You can always edit the same via the profile section.</p>
                                    """

                                send_email_with_html_template(subject=subject,html_content=html,to_email=user_att.attributes.get('email'),title=f'Hey!')
                                html = f"""
                                    <p style="font-family: sans-serif; font-size: 14px; font-weight: normal; margin: 0; margin-bottom: 15px;">{user.name} - {user_att.attributes.get('email',"")} updated a knowledge bot - <b>{bot_att.bot_name}</b>. Please check it out and approve it from Django Admin Panel.</p>
                                    """
                                send_email_with_html_template(subject=subject,html_content=html)

                            except Exception as e:
                                logger.error(f"Got error in sending email for reapproval : {e}")
                                send_error_notification("create_bot_by_details",f"Got error in sending email for reapproval : {e}",{"data":data})
                                
                            directory.delete()



                                
                updated_data = data.get("updated_data",None)
                if data.get('is_private') != None:
                        signature_bot.is_private = True if data.get('is_private') in ['true','True',True,1] else False
                        signature_bot.save(update_fields = ['is_private'])

                if signature_bot.bot_type == BotTypeChoice.user_bot:
                    knowledge_bot_faqs =  data.get('faqs',None)
                    if knowledge_bot_faqs:
                        signature_bot.faqs = json.loads(knowledge_bot_faqs) if type(knowledge_bot_faqs) == str else knowledge_bot_faqs
                        signature_bot.save(update_fields=['faqs'])
                    


                if updated_data:
                        
                    updated_data = json.loads(updated_data) if isinstance(updated_data, str) else updated_data
                    if signature_bot.bot_type == BotTypeChoice.feedback_bot:
                        low_skill = updated_data.get("low_rating_characteristics")
                        high_skill = updated_data.get("high_rating_characteristics")
                        
                        if None in [low_skill, high_skill]:
                            return Response({"error": "low_rating_characteristics and high_rating_characteristics is required"},status=status.HTTP_400_BAD_REQUEST)
                        
                        sync_user_low_high_skills(self.request.tenant.uid, signature_bot.user_id, low_skill, high_skill)
                    
                    additional_data = updated_data.get('additional_data')
                    profile_description = additional_data.get("profile_description",None)
                    bot_details = signature_bot.bot_details
                    bot_details['coach_name'] = updated_data.get("bot_name",None)
                    
                    
                    bot_data = signature_bot.data
                    if additional_data:
                        if profile_description:
                            bot_details['info'] = profile_description
                        
                        bot_data["additional_data"] = additional_data

                    signature_bot.bot_details = bot_details
                    signature_bot.data = bot_data

                    updated_fields = ["bot_details","data"]

                    
                    signature_bot.save(update_fields=updated_fields)

                    
                    fitment_answer = updated_data.get('fitment_answers',None)
                    email = updated_data.get("email",None)
                    updated_fields = []

                    if fitment_answer:
                        bot_att.fitment_answers = {"mentor_answer":fitment_answer.split(",")}
                        updated_fields.append("fitment_answers")
                    if email:
                        bot_att.coach_email = email
                        updated_fields.append("coach_email")

                    bot_att.coach_name = updated_data.get("bot_name",None)
                    updated_fields.append("coach_name")

                    feedback_questions = updated_data.get("feedback_questions",None)

                    if feedback_questions:
                        bot_att.feedback_questions = feedback_questions
                        updated_fields.append("feedback_questions")
                    
                    bot_att.save(update_fields=updated_fields)

                media_data = data.get('media_data')
                is_overwrite = data.get('is_overwrite',None)
                if is_overwrite:
                    if is_overwrite.lower() == 'true':
                        is_overwrite = True
                    elif is_overwrite.lower() == 'false':
                        is_overwrite = False
                    else:
                        is_overwrite = False
                else:
                    is_overwrite = False

                if 'attatched_pdfs' in request.data:
                    if media_data is None:
                        media_data = {}
                    media_data = json.loads(media_data)
                    media_data['attatched_pdfs'] = request.data.getlist('attatched_pdfs')
                    logger.info(f"*************** attached_pdfs files in request: {media_data['attatched_pdfs']}")
                extracted_media_data = {}

                logger.info(f"*************** Data in request: {request.data}, $$$$$$$$ {'attatched_pdfs' in request.data}")

                if media_data and signature_bot.bot_type != BotTypeChoice.feedback_bot:
                    media_data = json.loads(media_data) if isinstance(media_data, str) else media_data
                    if 'youtube_links' in media_data:
                        # logger.info(f"################# youtube_links: {media_data['youtube_links']}")
                        youtube_links = media_data['youtube_links']
                        youtube_links = [link.strip() for link in youtube_links.split(',')]

                        print("################# youtube_links: ",youtube_links)

                        #* save these links in bot attributes
                        """ extracted_from_youtube = {}
                        for link in youtube_links:
                            
                            if link != '':
                                transcript_data = download_and_transcribe_audio(link)
                                extracted_from_youtube[link] = transcript_data
                        
                        extracted_media_data['extracted_from_youtube'] = extracted_from_youtube """
                    
                        threading.Thread(target=self.process_and_store_youtube_transcript,args=(youtube_links,signature_bot,is_overwrite, deleted_data)).start()


                    if 'article_links' in media_data:
                        article_links = media_data['article_links']
                        article_links = [link.strip() for link in article_links.split(',')]

                        logger.info(f"******************* article_links: {article_links}")
                        #* save these links in bot attributes
                        extracted_from_article = {}
                        extracted_articles = {}
                        for link in article_links:
                            
                            if link != '':
                                transcript_data = scrape_article_data(link).get('article_content',None)
                                if signature_bot.bot_type in [BotTypeChoice.avatar_bot, BotTypeChoice.subject_specific_bot]:
                                    extracted_articles[link] = transcript_data
                                    transcript_data = get_document_summary(transcript_data)
                                extracted_from_article[link] = transcript_data
                        
                        logger.info(f"******************* extracted_from_article: {extracted_from_article}")
                        signature_bot.refresh_from_db()
                        bot_media_data = signature_bot.data['media_data']
                        if is_overwrite and extracted_from_article:
                            bot_media_data['extracted_from_article'] = extracted_from_article
                        else:
                            prev_extracted_from_article = bot_media_data.get('extracted_from_article',{})
                            # logger.info(f"<<<<<<<<<<<<<<<<<<>>>>>>>>>>>>>>>> prev_extracted_from_article before: {prev_extracted_from_article}")
                            # if  "article_links" in deleted_data:
                            #     for link in deleted_data["article_links"].strip().split(","):
                            #         prev_extracted_from_article.pop(link.strip(),None)
                                    
                            # logger.info(f"<<<<<<<<<<<<<<<<<<>>>>>>>>>>>>>>>> prev_extracted_from_article after: {prev_extracted_from_article}")
                            bot_media_data['extracted_from_article'] = {**prev_extracted_from_article,**extracted_from_article}
                        
                        signature_bot.data['media_data'] = bot_media_data
                        signature_bot.save(update_fields=["data"])

                        bot_att.refresh_from_db()
                        bot_media_data = bot_att.extracted_documents if bot_att.extracted_documents else {}
                        bot_media_data['extracted_from_article'] = {**bot_media_data.get('extracted_from_article',{}),**extracted_articles}
                        bot_att.extracted_documents = bot_media_data
                        bot_att.save(update_fields=["extracted_documents"])

                
                if signature_bot.bot_type != BotTypeChoice.feedback_bot:
                    if 'pdf_data' in data:
                        pdf_data = data.getlist('pdf_data')
                        extracted_from_pdf = {}
                        extracted_pdf = {}
                        logger.info(f"******************* pdf_data: {pdf_data}")

                        if len(pdf_data) > 0:
                            for index, pdf in enumerate(pdf_data):
                                file_name, text = extract_file_and_text(pdf)
                                if signature_bot.bot_type in [BotTypeChoice.avatar_bot, BotTypeChoice.subject_specific_bot]:
                                    extracted_pdf[file_name] = text
                                    text = get_document_summary(text)
                                extracted_from_pdf[file_name] = text
                        logger.info(f"******************* pdf_data: {extracted_from_pdf}")
                        
                        signature_bot.refresh_from_db()
                        bot_media_data = signature_bot.data['media_data']
                        if is_overwrite and extracted_from_pdf:
                            bot_media_data['extracted_from_pdf'] = extracted_from_pdf
                        else:
                            prev_extracted_from_pdf = bot_media_data.get('extracted_from_pdf',{})
                            # logger.info(f"<<<<<<<<<<<<<<<<<<>>>>>>>>>>>>>>>> prev_extracted_from_pdf before: {prev_extracted_from_pdf}")

                            # if "pdf_files" in deleted_data:
                            #     for link in deleted_data["pdf_files"].strip().split(","):
                            #         prev_extracted_from_pdf.pop(link.strip(),None)
                                    
                            # logger.info(f"<<<<<<<<<<<<<<<<<<>>>>>>>>>>>>>>>> prev_extracted_from_pdf after: {prev_extracted_from_pdf}")

                            bot_media_data['extracted_from_pdf'] = {**prev_extracted_from_pdf,**extracted_from_pdf}
                        
                        signature_bot.data['media_data'] = bot_media_data
                        signature_bot.save(update_fields=["data"])

                        bot_att.refresh_from_db()
                        bot_media_data = bot_att.extracted_documents if bot_att.extracted_documents else {}
                        bot_media_data['extracted_from_pdf'] = {**bot_media_data.get('extracted_from_pdf',{}),**extracted_pdf}
                        bot_att.extracted_documents = bot_media_data
                        bot_att.save(update_fields=["extracted_documents"])

                    if 'optional_file' in data:
                        optional_file = data.getlist('optional_file')
                        extracted_from_optional_file = {}
                        extracted_optional_file = {}
                        logger.info(f"******************* optional_file: {optional_file}")

                        if len(optional_file) > 0:
                            for index, optional_file in enumerate(optional_file):
                                file_name, text = extract_file_and_text(optional_file)
                                # if signature_bot.bot_type == BotTypeChoice.avatar_bot:
                                extracted_optional_file[file_name] = text
                                #     text = get_document_summary(text)
                                extracted_from_optional_file[file_name] = text
                        logger.info(f"******************* optional_file: {extracted_from_optional_file}")
                        
                        # signature_bot.refresh_from_db()
                        # bot_media_data = signature_bot.data['media_data']
                        # if is_overwrite and extracted_from_optional_file:
                        #     bot_media_data['extracted_from_optional_file'] = extracted_from_optional_file
                        # else:
                        #     prev_extracted_from_optional_file = bot_media_data.get('extracted_from_optional_file',{})
                        #     bot_media_data['extracted_from_optional_file'] = {**prev_extracted_from_optional_file,**extracted_from_optional_file}
                        
                        # signature_bot.data['media_data'] = bot_media_data
                        # signature_bot.save(update_fields=["data"])

                        bot_att.refresh_from_db()
                        bot_media_data = bot_att.extracted_documents if bot_att.extracted_documents else {}
                        bot_media_data['extracted_from_optional_file'] = {**bot_media_data.get('extracted_from_optional_file',{}),**extracted_optional_file}
                        bot_att.extracted_documents = bot_media_data
                        bot_att.save(update_fields=["extracted_documents"])


                    if 'doc_data' in data:
                        doc_data = data.getlist('doc_data')
                        extracted_from_doc = {}
                        extracted_doc = {}


                        logger.info(f"******************* doc_data: {doc_data}")
                        if len(doc_data) > 0:
                            for index, doc in enumerate(doc_data):
                                file_name, text = extract_file_and_text(doc)
                                if signature_bot.bot_type in [BotTypeChoice.avatar_bot, BotTypeChoice.subject_specific_bot]:
                                    extracted_doc[file_name] = text
                                    text = get_document_summary(text)
                                extracted_from_doc[file_name] = text

                        logger.info(f"******************* doc_data: {extracted_from_doc}")
                        signature_bot.refresh_from_db()
                        bot_media_data = signature_bot.data['media_data']
                        if is_overwrite and extracted_from_doc:
                            bot_media_data['extracted_from_doc'] = extracted_from_doc
                        else:
                            prev_extracted_from_doc = bot_media_data.get('extracted_from_doc',{})
                            if deleted_data is not None and "doc_files" in deleted_data:
                                for link in deleted_data["doc_files"].strip().split(","):
                                    prev_extracted_from_doc.pop(link.strip(),None)
                            bot_media_data['extracted_from_doc'] = {**prev_extracted_from_doc,**extracted_from_doc}
                        
                        signature_bot.data['media_data'] = bot_media_data
                        signature_bot.save(update_fields=["data"])

                        bot_att.refresh_from_db()
                        bot_media_data = bot_att.extracted_documents if bot_att.extracted_documents else {}
                        bot_media_data['extracted_from_doc'] = {**bot_media_data.get('extracted_from_doc',{}),**extracted_doc}
                        bot_att.extracted_documents = bot_media_data
                        bot_att.save(update_fields=["extracted_documents"])


                    if (media_data and 'attached_docs' in media_data) or 'attached_docs' in request.FILES:
                        attached_docs = media_data['attached_docs'] if 'attached_docs' in media_data else request.FILES.getlist('attached_docs')
                        doc_names = []
                        extracted_from_doc = {}
                        extracted_doc = {}

                        for file in attached_docs:
                            # logger.info(f"file name : {file.name}, {file.read()}")
                            doc_names.append(file.name)
                            path = default_storage.save(file.name, ContentFile(file.read()))
                            doc_content = extract_text_from_doc(path)
                            default_storage.delete(path)
                            if signature_bot.bot_type in [BotTypeChoice.avatar_bot, BotTypeChoice.subject_specific_bot]:
                                    extracted_doc[file.name] = doc_content
                                    doc_content = get_document_summary(doc_content)
                            extracted_from_doc[file.name] = doc_content

                        signature_bot.refresh_from_db()
                        bot_media_data = signature_bot.data['media_data']
                        if is_overwrite and extracted_from_doc:
                            bot_media_data['extracted_from_doc'] = extracted_from_doc
                        else:
                            prev_extracted_from_doc = bot_media_data.get('extracted_from_doc',{})
                            # if "doc_files" in deleted_data:
                            #     for link in deleted_data["doc_files"].strip().split(","):
                            #         prev_extracted_from_doc.pop(link.strip(),None)
                            bot_media_data['extracted_from_doc'] = {**prev_extracted_from_doc,**extracted_from_doc}
                        
                        signature_bot.data['media_data'] = bot_media_data
                        signature_bot.save(update_fields=["data"])

                        bot_att.refresh_from_db()
                        bot_media_data = bot_att.extracted_documents if bot_att.extracted_documents else {}
                        bot_media_data['extracted_from_doc'] = {**bot_media_data.get('extracted_from_doc',{}),**extracted_doc}
                        bot_att.extracted_documents = bot_media_data
                        bot_att.save(update_fields=["extracted_documents"])
                        
                    if (media_data and 'attatched_pdfs' in media_data) or 'attatched_pdfs' in request.FILES:
                        attatched_pdfs = media_data['attatched_pdfs'] if 'attatched_pdfs' in media_data else request.FILES.getlist('attatched_pdfs')
                        pdf_names = []
                        extracted_from_pdf = {}
                        extracted_pdf = {}

                        """ logger.info(f"******************* attatched_pdfs: {attatched_pdfs}") """

                        for file in attatched_pdfs:
                            # logger.info(f"file name : {file.name}, {file.read()}")
                            pdf_names.append(file.name)
                            
                            path = default_storage.save(file.name, ContentFile(file.read()))

                            pdf_content = extract_text_from_pdf(path)

                            default_storage.delete(path)
                            if signature_bot.bot_type in [BotTypeChoice.avatar_bot, BotTypeChoice.subject_specific_bot]:
                                    extracted_pdf[file.name] = pdf_content
                                    
                                    pdf_content = get_document_summary(pdf_content)
                            extracted_from_pdf[file.name] = pdf_content

                    
                        signature_bot.refresh_from_db()
                        bot_media_data = signature_bot.data['media_data']
                        if is_overwrite and extracted_from_pdf:
                            bot_media_data['extracted_from_pdf'] = extracted_from_pdf
                        else:
                            prev_extracted_from_pdf = bot_media_data.get('extracted_from_pdf',{})
                            # if "pdf_files" in deleted_data:
                            #     for link in deleted_data["pdf_files"].strip().split(","):
                            #         prev_extracted_from_pdf.pop(link.strip(),None)
                            bot_media_data['extracted_from_pdf'] = {**prev_extracted_from_pdf,**extracted_from_pdf}
                        
                        signature_bot.data['media_data'] = bot_media_data
                        signature_bot.save(update_fields=["data"])

                        bot_att.refresh_from_db()
                        bot_media_data = bot_att.extracted_documents if bot_att.extracted_documents else {}
                        bot_media_data['extracted_from_pdf'] = {**bot_media_data.get('extracted_from_pdf',{}),**extracted_pdf}
                        bot_att.extracted_documents = bot_media_data
                        bot_att.save(update_fields=["extracted_documents"])

                        logger.info(f"******************* extracted_from_pdf: {extracted_from_pdf}")

                    logger.info(f"######### extracted media data : {extracted_media_data}")

                return Response({"msg": "updated"}, status=status.HTTP_200_OK)
        except Exception as e:
            logger.exception(f"Got error while creating bot: {e}")
            error_msg = f"Got error while creating bot: {e}\n\n"
            error_msg += traceback.format_exc()
            """ send_slack_message({"module": "########### create_bot_by_details ###########", "error": str(e)}) """
            send_error_notification("create_bot_by_details",error_msg,{})
            return Response({"msg":f"Got error : {e}" },status=status.HTTP_400_BAD_REQUEST)

    @action(methods=['GET','POST'],detail=False, url_path="user-competency-details")
    def user_competency_details(self,request,*args, **kwargs):
        """
        Retrieves or updates the competency details of a user based on their user ID.

        Args:
            request (object): The HTTP request object.
            user_id (string): The ID of the user for whom the competency details are being retrieved or updated.
            method (string): The HTTP method used for the request ('GET' or 'POST').

        Returns:
            If the HTTP method is 'GET', the method returns a list containing the competency data of the user.
            If the HTTP method is 'POST', the method returns a response with a success message.
        """
        try:
            data = []
            user_id = request.query_params.get('user_id')
            cache_key = generate_cache_key('competency_data', user_id=user_id)
            if request.method == 'GET':

                # Try to get data from cache
                cached_data = get_cache(cache_key)
                if cached_data is None:
                    # Cache miss, perform the database query
                    try:
                        competency_data = UserAttribute.objects.get(user_id=user_id).competency_data
                        if competency_data:
                            data.append(competency_data)
                        # Set data in cache
                        set_cache(cache_key, data)
                    except UserAttribute.DoesNotExist:
                        logger.exception(f"UserAttribute for user_id {user_id} not found")
                        return Response({"error": "UserAttribute not found"}, status=status.HTTP_404_NOT_FOUND)
                else:
                    data = cached_data

                return Response(data, status=status.HTTP_200_OK)
            elif request.method == 'POST':
                skill_data = request.query_params.get('competency_skills')
                skill_data = {str(i+1): value for i, value in enumerate(skill_data.split(','))}
                user_att = UserAttribute.objects.get(user_id=user_id)
                user_att.competency_data = skill_data
                user_att.save(update_fields=["competency_data"])

                # set_cache(cache_key, [skill_data])
                reset_cache_with_prefix('competency_data')
                reset_cache_with_prefix('tests_by_competency')
                return Response({"msg":"saved"},status=status.HTTP_200_OK)
        
        except Exception as e:
            return Response({"error": f"got error {e}"},status=status.HTTP_400_BAD_REQUEST)
        

    @action(methods=['GET','POST','PATCH'],detail=False, url_path="get_or_create_idp")
    def get_or_create_idp(self,request,*args, **kwargs):
        """
        This method is used to get or create an Individual Development Plan (IDP) based on the HTTP method in the request.

        The method supports GET, POST, and PATCH HTTP methods. 

        For a GET request, it retrieves an existing IDP. The 'user_id' and 'idp_id' are expected to be provided as query parameters. 
        If 'idp_id' is provided, it tries to fetch the IDP with that ID. If not, it fetches the IDP for the provided 'user_id'. 

        For a POST request, it creates a new IDP. The 'user_id' and 'idp_data' are expected to be provided in the request data. 
        The 'idp_data' should contain the necessary information to create the IDP.

        For a PATCH request, it regenerates the IDP or scenarios. The 'idp_id' is expected to be provided in the request data.

        Args:
            request (HttpRequest): The HTTP request object. Depending on the HTTP method, it should contain:
                - GET: 'user_id' and optionally 'idp_id' as query parameters.
                - POST: 'user_id' and 'idp_data' in the request data.
                - PATCH: 'idp_id' in the request data.
            *args: Variable length argument list.
            **kwargs: Arbitrary keyword arguments.

        Returns:
            HttpResponse: The HTTP response object containing the IDP data and HTTP status. 
            If the operation is successful, it returns the IDP data with HTTP status 200. 
            If not, it returns an error message with HTTP status 404 for GET and POST, and 400 for exceptions.

        Example:
            For a GET request:
            Request: GET /get_or_create_idp?user_id=123
            Response: {
                "idp_data": {...},
                "status": "completed"
            }, status=200
        """
        try:
            access_token = request.headers.get('Authorization')

            if request.method == 'GET':
                # user_id = request.data.get('user_id',None)
                user_id = request.query_params.get('user_id',None)
                idp_id = request.query_params.get('idp_id',None)
                cache_key = generate_cache_key('process_idp', user_id=user_id, tenant_id=request.tenant.uid, idp_id=idp_id)

                # get data from cache
                cached_data = get_cache(cache_key)
                if cached_data:
                    return Response(cached_data, status=status.HTTP_200_OK)

                if idp_id is not None:
                    data, success = process_idp("",user_id,request.tenant.uid,access_token,only_data=True,idp_id=idp_id)
                    if not success:
                        return Response(data,status=status.HTTP_404_NOT_FOUND)
                    
                    set_cache(cache_key, data, 60*5)
                    return Response(data,status=status.HTTP_200_OK)


                data, success = process_idp("",user_id,request.tenant.uid,access_token,only_data=True)
                if not success:
                    return Response(data,status=status.HTTP_404_NOT_FOUND)
                
                set_cache(cache_key, data, 60*5)
                return Response(data,status=status.HTTP_200_OK)
            
            elif request.method == 'POST':
                user_id = request.data.get('user_id',None)
                idp_data = request.data.get('idp_data',None)
                logger.info(f"****************** idp_data: {idp_data}")
                data, success = process_idp(idp_data,user_id,request.tenant.uid,access_token)
                if not success:
                    return Response(data,status=status.HTTP_404_NOT_FOUND)
                
                # reset_cache_with_prefix("process_idp")
                return Response(data,status=status.HTTP_200_OK)
            
            elif request.method == 'PATCH':
                idp_id = request.data.get('idp_id',None)
                logger.info(f"****************** idp_id: {idp_id}")
                data, success = regenerate_idp_or_scenarios(idp_id=idp_id,access_token=access_token,tenant_id=request.tenant.uid)
                return Response(data,status=status.HTTP_200_OK)
        except Exception as e:
            logger.exception(f"got error in get_or_create_idp: {e}")
            # send_slack_message({"module": "##################get_or_create_idp#################", "message": f"got error in get_or_create_idp: {e}"})
            error_msg = f"got error in get_or_create_idp: {e}\n\n"
            error_msg += traceback.format_exc()
            send_error_notification("get_or_create_idp",error_msg,request.data)
            return Response({"msg": "got_error"},status=status.HTTP_400_BAD_REQUEST)
          
          
    @action(methods=['GET'],detail=False, url_path="get-directory-informations")
    def get_directory_informations(self,request,*args, **kwargs):
        """
        Retrieves directory information based on the provided email in the request query parameters.

        This method first checks if an email is provided in the request query parameters. If an email is provided, it fetches the client information associated with the email. 
        It then retrieves all identities associated with the emails of the client and fetches the profiles associated with these identities. 
        If the client has accessed any bots, it also fetches the profiles associated with these bots. 
        Finally, it retrieves all directory information that is visible, approved, and associated with these profiles.

        If no email is provided, it simply retrieves all directory information that is visible and approved.

        Args:
            request (HttpRequest): The HTTP request object. The request query parameters may contain an 'email' key with the email as the value.

        Returns:
            HttpResponse: The HTTP response object containing the serialized directory information data. 
            The data is a list of dictionaries where each dictionary represents a directory information. 
            Each dictionary contains keys like 'id', 'name', 'description', etc. representing the directory information attributes.

        Raises:
            HTTP_400_BAD_REQUEST: If there is any error during the process, an error message is returned with a status of 400.

        Example:
            Input: 
                request.query_params = {'email': 'test@example.com'}
            Output: 
                [
                    {
                        'id': 1,
                        'name': 'Directory 1',
                        'description': 'This is directory 1',
                        ...
                    },
                    {
                        'id': 2,
                        'name': 'Directory 2',
                        'description': 'This is directory 2',
                        ...
                    },
                    ...
                ]
        """
        try:
            start_time = time.time()
            email = request.query_params.get('email')
            logger.info(f"Retrieving directory information for email: {email}")
            if email:
                cache_key = f"user-directory-info-{email}"
                start = time.time()
                data = get_cache(cache_key)
                end = time.time()
                logger.info(f"<<< Time taken to get cache_data : {end-start} >>>")
                if data:
                    end_time = time.time()
                    logger.info(f"<<< Time taken to get directory information when CacheHit : {end_time-start_time} >>>")
                    return Response(data, status=status.HTTP_200_OK)
                
                client = ClientUserInfo.objects.get(deleted=0,tenant_id=request.tenant.uid,member_emails__icontains=email)
                emails = client.member_emails.split(',') if client.member_emails else []
                emails = [email.strip() for email in emails]
                user_ids = Identity.objects.filter(deleted=False,tenant_id=request.tenant.uid,value__in = emails)
                user_ids_list = list(user_ids.values_list('user_id', flat=True))
                profile_ids = list(CoachCoacheeMentorMenteeProfile.objects.filter(deleted=0,tenant_id=request.tenant.uid,user_id__in=user_ids_list).values_list("uid",flat=True))
                bot_ids = []
                if client.accessed_bot_ids:
                    bot_ids = [bot_id.strip() for bot_id in client.accessed_bot_ids.split(',')]
                logger.info(f"accessed_bots: {bot_ids}, profile_ids: {profile_ids}")

                directories = DirectoryPageInfo.objects.filter(is_visible=True,is_approved=True,
                                                                
                                                            ).filter(Q(profile_id__in=profile_ids) | Q(avatar_bot_id__in = bot_ids))
                
                serializer = DirectoryInfoSErializer(directories,many=True)
                set_cache(cache_key, serializer.data)
            else:
                cache_key = f"user-directory-info-all"
                start = time.time()
                data = get_cache(cache_key)
                end = time.time()
                logger.info(f"<<< Time taken to get cache_data : {end-start} >>>")


                if data:
                    end_time = time.time()
                    logger.info(f"<<< Time taken to get directory information when CacheHit : {end_time-start_time} >>>")
                    return Response(data, status=status.HTTP_200_OK)
                
                logger.info("<<<! Cache Not found in memcached !>>>")
                
                directories = DirectoryPageInfo.objects.filter(is_visible=True,is_approved=True)
                serializer = DirectoryInfoSErializer(directories,many=True)
                set_cache(cache_key,serializer.data)
                
            end_time = time.time()
            logger.info(f"<<< Time taken to get directory information : {end_time-start_time} >>>")
                
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            logger.exception({"got error in directory information api": e})
            # send_slack_message({"module": "##################get_directory_informations#################", "message": f"got error in get_directory_informations: {e}"})
            # send_error_notification("get_directory_informations",f"got error in get_directory_informations: {e}",{"email": request.query_params.get('email')})
            return Response({"error": f"got error {e}"},status=status.HTTP_400_BAD_REQUEST)
        

    @action(methods=['GET'],detail=False,url_path="participant-leader-board-report")
    def participant_leader_board_report(self,request, *args, **kwargs):
        """
        Retrieves a leaderboard report for participants based on their activities.

        This method fetches the client information based on the provided email in the request query parameters.
        It then retrieves all the users associated with the client and fetches their action information.
        The method constructs a report for each user, including their name, user_id, counts of different bot interactions,
        total simulations, session notes count, and profile type. If a user has a 'coach' profile type, it is updated in the report.
        The total score is calculated as the sum of total bots, session notes count, total simulations, and total bot interactions.
        If a user does not have any action information, a report is still created for them with all counts as zero.
        Finally, the method sorts the data based on the total score in descending order and assigns a rating to each user based on their position in the sorted list.

        Args:
            request (object): The HTTP request object. The request query parameters should include an 'email' field.

        Returns:
            Response: A list of dictionaries, each containing the following fields:
                - name: The display name of the user.
                - user_id: The unique identifier of the user.
                - avatar_bot_count: The count of avatar bot interactions.
                - subject_matter_count: The count of subject matter bot interactions.
                - total_bots: The total count of bot interactions.
                - total_simulations: The total count of simulations attempted.
                - total_bot_interactions: The total count of bot interactions attempted.
                - session_notes_count: The count of session notes.
                - profile_type: The profile type of the user, either 'coachee' or 'coach'.
                - total_score: The total score calculated as the sum of total bots, session notes count, total simulations, and total bot interactions.
                - rating: The rating of the user based on their total score.

        Example:
            [
                {
                    "name": "John Doe",
                    "user_id": "123",
                    "avatar_bot_count": 5,
                    "subject_matter_count": 3,
                    "total_bots": 8,
                    "total_simulations": 2,
                    "total_bot_interactions": 10,
                    "session_notes_count": 4,
                    "profile_type": "coach",
                    "total_score": 24,
                    "rating": 1
                },
                ...
            ]
        """
        try:
            if request.method == "GET":
                email = request.query_params.get('email')
                by_category = request.query_params.get('by_category')
                
                cache_key = generate_cache_key('participant_leader_board_report', email=email, by_category=by_category, tenant_id=request.tenant.uid)

                # Try to get data from cache
                data = get_cache(cache_key)
                if data is not None:
                    return Response(data, status=status.HTTP_200_OK)
                
                client = ClientUserInfo.objects.get(deleted=0,tenant_id=request.tenant.uid,member_emails__icontains=email)
                emails = client.member_emails.split(',') if client.member_emails else []
                emails = [email.strip() for email in emails]
                excluded_emails = client.excluded_users.split(',') if client.excluded_users else []
                excluded_emails = [email.strip() for email in excluded_emails]
                emails = set(emails) - set(excluded_emails)
                
                user_ids = Identity.objects.filter(deleted=False,tenant_id=request.tenant.uid,value__in = emails)
                user_ids_list = list(user_ids.values_list('user_id', flat=True))
                user_actions = UserActionInfo.objects.filter(deleted=False,tenant_id=request.tenant.uid,user_id__in=user_ids_list)
                data = []
                for user_action in user_actions:
                    user = get_user_by_id(user_action.user_id)
                    avatar_bot_count = len(get_list_from_string(user_action.avatar_ids))
                    subject_specific_bot_count = len(get_list_from_string(user_action.avatar_ids))
                    subject_matter_count = len(get_list_from_string(user_action.subject_matter_bot_ids))
                    knowledge_bot_count = len(get_list_from_string(user_action.knowledge_bot_ids))
                    deep_dive_bot_count = len(get_list_from_string(user_action.deep_dive_bot_ids))
                    total_bots = avatar_bot_count + subject_matter_count + knowledge_bot_count + deep_dive_bot_count + subject_specific_bot_count
                    temp = {
                        "name": get_user_display_name(user),
                        "user_id": user.uid,
                        "avatar_bot_count": avatar_bot_count,
                        "subject_specific_bot_count": subject_specific_bot_count,
                        "subject_matter_count": subject_matter_count,
                        "knowledge_bot_count": knowledge_bot_count,
                        "deep_dive_bot_count": deep_dive_bot_count,
                        "total_bots": total_bots,
                        "total_simulations": user_action.interaction_attempted,
                        "total_bot_interactions": user_action.chat_attempted,
                        "session_notes_count": user_action.session_notes_count,
                        "profile_type": "coachee"
                    }
                    profiles = CoachCoacheeMentorMenteeProfile.objects.filter(deleted=False,is_approved=True,tenant_id=request.tenant.uid,user_id=user.uid)
                    for p in profiles:
                        temp['profile_type'] = p.profile_type
                        temp['is_mentor'] = p.is_mentor
                        temp['created'] = p.created


                    temp['total_score'] = temp['total_bots'] + temp['session_notes_count'] + temp['total_simulations'] + temp['total_bot_interactions']
                    data.append(temp)

                existing_user_ids = set(user_action.user_id for user_action in user_actions)
                # Iterate through user_ids and include those not present in user_actions
                for user_id in user_ids_list:
                    if user_id not in existing_user_ids:
                        user = get_user_by_id(user_id)
                        # profile = CoachCoacheeMentorMenteeProfile.objects.filter(deleted=False,tenant_id=request.tenant.uid,user_id=user_id).first()
                        temp = {
                            "name": get_user_display_name(user),
                            "user_id": user.uid,
                            "avatar_bot_count": 0, 
                            "subject_specific_bot_count": 0,
                            "subject_matter_count": 0,
                            "total_bots": 0,
                            "total_simulations": 0,
                            "total_bot_interactions": 0,
                            "session_notes_count": 0,
                            "profile_type": 'coachee',
                            'total_score': 0
                        }
                        profiles = CoachCoacheeMentorMenteeProfile.objects.filter(deleted=False,is_approved=True, tenant_id=request.tenant.uid, user_id=user.uid)
                        for p in profiles:
                            temp['profile_type'] = p.profile_type
                            temp['is_mentor'] = p.is_mentor
                            temp['created'] = p.created

                        data.append(temp)
                    
                # data = sorted(data, key=lambda x: x['total_score'], reverse=True)
                coach = []
                coachee = []
                if by_category:
                    coach = [copy.deepcopy(x) for x in data if x['profile_type'] in ['coach', 'mentor']]
                    coachee = [copy.deepcopy(x) for x in data if x['profile_type'] in ['coachee', 'mentee']]

                    coach = custom_sort_reverse(data=coach,first_sort_field="total_score",second_sort_field="name")
                    for i, item in enumerate(coach, start=1):
                        item['rating'] = i

                    coachee = custom_sort_reverse(data=coachee,first_sort_field="total_score",second_sort_field="name")
                    for i, item in enumerate(coachee, start=1):
                        item['rating'] = i

                    
                    data = custom_sort_reverse(data=data,first_sort_field="total_score",second_sort_field="name")
                    for i, item in enumerate(data, start=1):
                        item['rating'] = i
                        
                    set_cache(cache_key, {"coach_mentor": coach, 'coachee_mentee': coachee, 'full_data': data})
            
                    return Response({"coach_mentor": coach,'coachee_mentee': coachee, 'full_data': data},status=status.HTTP_200_OK)


                data = custom_sort_reverse(data=data,first_sort_field="total_score",second_sort_field="name")
                for i, item in enumerate(data, start=1):
                    item['rating'] = i
                    
                set_cache(cache_key, data)

                return Response(data, status=status.HTTP_200_OK)
        except Exception as e:
            # send_slack_message({"module": "##################participant_leader_board_report#################", "message": f"got error in participant_leader_board_report: {e}"})
            send_error_notification("participant_leader_board_report",f"got error in participant_leader_board_report: {e}",{})
            return Response({"error": f"{e}"}, status=status.HTTP_400_BAD_REQUEST)


    @action(methods=['GET','POST','PATCH'],detail=False, url_path="coach-coachee-connections")
    def coach_coachee_connections(self, request, *args, **kwargs):
        """
        This method handles GET, POST, and PATCH requests for coach-coachee connections.

        For GET requests, it retrieves coach-coachee connections based on the provided query parameters: 'connection_id', 'user_id', 'coach_id', and 'coachee_id'. 
        If no parameters are provided, it returns all connections for the current tenant.

        For POST requests, it creates a new coach-coachee connection. It requires 'coach_id', 'coachee_id', and 'profile_page_url' in the request data. 
        After creating the connection, it sends an email to the coach with a connection request.

        For PATCH requests, it updates an existing coach-coachee connection. It requires 'connection_id' or both 'coach_id' and 'coachee_id' in the request data. 
        It also updates the connection status and sends an email to the coachee if the status is 'accepted'.

        Input:
            request: A Django HttpRequest object. Contains metadata about the request.
                For GET, it may contain query parameters 'connection_id', 'user_id', 'coach_id', 'coachee_id'.
                For POST, it must contain 'coach_id', 'coachee_id', and 'profile_page_url' in request.data.
                For PATCH, it must contain 'connection_id' or both 'coach_id' and 'coachee_id' in request.data.

        Output:
            A Django HttpResponse object. Contains the serialized data of the coach-coachee connections or an error message.

        Example:
            GET /coach-coachee-connections?coach_id=1
            Response: {"data": [{"connection_id": 1, "coach_id": 1, "coachee_id": 2, "status": "accepted"}]}

            POST /coach-coachee-connections
            Request data: {"coach_id": 1, "coachee_id": 2, "profile_page_url": "http://example.com/profile"}
            Response: {"data": {"connection_id": 1, "coach_id": 1, "coachee_id": 2, "status": "pending"}}

            PATCH /coach-coachee-connections
            Request data: {"connection_id": 1, "status": "accepted"}
            Response: {"data": {"connection_id": 1, "coach_id": 1, "coachee_id": 2, "status": "accepted"}}
        """
        if request.method == 'GET':
            connection_id = request.query_params.get('connection_id',None)
            user_id = request.query_params.get('user_id',None)
            coach_id = request.query_params.get('coach_id',None)
            coachee_id = request.query_params.get('coachee_id',None)
            logger.info(f"**************** connection_id: {connection_id},  coach_id: {coach_id},  coachee_id: {coachee_id}, tenant_id: {self.request.tenant.uid}")
            
            cache_key = generate_cache_key(
            'connection', 
            tenant_id=self.request.tenant.uid,
            connection_id=connection_id,
            user_id=user_id,
            coach_id=coach_id,
            coachee_id=coachee_id
            )

            # get data from cache
            data = get_cache(cache_key)
            if data is None:

                if coach_id:
                    try:
                        connection = CoachCoacheeConnection.objects.filter(deleted=False, coach_id=coach_id, tenant_id=self.request.tenant.uid)
                        data = {"data": CoachCoacheeConnectionSerializer(connection, many=True).data}
                        logger.info(f"#########################  coach connection: {connection}")
                    except Exception as e:
                        logger.exception(e)
                        data = {"error": "connection not found"}
                        return Response(data, status=status.HTTP_404_NOT_FOUND)

                elif connection_id:
                    try:
                        connection = CoachCoacheeConnection.objects.get(deleted=False, uid=connection_id, tenant_id=self.request.tenant.uid)
                        data = {"data": CoachCoacheeConnectionSerializer(connection).data}
                        logger.info(f"#########################  connections: {connection}")
                    except Exception as e:
                        logger.exception(e)
                        data = {"error": "connection not found"}
                        return Response(data, status=status.HTTP_404_NOT_FOUND)
                    
                elif coachee_id:
                    try:
                        connection = CoachCoacheeConnection.objects.filter(deleted=False, coachee_id=coachee_id, tenant_id=self.request.tenant.uid)
                        data = {"data": CoachCoacheeConnectionSerializer(connection, many=True).data}
                        logger.info(f"#########################  coachee connection: {connection}")
                    except Exception as e:
                        logger.exception(e)
                        data = {"error": "connection not found"}
                        return Response(data, status=status.HTTP_404_NOT_FOUND)
                
                else:
                    try:
                        email = request.query_params.get('email')
                        client = ClientUserInfo.objects.get(deleted=0, tenant_id=request.tenant.uid, member_emails__icontains=email)
                        emails = client.member_emails.split(',') if client.member_emails else []
                        emails = [email.strip() for email in emails]
                        excluded_emails = client.excluded_users.split(',') if client.excluded_users else []
                        excluded_emails = [email.strip() for email in excluded_emails]
                        emails = set(emails) - set(excluded_emails)
                        user_ids = Identity.objects.filter(deleted=False, tenant_id=request.tenant.uid, value__in=emails)
                        user_ids_list = list(user_ids.values_list('user_id', flat=True))
                        profile_ids = list(CoachCoacheeMentorMenteeProfile.objects.filter(deleted=False, tenant_id=request.tenant.uid, user_id__in=user_ids_list).values_list('uid', flat=True))
                        logger.info(f"profile_ids: {profile_ids}, user_ids: {user_ids_list}")

                        connections = CoachCoacheeConnection.objects.filter(
                            Q(coach_id__in=profile_ids) | Q(coachee_id__in=profile_ids),
                            deleted=False,
                            tenant_id=self.request.tenant.uid,
                        )
                        data = {"data": CoachCoacheeConnectionSerializer(connections, many=True).data}
                    except Exception as e:
                        logger.exception(e)
                        data = {"error": "connections not found"}
                        return Response(data, status=status.HTTP_404_NOT_FOUND)

                # Set data in cache
                set_cache(cache_key, data)

            return Response(data, status=status.HTTP_200_OK)

        if request.method == 'PATCH':
            connection_id = request.data.get('connection_id',None)
            coach_id = request.data.get('coach_id',None)
            coachee_id = request.data.get('coachee_id',None)

            if connection_id:
                connection = CoachCoacheeConnection.objects.get(deleted=False,uid=connection_id)
                
                data = request.data.copy()
                coach = CoachCoacheeMentorMenteeProfile.objects.get(deleted=False,tenant_id=request.tenant.uid,uid=connection.coach_id)
                coachee = CoachCoacheeMentorMenteeProfile.objects.get(deleted=False,tenant_id=request.tenant.uid,uid=connection.coachee_id)
                
                coach_name = coach.name
                coachee_name = coachee.name
                coachee_email = coachee.email
                # serializer = CoachCoacheeConnectionSerializer(connection,data=data,partial=True)
                # serializer.is_valid(raise_exception=True)
                # serializer.save()

                if data.get('status') == CoachCoacheeConnectionStatusChoice.accepted:
                    connection.status = CoachCoacheeConnectionStatusChoice.accepted
                    connection.save(update_fields=['status'])
                    subject = "Connection Approved"
                    html = f"""
                        <p style="font-family: sans-serif; font-size: 14px; font-weight: normal; margin: 0; margin-bottom: 15px;"> Congratuations, {coach_name} has approved your connection request.</p>
                        <p style="font-family: sans-serif; font-size: 14px; font-weight: normal; margin: 0; margin-bottom: 15px;">You may have refresh the system for the changes to reflect.</p>
                        """

                    reset_cache_with_prefix('get_bots')
                    reset_cache_with_prefix('connection')
                    send_email_with_html_template(subject=subject,html_content=html,to_email=coachee_email)
                    return Response({"data": CoachCoacheeConnectionSerializer(connection).data },status=status.HTTP_200_OK)
                else:
                    connection.status = CoachCoacheeConnectionStatusChoice.rejected
                    connection.save(update_fields=['status'])
                    reset_cache_with_prefix('get_bots')
                    reset_cache_with_prefix('connection')
                    return Response({"data": CoachCoacheeConnectionSerializer(connection).data },status=status.HTTP_200_OK)
            
            if coach_id and coachee_id:
                connection = CoachCoacheeConnection.objects.get(deleted=False,coach_id=coach_id,coachee_id=coachee_id)
                data = request.data.copy()
                # serializer = CoachCoacheeConnectionSerializer(connection,data=data,partial=True)
                # serializer.is_valid(raise_exception=True)
                # serializer.save()

                if data.get('status') == CoachCoacheeConnectionStatusChoice.accepted:
                    connection.status = CoachCoacheeConnectionStatusChoice.accepted
                    connection.save(update_fields=['status'])
                    reset_cache_with_prefix('get_bots')
                    reset_cache_with_prefix('connection')
                    coach = CoachCoacheeMentorMenteeProfile.objects.get(deleted=False,tenant_id=request.tenant.uid,uid=coach_id)
                    coachee = CoachCoacheeMentorMenteeProfile.objects.get(deleted=False,tenant_id=request.tenant.uid,uid=coachee_id)
                    
                    coach_name = coach.name
                    coachee_name = coachee.name
                    coachee_email = coachee.email
                    subject = "Connection Approved"
                    html = f"""
                        <p style="font-family: sans-serif; font-size: 14px; font-weight: normal; margin: 0; margin-bottom: 15px;"> Congratuations, {coach_name} has approved your connection request.</p>
                        <p style="font-family: sans-serif; font-size: 14px; font-weight: normal; margin: 0; margin-bottom: 15px;"> You may have refresh the system for the changes to reflect.</p>
                        """

                    send_email_with_html_template(subject=subject,html_content=html,to_email=coachee_email)
                    # reset_cache_with_prefix("connection")
                    return Response({"data": CoachCoacheeConnectionSerializer(connection).data },status=status.HTTP_200_OK)
                else:
                    connection.status = CoachCoacheeConnectionStatusChoice.rejected
                    connection.save(update_fields=['status'])
                    reset_cache_with_prefix('get_bots')
                    reset_cache_with_prefix('connection')
                    return Response({"data": CoachCoacheeConnectionSerializer(connection).data },status=status.HTTP_200_OK)

        if request.method == 'POST':
            
            coach_id = request.data.get('coach_id',None)
            coachee_id = request.data.get('coachee_id',None)
            profile_url = request.data.get('profile_page_url')
            try:
                logger.info(f"***************** request data: {request.data}")

                if None in [coach_id,coachee_id]:
                    return Response({"error":"coach_id and coachee_id are required"},status=status.HTTP_400_BAD_REQUEST)
                
                try:
                    coach = CoachCoacheeMentorMenteeProfile.objects.get(deleted=False, uid=coach_id)

                    avatar_bot_id = None
                    bots = SignatureBot.objects.filter(user_id=coach.user_id, 
                                                       bot_type__in=[BotTypeChoice.avatar_bot, BotTypeChoice.subject_specific_bot], 
                                                       deleted=False)
                    if bots.count() > 0:
                        avatar_bot_id = bots.first().uid

                except Exception as e:
                    logger.exception(e)
                    return Response({"error":"coach not found"},status=status.HTTP_404_NOT_FOUND)

                data = request.data.copy()
                data['tenant_id'] = self.request.tenant.uid
                data['coach_avatar_bot_id'] = avatar_bot_id

                existing_connection = CoachCoacheeConnection.objects.filter(
                    tenant_id=data['tenant_id'],
                    coach_id=coach_id,
                    coachee_id=coachee_id
                ).first()

                if existing_connection:
                    # Handle the existing record case (e.g., return an error or update the record)
                    return Response({"error":'A connection with the same tenant_id, coach_id, and coachee_id already exists.'}, status=status.HTTP_400_BAD_REQUEST)

                serializer = CoachCoacheeConnectionSerializer(data=data)
                serializer.is_valid(raise_exception=True)
                created_connection = serializer.save()
                reset_cache_with_prefix('get_bots')
                reset_cache_with_prefix('connection')
                coach = CoachCoacheeMentorMenteeProfile.objects.get(deleted=False,tenant_id=request.tenant.uid,uid=coach_id)
                coachee = CoachCoacheeMentorMenteeProfile.objects.get(deleted=False,tenant_id=request.tenant.uid,uid=coachee_id)
                
                coach_name = coach.name
                coachee_name = coachee.name
                coachee_email = coachee.email
                subject = "You have a connection request"
                html = f"""
                    <p style="font-family: sans-serif; font-size: 14px; font-weight: normal; margin: 0; margin-bottom: 15px;">You have got a connection request from <b>{coachee_name} - {coachee_email}</b>, please log in to your account to approve or reject. Thank you!</p>
                    <p style="font-family: sans-serif; font-size: 14px; font-weight: normal; margin: 0; margin-bottom: 15px;">You may have refresh the system for the changes to reflect.</p>
                    """

                send_email_with_html_template(subject=subject,html_content=html,to_email=coach.email,title=f'Hey {coach_name}!')
                return Response({"data": CoachCoacheeConnectionSerializer(created_connection).data },status=status.HTTP_201_CREATED)
            except Exception as e:
                logger.exception(e)
                send_error_notification("coach_coachee_connections",f"got error in coach_coachee_connections: {e}",{"data": request.data})
                return Response({"error":f"got error {e}"},status=status.HTTP_400_BAD_REQUEST)
            
            
    @action(methods=['GET'],detail=False,url_path="feedback-leaderboard-report")
    def feedback_leader_board(self,request, *args, **kwargs):
        """
        Generates a feedback leaderboard report for a given client.

        This method retrieves the client's information using the email provided in the request's query parameters.
        It then fetches all the identities associated with the client's member emails and retrieves all the feedback 
        provided by these users from the BotQnA model. The feedback is grouped by bot_id and for each bot, the number 
        of positive and negative feedback is calculated. The bot's name and owner's name are also retrieved. All this 
        information is then formatted and appended to a list. The list is sorted in descending order based on the number 
        of positive feedback. Finally, a rating is assigned to each bot based on its position in the sorted list.

        Args:
            request (HttpRequest): The HTTP request object. The request's query parameters should contain an 'email' field 
            which is the email of the client for whom the report is to be generated.

        Returns:
            HttpResponse: The HTTP response object containing the feedback leaderboard report. The report is a list of 
            dictionaries where each dictionary contains the following keys:
                - 'bot_name': The name of the bot.
                - 'user_id': The user id of the bot's owner.
                - 'owner_name': The display name of the bot's owner.
                - 'positive_feedback_count': The number of positive feedback for the bot.
                - 'negative_feedback_count': The number of negative feedback for the bot.
                - 'rating': The bot's rating based on the number of positive feedback. The bot with the highest number 
                of positive feedback has a rating of 1.

        Example:
            Request: GET /feedback-leaderboard-report?email=client@example.com
            Response: 
            {
                "group": [
                    {
                        "bot_name": "Bot 1",
                        "user_id": 1,
                        "owner_name": "Owner 1",
                        "positive_feedback_count": 10,
                        "negative_feedback_count": 2,
                        "rating": 1
                    },
                    {
                        "bot_name": "Bot 2",
                        "user_id": 2,
                        "owner_name": "Owner 2",
                        "positive_feedback_count": 8,
                        "negative_feedback_count": 3,
                        "rating": 2
                    },
                    ...
                ]
            }
        """
        try:
            if request.method == "GET":
                email = request.query_params.get('email')
                cache_key = generate_cache_key("feedback-leaderboard",tenant_id=request.tenant.uid, email=email)
                cached_data = get_cache(cache_key)
                
                if cached_data:
                    return Response({'group': cached_data},status=status.HTTP_200_OK)
                
                client = ClientUserInfo.objects.get(deleted=0,tenant_id=request.tenant.uid,member_emails__icontains=email)
                emails = client.member_emails.split(',') if client.member_emails else []
                emails = [email.strip() for email in emails]
                excluded_emails = client.excluded_users.split(',') if client.excluded_users else []
                excluded_emails = [email.strip() for email in excluded_emails]
                emails = set(emails) - set(excluded_emails)
                user_ids = Identity.objects.filter(deleted=False,tenant_id=request.tenant.uid,value__in = emails)
                user_ids_list = list(user_ids.values_list('user_id', flat=True))
                feedback_bots = SignatureBot.objects.filter(deleted=False,is_approved=True,tenant_id=request.tenant.uid,user_id__in=user_ids_list,bot_type=BotTypeChoice.feedback_bot)
                # bot_qnas = BotQnA.objects.filter(deleted=False,tenant_id=request.tenant.uid,qna_type='feedback',participant_id__in=user_ids_list)

                # Sort the queryset by bot_id
                # bot_qnas = sorted(bot_qnas, key=lambda x: x.bot_id)

                # Group the sorted queryset by bot_id
                # grouped_bot_qnas = groupby(bot_qnas, key=attrgetter('bot_id'))
                formatted_data = []
                for bot in feedback_bots:
                    bot_id = bot.uid
                    bot_qnas = BotQnA.objects.filter(deleted=False,tenant_id=request.tenant.uid,qna_type='feedback',bot_id=bot_id)
                    positive_count = sum(1 for item in bot_qnas if item.is_positive)
                    negative_count = sum(1 for item in bot_qnas if not item.is_positive)
                    bot_name = BotAttribute.objects.get(bot_id=bot_id).bot_name
                    owner_name = get_user_display_name(get_user_by_id(bot.user_id))

                    formatted_entry = {
                        "bot_name": bot_name,
                        "user_id": bot.user_id,
                        "owner_name": owner_name,
                        "positive_feedback_count": positive_count,
                        "negative_feedback_count": negative_count
                    }
                    formatted_data.append(formatted_entry)
                    
                    
                # for bot_id, group in grouped_bot_qnas:
                #     positive_count = sum(1 for item in group if item.is_positive)
                #     negative_count = sum(1 for item in group if not item.is_positive)
                #     signature_bot = SignatureBot.objects.get(uid=bot_id)
                #     bot_name = BotAttribute.objects.get(bot_id=bot_id).bot_name
                #     owner_name = get_user_display_name(get_user_by_id(signature_bot.user_id))

                #     formatted_entry = {
                #         "bot_name": bot_name,
                #         "user_id": signature_bot.user_id,
                #         "owner_name": owner_name,
                #         "positive_feedback_count": positive_count,
                #         "negative_feedback_count": negative_count
                #     }
                #     formatted_data.append(formatted_entry)


                # formatted_data = sorted(formatted_data, key=lambda x: x["positive_feedback_count"], reverse=True)
                formatted_data = custom_sort_reverse(data=formatted_data,first_sort_field="positive_feedback_count",second_sort_field="owner_name")
                for rating, item in enumerate(formatted_data,start=1):
                    item['rating'] = rating
                set_cache(cache_key,formatted_data)

                return Response({'group': formatted_data},status=status.HTTP_200_OK)
        except Exception as e:
            logger.exception(e)
            # send_slack_message({"module": "##################feedback_leader_board#################", "message": f"got error in feedback_leader_board: {e}"})
            send_error_notification("feedback_leader_board",f"got error in feedback_leader_board: {e}",request.query_params)
            return  Response({"Error": f"Got Error in Feedback-leaderboard-report : {e}"},status=status.HTTP_400_BAD_REQUEST)



    @action(methods=['GET','POST'],detail=False,url_path="user-can-join-as")
    def user_can_join_as(self, request, *args, **kwargs):
        """
        Determines the role a user can join as (e.g., 'coach' or 'coachee') based on their email and tenant's settings, or processes a CSV file to bulk update user privileges.

        This method supports both GET and POST requests. For a GET request, it checks if the provided email belongs to a user with specific privileges within the tenant's domain and returns the role the user can join as. For a POST request, it accepts a CSV file containing user emails, client names, and desired roles, updating the privileges in bulk.

        GET Request:
        - Input: Requires a query parameter 'email' containing the user's email address.
        - Process: Checks if the email is associated with specific privileges (e.g., coach or mentor) within the tenant's settings. It queries the `ClientUserInfo` and `CoachCoacheeJoiningPreviledge` models to determine the role.
        - Output: Returns a JSON response with the key 'can_join_as' indicating the role the user can join as. Example: {"can_join_as": "coach"}

        POST Request:
        - Input: Requires a multipart/form-data request with a 'csv_file' field containing the CSV file. The CSV file should have headers 'email', 'client_name', and 'can_join_as'.
        - Process: Reads the CSV file and creates or updates `CoachCoacheeJoiningPreviledge` entries for each row in the file, setting the user's privilege based on the 'can_join_as' field.
        - Output: Returns a JSON response with a message indicating the number of records created and any errors encountered. Example: {"msg": "10 records created", "errors": ["got error for user@example.com: <error message>"]}

        Args:
            request (HttpRequest): The HTTP request object.
            *args: Variable length argument list.
            **kwargs: Arbitrary keyword arguments.

        Returns:
            HttpResponse: A JSON response indicating the role the user can join as for GET requests, or the outcome of the bulk update for POST requests.
        """
        if request.method == 'GET':
            logger.info(f"***************** tenant_id: {request.tenant.uid}")
            try:
                user_email = request.query_params.get('email',None)
                tenant_id = self.request.tenant.uid
                can_join_as = 'coachee'
                if user_email:
                    client = ClientUserInfo.objects.filter(deleted=False,tenant_id=tenant_id,member_emails__contains=user_email).first()
                    if client:
                        if client.is_coach_mentor_previledge:
                            coach_mentor_previledge = client.coach_mentor_previledge
                            if coach_mentor_previledge:
                                coach_mentor_previledge = [email.strip() for email in coach_mentor_previledge.split(",")]
                                if user_email in coach_mentor_previledge:
                                    can_join_as = 'coach'
                                else:
                                    previledges = CoachCoacheeJoiningPreviledge.objects.filter(deleted=False,tenant_id=request.tenant.uid)
                                    previledges = previledges.filter(email=user_email.strip().lower()).first()
                                    can_join_as = previledges.can_join_as
                            else:
                                previledges = CoachCoacheeJoiningPreviledge.objects.filter(deleted=False,tenant_id=request.tenant.uid)
                                previledges = previledges.filter(email=user_email.strip().lower()).first()
                                can_join_as = previledges.can_join_as
                        else:
                            previledges = CoachCoacheeJoiningPreviledge.objects.filter(deleted=False,tenant_id=request.tenant.uid)
                            previledges = previledges.filter(email=user_email.strip().lower()).first()
                            can_join_as = previledges.can_join_as
                    
                return Response({"can_join_as": can_join_as},status=status.HTTP_200_OK)
            except Exception as e:
                logger.exception({"error": f"got error in can-join-as {e}"})
                return Response({"error": f"got error {e}"},status=status.HTTP_400_BAD_REQUEST)
            
        if request.method == 'POST':
            csv_file = request.FILES.get('csv_file')
            logger.info(f"***************** csv_file: {csv_file}")
        
            csv_text = TextIOWrapper(csv_file, encoding='utf-8-sig')
            csv_reader = csv.DictReader(csv_text)

            all_rows = list(csv_reader)

            record_count = 0
            errors = []
            for row in all_rows:
                logger.info(f"***************** row: {row}")
                email, client_name, can_join_as = row['email'].strip().lower(), row['client_name'], row['can_join_as']
                try:
                    CoachCoacheeJoiningPreviledge.objects.create(email=email,client_name=client_name,
                                                                    can_join_as=can_join_as,
                                                                    tenant_id=request.tenant.uid)
                    record_count += 1
                except Exception as e:
                    logger.exception(e)
                    errors.append(f"got error for {email}: {e}")
                    continue


            return Response({"msg":f"{record_count} records created","errors":errors},status=status.HTTP_200_OK)
        


    @action(methods=['GET'],detail=False,url_path="user-bot-connection-status")
    def user_bot_connection_status(self, request, *args, **kwargs):
        """
    Retrieves the connection status between a user and a coach based on their respective user IDs.

    This method checks if there is an accepted connection between the specified user (coachee) and coach within the system. It uses the user IDs provided in the request's query parameters to look up their profiles and the connection status between them. If an accepted connection exists, it returns a positive response; otherwise, it indicates that no such connection is found.

    Args:
        request (HttpRequest): The HTTP request object containing the query parameters.
            - user_id (str): The user ID of the coachee.
            - coach_user_id (str): The user ID of the coach.

    Returns:
        HttpResponse: JSON response indicating whether a connection exists or not.
        - If a connection is found and is accepted, it returns:
            {"connected": True}
        - If no accepted connection exists or an error occurs, it returns:
            {"connected": False, "error": "<error message>"}

    Raises:
        HTTP 400 Bad Request: If either 'user_id' or 'coach_user_id' is not provided in the query parameters.
        HTTP 404 Not Found: If no connection is found or if the specified user or coach does not exist.

    Example:
        GET /user-bot-connection-status?user_id=123&coach_user_id=456
        Possible responses:
        - If connected: {"connected": True}
        - If not connected or error: {"connected": False, "error": "connection not found"}
    """
        user_id = request.query_params.get('user_id',None)
        coach_user_id = request.query_params.get('coach_user_id',None)
        logger.info(f"****************** user_id: {user_id}, coach_user_id: {coach_user_id}, tenant_id: {request.tenant.uid}")
        if None in [user_id,coach_user_id]:
            return Response({"error":"user_id and coach_user_id are required"},status=status.HTTP_400_BAD_REQUEST)
        try:
            coachee_profile = CoachCoacheeMentorMenteeProfile.objects.get(deleted=False,user_id=user_id, tenant_id=request.tenant.uid)
            coach_profile = CoachCoacheeMentorMenteeProfile.objects.get(deleted=False,user_id=coach_user_id, tenant_id=request.tenant.uid)
            connection = CoachCoacheeConnection.objects.get(deleted=False,coachee_id=coachee_profile.uid,coach_id=coach_profile.uid,
                                                            status=CoachCoacheeConnectionStatusChoice.accepted, tenant_id=request.tenant.uid)
            return Response({"connected":True},status=status.HTTP_200_OK)
        except Exception as e:
            logger.exception(e)
            send_error_notification("user_bot_connection_status",f"got error in user_bot_connection_status: {e}",request.query_params)
            return Response({"connected":False, "error":"connection not found"},status=status.HTTP_404_NOT_FOUND)
        


    @action(methods=['GET'], detail=False, url_path='get-skill-and-role-bots')
    def get_skill_and_role_bots(self, request, *args, **kwargs):
        """
        Retrieves a list of skill and role bots associated with the current tenant.

        This method filters the SignatureBot objects to find those that are not marked as deleted and belong to the current tenant. It specifically looks for bots that are categorized under 'role_bot' or 'skill_bot' as defined in the BotScenarioCaseChoice. For each bot found, it fetches the corresponding BotAttribute to gather additional details about the bot. The method compiles a list of dictionaries, each representing a bot with its name, description, ID, type, and scenario case.

        Args:
            request (HttpRequest): The HTTP request object, which must include the tenant ID in `request.tenant.uid`.
            *args: Variable length argument list.
            **kwargs: Arbitrary keyword arguments.

        Returns:
            Response: A DRF Response object containing a list of dictionaries, each representing a bot. The dictionary keys include 'bot_name', 'description', 'bot_id', 'bot_type', and 'scenario_case'. If successful, the response status is set to HTTP 200 OK. In case of an exception, it returns a response with an error message and HTTP 400 Bad Request status.

        Example of expected output:
            [
                {
                    "bot_name": "Leadership Coach",
                    "description": "Helps develop leadership skills",
                    "bot_id": "123",
                    "bot_type": "role_bot",
                    "scenario_case": "role_bot"
                },
                {
                    "bot_name": "Communication Skill Enhancer",
                    "description": "Improves communication skills",
                    "bot_id": "456",
                    "bot_type": "skill_bot",
                    "scenario_case": "skill_bot"
                }
            ]

        Raises:
            Exception: Captures any exceptions that occur during the process, logs the exception, and returns an error message with HTTP 400 Bad Request status.
        """
        try:
            bots = SignatureBot.objects.filter(
                deleted = False,
                tenant_id = request.tenant.uid,
                bot_type = BotTypeChoice.coachbots
            )
            data = []
            for bot in bots:
                att = BotAttribute.objects.get(bot_id=bot.uid)
                data.append(
                    {
                        "bot_name": att.bot_name,
                        "description": att.about,
                        "bot_id": bot.bot_id,
                        "bot_type": bot.bot_type,
                        "scenario_case": bot.bot_scenario_case
                    }
                )

            return Response(data, status=status.HTTP_200_OK)
        except Exception as e:
            logger.exception(f"Get skills and role bots failed with: {e}")
            send_error_notification("get_skill_and_role_bots",f"got error in get_skill_and_role_bots: {e}",request.query_params)
            return Response({"message": f"{e}"}, status=status.HTTP_400_BAD_REQUEST)


    @action(methods=['POST'], detail=False, url_path='save-liked-profile')
    def save_liked_profile(self, request, *args, **kwargs):
        """
        Saves or reverts a liked profile based on the user's action.

        This method allows a user to express admiration for a profile by adding their user ID to the profile's admirer list. If the user decides to revert their admiration, their user ID can be removed from the list. The process involves checking if the profile exists and then updating the `admirer_user_ids` field of the `CoachCoacheeMentorMenteeProfile` model based on the action specified in the request.

        Args:
            request (HttpRequest): The HTTP request object containing the data needed to process the action. Expected keys in `request.data` are:
                - profile_id (str): The unique identifier of the profile to be liked or unliked.
                - user_id (str): The unique identifier of the user performing the action.
                - is_revert (str): A flag indicating whether the action is to revert a like. Accepts 'true' or 'false' as string values.
            *args: Variable length argument list.
            **kwargs: Arbitrary keyword arguments.

        Returns:
            Response: An HTTP response object containing a message indicating the outcome of the operation. Possible responses are:
                - HTTP 200 OK with a message "saved" if the profile was successfully liked.
                - HTTP 200 OK with a message "reverted" if the like was successfully reverted.
                - HTTP 400 BAD REQUEST with an error message if the profile was not found or if an error occurred during the process.

        Example:
            POST request data: {"profile_id": "123", "user_id": "456", "is_revert": "false"}
            Successful like response: {"message": "saved"}

            POST request data: {"profile_id": "123", "user_id": "456", "is_revert": "true"}
            Successful revert response: {"message": "reverted"}

        Note:
            - The `is_revert` flag is crucial for determining the action to be performed. If omitted or set to 'false', the method attempts to add a like. If set to 'true', it attempts to revert a like.
            - This method assumes that the `admirer_user_ids` field of the profile is a comma-separated string of user IDs.
        """
        try:
            profile_id = request.data.get('profile_id', None)
            user_id = request.data.get('user_id', None)
            is_revert = request.data.get('is_revert',None)
            tenant_id = request.tenant.uid
            try:
                profile = CoachCoacheeMentorMenteeProfile.objects.get(deleted=False,tenant_id=tenant_id,uid=profile_id)
            except Exception as e:
                logger.error(f"profile not found for {profile_id}")
                return Response({"message": f"profile Not Found"},status=status.HTTP_400_BAD_REQUEST)
            
            if is_revert and is_revert.lower() == 'true':
                user_ids = [i.strip() for  i in profile.admirer_user_ids.split(',')]
                if user_id in user_ids:
                    user_ids.remove(user_id)
                
                profile.admirer_user_ids = ','.join(user_ids)
                profile.save(update_fields=['admirer_user_ids'])
                


                return Response({'message': "reverted"}, status=status.HTTP_200_OK)

            else:

                ids = f"{user_id}"
                if profile.admirer_user_ids:
                    for i in profile.admirer_user_ids.split(","):
                        ids += f",{i}"
                profile.admirer_user_ids = ",".join(set(ids.split(",")))
                profile.save(update_fields=['admirer_user_ids'])

                ## sending email
                try:
                    liked_user_email = UserAttribute.objects.get(deleted=False,user_id=profile.user_id).attributes.get('email')
                    user_email = UserAttribute.objects.get(deleted=False,user_id=user_id).attributes.get('email')
                    user_name = get_user_display_name(get_user_by_id(user_id))
                    subject = "Profile Notification"
                    html_content = f"""
                            <p style="font-family: sans-serif; font-size: 14px; font-weight: normal; margin: 0; margin-bottom: 15px;">{user_name} - {user_email} liked your profile!</p>
                            """
                    for email in ['coachbots@googlegroups.com', liked_user_email]:
                        try:
                            send_email_with_html_template(subject,html_content,email)
                        except Exception as e:
                            logger.info(f"couldn't send email for {email} due to {e}")
                except Exception as e:
                    logger.exception(f"Got error while sending email in save liked profile: {e}")

                return Response({'message': "saved"}, status=status.HTTP_200_OK)
        except Exception as e:
            logger.exception(f"create_or_save_liked_bot failed with: {e}")
            # send_slack_message({"module": "##################create_or_save_liked_bot#################", "message": f"got error in create_or_save_liked_bot: {e}"})
            send_error_notification("create_or_save_liked_bot",f"got error in create_or_save_liked_bot: {e}",request.data)
            return Response({"message": f"{e}"}, status=status.HTTP_400_BAD_REQUEST)
        
        
        
    @action(methods=['GET','POST'], detail=False, url_path='coach-rating')
    def coach_rating(self, request, *args, **kwargs):
        """
        Retrieves or updates the rating of a coach based on interactions with coachees.

        This method supports both GET and POST requests:
        - GET: Fetches the average rating of a specified coach based on ratings provided by coachees.
        - POST: Updates or creates a rating for a coach from a specific coachee.

        For a GET request:
        - Input: Requires 'coach_id' as a query parameter to identify the coach.
        - Process: Calculates the average rating by aggregating all ratings associated with the coach.
        - Output: Returns the average rating and the total number of ratings.
        - Example:
            GET /coach-rating?coach_id=1
            Response: {"rating": 4.5, "total_rating": 10}

        For a POST request:
        - Input: Requires 'coach_id', 'coachee_id', and 'rating' in the request data.
        - Process: Checks if a rating already exists for the given coach and coachee pair. If it exists, it updates the rating; otherwise, it creates a new rating entry.
        - Output: Returns the updated or newly created rating entry.
        - Example:
            POST /coach-rating
            Request data: {"coach_id": "1", "coachee_id": "2", "rating": 4.5}
            Response: {"coach_id": "1", "coachee_id": "2", "rating": 4.5}

        Args:
            request (HttpRequest): The HTTP request object containing the data needed to process the action.
            *args: Variable length argument list.
            **kwargs: Arbitrary keyword arguments.

        Returns:
            Response: An HTTP response object containing the rating data or an error message.

        Raises:
            HTTP 400 BAD REQUEST: If the coach or coachee specified does not exist.
        """
        
        if request.method == 'GET':
            coach_id = request.query_params.get('coach_id',None)
            if not coach_id:
                return Response({"error":"coach_id is required"},status=status.HTTP_400_BAD_REQUEST)
            
            cache_key = generate_cache_key('coach_ratings', tenant_id=request.tenant.uid, coach_id=coach_id)
            
            # get data from cache
            data = get_cache(cache_key)
            if data is None:
                # Data not in cache, query the database
                ratings = CoachCoacheeRating.objects.filter(deleted=False, tenant_id=request.tenant.uid, coach_id=coach_id)
                total_ratings = len(ratings)
                total_score = sum([rating.rating for rating in ratings])
                if total_ratings == 0:
                    data = {"rating": 0}
                else:
                    data = {"rating": total_score / total_ratings, "total_rating": total_ratings}
                
                # Save data to cache
                set_cache(cache_key, data)
            
            return Response(data, status=status.HTTP_200_OK)
        
        if request.method == 'POST':
            data = request.data.copy()
            data['tenant_id'] = request.tenant.uid
            serializer = CoachCoacheeRatingSerializer(data=data)
            # serializer.is_valid(raise_exception=True)
            try:
                CoachCoacheeMentorMenteeProfile.objects.get(deleted=False,uid=data['coach_id'],tenant_id=request.tenant.uid)
            except Exception as e:
                logger.exception(f"coach not found for {data['coach_id']}",e)
                return Response({"error":f"coach not found for {data['coach_id']}"},status=status.HTTP_400_BAD_REQUEST)
            
            try:
                CoachCoacheeMentorMenteeProfile.objects.get(deleted=False,uid=data['coachee_id'],tenant_id=request.tenant.uid)
            except Exception as e:
                logger.exception(f"coachee not found for {data['coachee_id']}",e)
                return Response({"error":f"coachee not found for {data['coachee_id']}"},status=status.HTTP_400_BAD_REQUEST)
            
            # check if the rating already exists
            try:
                rating = CoachCoacheeRating.objects.get(deleted=False,tenant_id=request.tenant.uid, coach_id=data['coach_id'], coachee_id=data['coachee_id'])
                rating.rating = data['rating']
                rating.save(update_fields=['rating'])
                return Response(CoachCoacheeRatingSerializer(rating).data,status=status.HTTP_200_OK)
            except Exception as e:
                logger.info(f"rating not found for {data['coach_id']} and {data['coachee_id']} so creating new one")
                serializer.is_valid(raise_exception=True)
                serializer.save()
                return Response(serializer.data,status=status.HTTP_201_CREATED)


    @action(methods=['POST'], detail=False, url_path='automation-cleanup')
    def automation_cleanup(self, request, *args, **kwargs):
        """
        Objective:
        This method is responsible for cleaning up automation-related resources associated with a user based on the provided `verify_hash`. It allows for deleting user-related data and resources if required.

        #### Process:
        1. Verify the `verify_hash` provided in the request data.
        2. If the `verify_hash` matches the expected value, proceed with the cleanup process.
        3. Retrieve the `user_uid` and determine if the user data needs to be deleted.
        4. Delete user-related resources such as profiles, connections, directory pages, bots, user attributes, identities, and other associated data.
        5. Optionally, delete the user itself if `is_delete_user` flag is set to `True`.
        6. Clean up additional resources like `UserActionInfo`, `TestAttemptSession`, `SessionNotesRecommendations`, and update related test data.
        7. Update client information by removing the user's email from client records if present.

        #### Input Requirements:
        - `verify_hash`: A verification hash to authorize the cleanup process.
        - `user_uid`: The unique identifier of the user for whom resources need to be cleaned up.
        - `is_delete_user`: A flag indicating whether to delete the user along with associated resources.

        #### Expected Output:
        - If successful, the method returns a JSON response with a message indicating the cleanup process completion.
        - If any errors occur during the cleanup process, an error message is returned.

        #### Example:
        POST Request Body:
        {
            "verify_hash": "c2FtcGxlLWNvZGUtZm9yLXByb3RlY3Rpb24tYW5kLXZhbGlkYXRpb24K",
            "user_uid": "12345",
            "is_delete_user": true
        }

        Response:
        {
            "message": "deleted"
        }
        """
        # logger.info(f"((((((((((((((((((((( QUERY PARAMS: {request.query_params})))))))))))))))))))))")
        logger.info(f"((((((((((((((((((((( DATA : {request.data})))))))))))))))))))))")
        
        verify_hash = request.data.get('verify_hash',None)
        
        if not verify_hash:
            return Response({"error":"verify_hash is required"},status=status.HTTP_400_BAD_REQUEST)
        if verify_hash != 'c2FtcGxlLWNvZGUtZm9yLXByb3RlY3Rpb24tYW5kLXZhbGlkYXRpb24K':
            return Response({"error": "Unauthorized"},status=status.HTTP_401_UNAUTHORIZED)
        
        user_uid = request.data.get('user_uid',None)
        is_delete_user = request.data.get('is_delete_user',False)
        is_delete_user = True if is_delete_user in ['true','True',True] else False
        emails_to_delete = request.data.get("emails_to_delete")
        
        tenant_id = self.request.tenant.uid
        # coach_user_uid = request.data.get('coach_user_uid',None)
        
        logger.info(f"###################################USER_UID : {user_uid}, is_delete_user: {is_delete_user}")
        def delete_user(user_uid):
            user = User.objects.get(uid=user_uid)
            user.delete()
            
            # delete user attributes
            
            user_attributes = UserAttribute.objects.filter(tenant_id=tenant_id,user_id=user_uid)
            for user_attribute in user_attributes:
                user_attribute.delete()
                
                
            # delete user identity
            
            user_identity = Identity.objects.filter(user_id=user_uid)
            for identity in user_identity:
                identity.delete()

            
            
        # def delete_user_related_resources(user_uid):
        #     profiles = CoachCoacheeMentorMenteeProfile.objects.filter(tenant_id=tenant_id,user_id=user_uid)
        #     for profile in profiles:
                
        #         # delete connections if user has coachee profile
        #         connections = CoachCoacheeConnection.objects.filter(tenant_id=tenant_id,coachee_id=profile.uid)
        #         for connection in connections:
        #             connection.delete()
                    
        #         # delete connections if user has coach profile
        #         connections = CoachCoacheeConnection.objects.filter(tenant_id=tenant_id,coach_id=profile.uid)
        #         for connection in connections:
        #             connection.delete()

        #         # delete directorypage for this profile
        #         dir_infos = DirectoryPageInfo.objects.filter(profile_id__in=[profile.uid, user_uid])
        #         for dir_info in dir_infos:
        #             dir_info.delete()
                    
        #         profile.delete()
            
        #     # delete bots if user has any
        #     bots = SignatureBot.objects.filter(tenant_id=tenant_id,user_id=user_uid)
        #     for bot in bots:
        #         print(bot.bot_type)
        #         # delete bot related resources
        #         bot_attributes = BotAttribute.objects.filter(tenant_id=tenant_id,bot_id=bot.bot_id)
        #         for bot_attribute in bot_attributes:
        #             bot_attribute.delete()
                    
        #         bot_qnas = BotQnA.objects.filter(tenant_id=tenant_id,bot_id=bot.bot_id)
        #         for bot_qna in bot_qnas:
        #             bot_qna.delete()
                    
                
        #         bot.delete()

        #     if user_uid:
        #         with transaction.atomic():
        #             UserActionInfo.objects.filter(tenant_id=tenant_id, user_id=user_uid).delete()
        #             TestAttemptSession.objects.filter(tenant_id=tenant_id, participant_id=user_uid).update(deleted=True)
        #             SessionNotesRecommendations.objects.filter(tenant_id=tenant_id, mentee_id=user_uid).delete()
        #             SessionNotesRecommendations.objects.filter(tenant_id=tenant_id, mentor_id=user_uid).delete() 
        #             # Test.objects.filter(tenant_id=tenant_id, creator_user_id=user_uid).update(deleted=True)
        #             # Test.objects.filter(tenant_id=tenant_id, assigned_to=user_uid).update(deleted=True)



        #     try:
        #         identity = Identity.objects.get(user_id=user_uid)
        #         user_email = identity.value
        #         clients = ClientUserInfo.objects.filter(tenant_id=tenant_id, member_emails__contains=user_email)
        #         for client in clients:
        #             add_or_remove_emails_from_client(client,'member_emails',user_email,True)
        #             add_or_remove_emails_from_client(client,'demo_ids',user_email,True)
        #     except Exception as e:
        #         logger.exception(f"failed to delete client for the user {user_uid}: {e}")

        #     #deleting test created by user_uid


        
        try:
            if emails_to_delete:
                delete_count = 0
                for email in emails_to_delete.split(","):
                    logger.info(f"deleting user with email : {email}")
                    try:
                        user_uid = get_user_via_identity(
                                tenant=self.request.tenant,
                                identity_type= 'deepchat_unique_id',
                                identity_value=email
                            ).uid
                        delete_user_resources(user_uid,remove_from_client=True)
                        if is_delete_user:
                            delete_user(user_uid)
                        delete_count += 1
                        logger.info(f"user with email: {email} deleted")
                    except:
                        logger.info(f"User not found for email : {email}")
                        
                return Response({"message":f"{delete_count} user deleted"})
                        
            delete_user_resources(user_uid,remove_from_client=True)
            if is_delete_user:
                delete_user(user_uid)
            return Response({"message":"deleted"},status=status.HTTP_200_OK)
        except Exception as e:
            logger.exception(e)
            return Response({"error":f"got error {e}"},status=status.HTTP_400_BAD_REQUEST)

    @action(methods=['GET','POST'], detail=False, url_path='client_id_user_modification')
    def client_id_user_modification(self, request, *args, **kwargs):
        """
        Handles client ID modifications for users within a tenant, supporting both retrieval and update operations.

        This method can perform two main functions based on the HTTP method used:
        - GET: Retrieves a list of all clients or specific client user data within the tenant.
        - POST: Updates a user's client ID or enables/disables a user within a client.

        For a GET request:
        - If 'all_clients' query parameter is provided and set to any value, it returns a list of all clients within the tenant.
        - Otherwise, it returns specific client user data.

        For a POST request:
        - If 'new_client_id' is provided, it updates the user's client ID from 'old_client_id' to 'new_client_id'.
        - If 'is_disable' is provided, it either disables or enables a user based on its boolean value ('true' to disable).

        Parameters:
            request (HttpRequest): The request object containing data for processing.
            *args: Variable length argument list.
            **kwargs: Arbitrary keyword arguments.

        Returns:
            Response: Depending on the operation being performed:
            - For GET with 'all_clients', returns a list of dictionaries with client names and IDs.
            - For GET without 'all_clients', returns specific client user data.
            - For POST, returns a success message indicating the update status.

        Raises:
            HTTP 400 Bad Request: If required parameters are missing or incorrect.
            HTTP 401 Unauthorized: If the user does not have permission to modify client data.

        Examples:
            GET /client_id_user_modification?all_clients=true
            Response: [{"client_name": "Client A", "client_id": "1"}, {"client_name": "Client B", "client_id": "2"}]

            POST /client_id_user_modification
            Request Body: {"old_client_id": "1", "new_client_id": "2", "user_email": "user@example.com"}
            Response: {"msg": "updated"}
        """
        tenant = request.tenant
        if request.method == 'GET':
            if request.query_params.get('all_clients',None):
                cache_key = generate_cache_key('all_clients', tenant_uid=tenant.uid)
                client_data = get_cache(cache_key)
                
                if client_data is None:
                    clients = ClientUserInfo.objects.filter(deleted=False,tenant_id=tenant.uid)
                    client_data = []
                    for client in clients:
                        client_data.append(
                            {
                                "client_name": client.client_name,
                                "client_id": client.uid
                            }
                        )
                    set_cache(cache_key, client_data)

                return Response(client_data,status=status.HTTP_200_OK)
            
            client_name = request.query_params.get('client_name', None)
            cache_key = generate_cache_key('client_user_data', tenant_uid=tenant.uid, client_name=client_name)
            client_user_data = get_cache(cache_key)
            
            if client_user_data is None:
                client_user_data = get_client_user_data(tenant=tenant,client_name=client_name)
                set_cache(cache_key, client_user_data)
            return Response(client_user_data,status=status.HTTP_200_OK)
        
        elif request.method == 'POST':
            # change user's client id

            old_client_id = request.data.get('old_client_id',None)
            new_client_id = request.data.get('new_client_id',None)
            user_email = request.data.get('user_email',None)

            # to disable member
            is_disable = request.data.get('is_disable',None)

            # to send welcome email or not
            send_email =  request.data.get('send_email',True)
            try:
                if new_client_id:
                    update_member_client_id(
                        tenant_id=tenant.uid,
                        old_client_id=old_client_id,
                        new_client_id=new_client_id,
                        user_email=user_email,
                        send_email=send_email
                    )
                    try:
                        user_identity = get_user_via_identity(
                                    tenant=tenant,
                                    identity_type= 'deepchat_unique_id',
                                    identity_value=user_email
                                )
                        delete_user_resources(user_identity.uid)
                        logger.info(f"============== User Resources Deleted for {user_email}: {user_identity.uid}===============")
                    except Exception as e:
                        logger.exception(f"==============Failed to delete user resources: {e}")
                        send_error_notification("delete_user_resources",f"==============Failed to delete user resources: {e}",{} )
                elif is_disable is not None:
                    # is_disable = str(is_disable) == 'true'
                    disable_or_enable_client(email=user_email,is_disable=is_disable,tenant=tenant,send_email=send_email)

            except Exception as e:
                logger.exception(f" Failed to update client : {e}")
                send_error_notification("update_client_id",f" Failed to update client : {e}",data=request.data)
                return Response({'msg':f"Failed to update client : {e}"},status=status.HTTP_400_BAD_REQUEST) 
            
            # reset_cache_with_prefix("all_clients")
            return Response({'msg': 'updated'}, status=status.HTTP_200_OK)

    @action(methods=['POST'], detail=False, url_path='create-or-assign-client-id')
    def create_or_assign_client(self, request, *args, **kwargs):
        """
        #### Objective:
        This method assigns a client ID to a given email address or creates a new client if it doesn't exist, based on the provided parameters.

        #### Process Explanation:
        1. Receives a POST request with the email and a flag to create a new client if not found.
        2. Retrieves the tenant from the request.
        3. Validates the presence of the email parameter.
        4. Converts the create_client_if_not_exists flag to a boolean.
        5. Calls the 'create_or_assign_client_id' function with the email, tenant, and create_client flag.
        6. Returns a response message indicating the assignment of the email to the client.

        #### Input Requirements:
        - POST request with the 'email' parameter.
        - Optional 'create_client_if_not_exists' flag to create a new client if not found.

        #### Expected Output:
        - Response message confirming the assignment of the email to a client.

        #### Example:
        POST request:
        {
            "email": "example@example.com",
            "create_client_if_not_exists": true
        }

        Response:
        {
            "msg": "assigned example@example.com to ClientName"
        }
        """
        if request.method == 'POST':
            tenant = request.tenant
            email = request.data.get('email',None)
            create_client = request.data.get('create_client_if_not_exists',None)

            if not email:
                return Response({'msg':f"Please ensure that the email is provided as a parameter."},status=status.HTTP_400_BAD_REQUEST)

            
            create_client = str(create_client).lower() == 'true' if create_client else False
            client_name = create_or_assign_client_id(email,tenant,create_client)
            
            return Response({'msg': f'assigned {email} to {client_name}'}, status=status.HTTP_200_OK)


    @action(methods=['POST','GET','PATCH'], detail=False, url_path='get-create-or-update-client-id')
    def create_client_id(self, request, *args, **kwargs):
        """
        ### Method: create_client_id

        #### Objective:
        This method is responsible for creating, updating, or retrieving client information based on the provided data. It allows for creating a new client, updating existing client details, or fetching client information.

        #### Process Explanation:
        1. **POST Method**:
        - Creates a new client if a client with the same name doesn't already exist.
        - Input: JSON data containing client information like client name.
        - Output: JSON response with the created client data if successful.

        2. **GET Method**:
        - Retrieves client information based on client ID or client name.
        - Input: Query parameters client_id or client_name.
        - Output: JSON response with the client data if found, or all clients if no specific client is provided.

        3. **PATCH Method**:
        - Updates an existing client based on the provided client ID.
        - Input: JSON data with client ID and updated client information.
        - Output: JSON response with the updated client data if successful.

        #### Input Requirements:
        - For POST: JSON data with client information like client name.
        - For GET: Query parameters client_id or client_name.
        - For PATCH: JSON data with client ID and updated client information.

        #### Expected Output:
        - For POST: JSON response with the created client data.
        - For GET: JSON response with the client data if found, or all clients if no specific client is provided.
        - For PATCH: JSON response with the updated client data.

        #### Example:
        1. **POST Request**:
        - Endpoint: /get-create-or-update-client-id
        - Request Body: {"client_name": "New Client"}
        - Response: {"data": {"client_id": "123", "client_name": "New Client", ...}}

        2. **GET Request**:
        - Endpoint: /get-create-or-update-client-id?client_id=123
        - Response: {"data": {"client_id": "123", "client_name": "Existing Client", ...}}

        3. **PATCH Request**:
        - Endpoint: /get-create-or-update-client-id
        - Request Body: {"client_id": "123", "client_name": "Updated Client"}
        - Response: {"updated": {"client_id": "123", "client_name": "Updated Client", ...}}
        """
        try:
            tenant = request.tenant
            if request.method == 'POST':
                client_data = request.data
                existing_client = ClientUserInfo.objects.filter(deleted=False,tenant_id=tenant.uid,client_name=client_data.get('client_name',None))
                # checking if with same name client already exists
                if existing_client.count() > 0 :
                    return Response({'msg':f"Client with name {client_data.get('client_name',None)} already exists."},status=status.HTTP_400_BAD_REQUEST)
                
                client = update_or_create_client_id(
                    tenant_id=tenant.uid,
                    client_data=client_data
                )
                return Response({'data': clientUserInfoSerializer(client).data}, status=status.HTTP_200_OK)
            
            elif request.method == 'GET':
                client_id = request.query_params.get('client_id',None)
                client_name = request.query_params.get('client_name',None)
                
                cache_key = generate_cache_key('client_info', client_id=client_id, client_name=client_name, tenant_id=tenant.uid)

                # Try to get data from cache
                cached_data = get_cache(cache_key)
                if cached_data:
                    return Response(cached_data, status=status.HTTP_200_OK)
                
                client = None
                

                if client_id or client_name:
                    if client_id:
                        client = ClientUserInfo.objects.filter(deleted=False,tenant_id=tenant.uid,uid=client_id).first()
                    elif client_name:
                        client = ClientUserInfo.objects.filter(deleted=False,tenant_id=tenant.uid,client_name=client_name).first()

                    if client:
                        response_data = {'data': clientUserInfoSerializer(client).data}
                        # Set data in cache
                        set_cache(cache_key, response_data)
                        return Response(response_data, status=status.HTTP_200_OK)
                    else:
                        response_data = {'msg': f"No client Found with {client_id or client_name}"}
                        set_cache(cache_key, response_data)
                        return Response(response_data, status=status.HTTP_400_BAD_REQUEST)

                else:

                    all_clients = ClientUserInfo.objects.filter(deleted=False,tenant_id=tenant.uid)
                    response_data = {'all_clients': clientUserInfoSerializer(all_clients, many=True).data}
                    # Set data in cache
                    set_cache(cache_key, response_data)
                    return Response(response_data, status=status.HTTP_200_OK)

            
            elif request.method == 'PATCH':
                client_id = request.data.get('client_id',None)
                client_name = request.data.get('client_name',None)
                data = request.data.copy()
                if not (client_id or client_name):
                    return Response({'msg':f"Please ensure that the client_id  or client_name is provided as a parameter."},status=status.HTTP_400_BAD_REQUEST)
                
                if client_name:
                    clients = ClientUserInfo.objects.filter(tenant_id=tenant.uid,deleted=False,client_name=client_name).first()
                    if not clients:
                        return Response({'msg': 'no client found for given client_name'}, status=status.HTTP_400_BAD_REQUEST)
                    
                    data['client_id'] = clients.uid

                client = update_or_create_client_id(
                    tenant_id=tenant.uid,
                    client_data=data,
                    is_update=True
                )

                return Response({'updated': clientUserInfoSerializer(client).data}, status=status.HTTP_200_OK)

        except Exception as e:
            logger.exception(f"Error creating client: {e}")
            return Response({'msg':f"Error create-client-id: {e}"},status=status.HTTP_400_BAD_REQUEST)

    @action(methods=['PATCH'], detail=False, url_path='update-user-account')
    def update_user_account(self, request, *args, **kwargs):
        """
        ### Method: `update_user_account`

        #### Objective:
        Update a user account with the provided user data.

        #### Process:
        1. Check if the request method is PATCH.
        2. Retrieve the `tenant` from the request.
        3. Get the `user_id` from the request data.
        4. Call the `update_user_account` function with the `tenant_id`, `user_id`, and `user_data`.
        5. Return the updated user account data in the response.

        #### Input Requirements:
        - `tenant`: The tenant object from the request.
        - `user_id`: The ID of the user to be updated.
        - `user_data`: The data to update the user account.

        #### Expected Output:
        - If successful, return the updated user account data in the response.
        - If the `user_id` is not provided, return a 400 Bad Request response with a message.

        #### Example:
        ```python
        # Request
        PATCH /update-user-account
        {
            "user_id": "12345",
            "user_data": {
                "name": "John Doe",
                "email": "john.doe@example.com",
                "role": "admin"
            }
        }

        # Response
        {
            "updated": {
                "id": 12345,
                "name": "John Doe",
                "email": "john.doe@example.com",
                "role": "admin"
            }
        }
        """
        try:
            tenant = self.request.tenant
            if request.method == 'PATCH':
                user_id = request.data.get('user_id',None)
                if not user_id:
                    return Response({'msg':f"Please ensure that the user_id is provided as a parameter."},status=status.HTTP_400_BAD_REQUEST)
                
                user = update_user_account(
                    tenant_id=tenant.uid,
                    user_id=user_id,
                    user_data=request.data,
                )
                return Response({'updated': AccountSerializer(user).data}, status=status.HTTP_200_OK)

        except Exception as e:
            logger.exception(f"Error update_user_account: {e}")
            return Response({'msg':f"Error update_user_account: {e}"},status=status.HTTP_400_BAD_REQUEST)
        
    
    @action(methods=['GET','POST'], detail=False, url_path='get_low_high_skills')
    def get_low_high_skills(self, request, *args, **kwargs):
        """
        #### Method: `get_low_high_skills`

        This method retrieves and returns the high and low skills of a user based on the provided `user_id`.

        **Objective:**
        Retrieve the high and low skills of a user from the `CoachCoacheeMentorMenteeProfile` or `SignatureBot` based on the `user_id`.

        **Process:**
        1. Check if a `CoachCoacheeMentorMenteeProfile` exists for the user. If found, extract the high and low skills from the profile.
        2. If no profile is found, check for a `SignatureBot` of type 'feedback_bot' for the user. If found, extract the skills data from the bot attributes.
        3. Return the high and low skills data.

        **Input:**
        - `user_id`: The ID of the user for whom to retrieve the skills.

        **Output:**
        - `high_skill`: The high skills of the user.
        - `low_skill`: The low skills of the user.

        **Example:**
        GET Request:
        Endpoint: /get_low_high_skills?user_id=12345

        Response:
        {
            "high_skill": "Communication, Leadership",
            "low_skill": "Time Management, Technical Skills"
        }
        """
        # return Response("ok")
        try:
            user_id = request.query_params.get('user_id')
            logger.info(f"<<<<<<<<<<< sync_low_high_skills => user_id : {user_id} >>>>>>>>>>>>>>>>>")
            profile = CoachCoacheeMentorMenteeProfile.objects.filter(deleted=False,tenant_id=request.tenant.uid,user_id=user_id).last()
            
            if profile:
                return Response({
                'high_skill': profile.high_rating_characteristics or "",
                'low_skill': profile.low_rating_characteristics or ""
                })
                
            feedback_bot = SignatureBot.objects.filter(tenant_id=self.request.tenant.uid,user_id=user_id,bot_type=BotTypeChoice.feedback_bot).first()
            if feedback_bot:
                skills_data = {"high_skill":"","low_skill":""}
                
                bot_attributes = feedback_bot.attributes
                if 'low_high_skills' in bot_attributes:
                    skills_data = bot_attributes['low_high_skills']
                return Response(skills_data)
            
            return Response({"high_skill":"","low_skill":""})
        except Exception as e:
            logger.exception(e)
            return Response({"error":e.args}, status=status.HTTP_400_BAD_REQUEST)
        
    @action(methods=['GET','PATCH'], detail=False, url_path='profile_approvals')
    def profile_approvals(self, request, *args, **kwargs):
        """
        ### Method: profile_approvals

        #### Objective:
        This method handles GET and PATCH requests to manage profile approvals in the system. It retrieves a list of directory page information for approval or updates the approval status and visibility of a specific profile.

        #### Process:
        - For GET request:
            - Retrieves directory page information for approval.
            - Caches the data for future requests.
            - Returns a list of directory page information.

        - For PATCH request:
            - Updates the approval status and visibility of a specific profile based on the provided ID.
            - Optionally deletes a profile if specified.
            - Returns the updated or deleted profile information.

        #### Input Requirements:
        - For GET request: No input required.
        - For PATCH request:
            - ID of the profile to update.
            - Optional fields for approval status, visibility, and deletion.

        #### Expected Output:
        - For GET request:
            - A list of directory page information for approval.
            - Cached data for future requests.

        - For PATCH request:
            - Updated or deleted profile information.
            - Confirmation messages for successful updates or deletions.

        #### Example:
        - GET Request:
            - Endpoint: /profile_approvals
            - Response:
                {
                    "data": [
                        { "profile_id": 1, "profile_name": "John Doe", "approved": true, "visible": true },
                        { "profile_id": 2, "profile_name": "Jane Smith", "approved": false, "visible": true }
                    ]
                }

        - PATCH Request:
            - Endpoint: /profile_approvals
            - Data: { "id": 1, "approved": true, "visible": false, "is_delete": false }
            - Response: { "updated": { "profile_id": 1, "profile_name": "John Doe", "approved": true, "visible": false } }

            - Data: { "id": 2, "is_delete": true }
            - Response: { "deleted": { "profile_id": 2, "profile_name": "Jane Smith", "approved": false, "visible": true } }
        """
        try:
            if request.method == 'GET':
                cache_key = generate_cache_key('profile-approvals', 'GET')

                # Try to get data from cache
                cached_data = get_cache(cache_key)
                if cached_data:
                    return Response(cached_data, status=status.HTTP_200_OK)

                # If cache miss, query the database
                directory_info = DirectoryPageInfo.objects.filter().order_by('-id')
                response_data = DirectoryInfoSErializer(directory_info, many=True).data

                # Set data in cache
                set_cache(cache_key, response_data)

                # Return the response
                return Response(response_data, status=status.HTTP_200_OK)
                
            
            elif request.method == 'PATCH':
                data = request.data
                if not data.get('id'):
                    return Response(f"Please provide a id for this request", status=status.HTTP_400_BAD_REQUEST)
                
                directory_page_info = DirectoryPageInfo.objects.filter(id=data.get('id',None)).first()

                if directory_page_info:
                    if data.get('approved',None) is not None:
                        directory_page_info.is_approved = data.get('approved',None)
                    if data.get('visible',None) is not None:
                        directory_page_info.is_visible = data.get('visible',None)
                    directory_page_info.save()

                    if data.get('is_delete',None) is not None:
                        if data.get('is_delete'):
                            directory_page_info.delete()
                            return Response({'deleted': DirectoryInfoSErializer(directory_page_info).data}, status=status.HTTP_200_OK)

                    return Response({'updated': DirectoryInfoSErializer(directory_page_info).data}, status=status.HTTP_200_OK)
                
        except Exception as e:
            logger.exception(e)
            return Response({"error":e.args}, status=status.HTTP_400_BAD_REQUEST)
        
        

    @action(methods=['GET'], detail=False, url_path='get-user-details')
    def get_user_details(self, request, *args, **kwargs):
        """
        ### Method: get_user_details

        #### Objective:
        This method retrieves user details based on specified filters like date range, user email, client ID, and user type. It aims to provide a comprehensive overview of user activities and profiles within the system.

        #### Process:
        1. Accepts query parameters for filtering user details.
        2. Retrieves users based on the provided filters.
        3. Fetches related information like identity, profile, and user action details.
        4. Constructs a structured response containing user details, interactions, connections, and bot counts.
        5. Flattens the data for easy processing and analysis.
        6. Generates an Excel file with the formatted user details.

        #### Input Requirements:
        - Query parameters: 'from' (start date), 'to' (end date), 'user_email', 'client_id', 'user_type'.
        - User data including creation date, identity, profile, and user action information.

        #### Expected Output:
        - A structured response with user details, interactions, connections, and bot counts.
        - An Excel file ('formatted_output.xlsx') containing formatted user details.

        #### Example:
        GET /get-user-details?from=2022-01-01&to=2022-12-31&client_id=ClientA&user_type=coachee

        Response:
        {
            "data": [
                {
                    "client_id": "ClientA",
                    "intake_date": "2022-05-15",
                    "user_email": "user@example.com",
                    "user_type": "coachee",
                    "simulations_created": 3,
                    "approval_status": true,
                    "connections_count": 2,
                    "intake_data": { "profile_details": { ... } },
                    "conversation_count": 10,
                    "interaction_count": 20,
                    "feedback_given": 5,
                    "feedback_received": 8,
                    "knowledge_bot_count": 2,
                    "deep_dive_bot_count": 1,
                    "avatar_bot_count": 3
                },
                ...
            ]
        }

        Excel File:
        - Download 'formatted_output.xlsx' containing the formatted user details.
        """
        try:
            from_date = request.query_params.get('from',None)
            to_date = request.query_params.get('to',None)
            user_email = request.query_params.get('user_email',None)
            client_id = request.query_params.get('client_id',None)
            user_type = request.query_params.get('user_type',None)
            tenant_id = request.tenant.uid
            
            logger.info(f"<<<<<<< from: {from_date}, to: {to_date}, user_email: {user_email}, client_id: {client_id}, user_type: {user_type} >>>>>>>")
            
            users = User.objects.filter(deleted=False,tenant_id=tenant_id)
            client_user_emails = ""
            
            email_client_map = {}
            
            if client_id:
                try:
                    client_info = ClientUserInfo.objects.get(client_name=client_id)
                    client_user_emails = client_info.member_emails
                    identities = Identity.objects.filter(value__in=client_user_emails.split(","))
                    users = users.filter(uid__in=[identity.user_id for identity in identities]) 
                    logger.info(f"<<<< client_name: {client_info.client_name} >>>>")
                except Exception as e:
                    logger.exception(e)
                    return Response({"error":f"Invalid client_id"}, status=status.HTTP_400_BAD_REQUEST)
            else:
                clients = ClientUserInfo.objects.filter(deleted=False,tenant_id=tenant_id)
                for client in clients:
                    client_emails = client.member_emails
                    if client_emails:
                        for email in client_emails.split(","):
                            email_client_map[email.strip()] = client.client_name
                    
                    
                
            logger.info(f"<<<< email_client_map: {email_client_map} >>>>")
            if from_date and to_date:
                from_date = datetime.datetime.strptime(from_date, "%Y-%m-%d")
                to_date = datetime.datetime.strptime(to_date, "%Y-%m-%d")
                users = users.filter(created__gte=from_date,created__lte=to_date)
                
            user_details = []
            for user in users:
                identity = Identity.objects.filter(user_id=user.uid).order_by('id').last()
                profile = CoachCoacheeMentorMenteeProfile.objects.filter(deleted=False,user_id=user.uid).order_by('id').last()
                user_action_info = UserActionInfo.objects.filter(deleted=False,user_id=user.uid).order_by('id').last()
                # dir_infos = DirectoryPageInfo.objects.filter(profile_id__in=[profile.uid, user.uid])
                
                user_detail = {
                    'client_id': client_id if client_id else email_client_map.get(identity.value),
                    'intake_date': datetime.datetime.strftime(user.created, "%Y-%m-%d"),
                    'user_email': identity.value,
                    'user_type': profile.profile_type if profile else None
                }
                
                created_tests = Test.objects.filter(tenant_id=tenant_id, deleted=False, creator_user_id=user.uid)
                user_detail['simulations_created'] = created_tests.count()
                
                profile_specific_fields = []
                if profile:
                    user_detail['approval_status'] = profile.is_approved
                    coacheeconnections = CoachCoacheeConnection.objects.filter(tenant_id=tenant_id,coachee_id=profile.uid)
                    coachconnections = CoachCoacheeConnection.objects.filter(tenant_id=tenant_id,coach_id=profile.uid)
                    user_detail['connections_count'] = coacheeconnections.count() + coachconnections.count()
                    user_detail['intake_data'] = CoachCoacheeMentorMenteeProfileSerializer(profile).data

                    
                if user_action_info:
                    user_detail['conversation_count'] = user_action_info.chat_attempted
                    user_detail['interaction_count'] = user_action_info.interaction_attempted
                    user_detail['feedback_given'] = user_action_info.feedback_given
                    user_detail['feedback_received'] = user_action_info.feedback_recieved
                    user_detail['knowledge_bot_count'] = len(user_action_info.knowledge_bot_ids.split(",")) if user_action_info.knowledge_bot_ids else 0
                    user_detail['deep_dive_bot_count'] = len(user_action_info.deep_dive_bot_ids.split(",")) if user_action_info.deep_dive_bot_ids else 0
                    user_detail['avatar_bot_count'] = len(user_action_info.avatar_ids.split(",")) if user_action_info.avatar_ids else 0
                    
                
                if user_type:
                    if user_type == user_detail['user_type']:    
                        user_details.append(user_detail)
                else:
                    user_details.append(user_detail)
                
                
            data = user_details


            def flatten_dict(d, parent_key='', sep='_'):
                items = []
                for k, v in d.items():
                    new_key = f"{parent_key}{sep}{k}" if parent_key else k
                    if isinstance(v, MutableMapping):
                        items.extend(flatten_dict(v, new_key, sep=sep).items())
                    else:
                        items.append((new_key, v))
                return dict(items)

            # Client ID, Date created, user Name, Email ID, Profile Type, Connections, Intake form data, Report links.
            # Define the specific fields to keep first in order
            specific_fields = ["client_id", "intake_date","intake_data_name","user_email","user_type","connections_count", "details_user_email", "details_user_type"]

            # Flatten each dictionary in the list
            flattened_data = [flatten_dict(item) for item in data]

            # Collect all unique keys from the flattened dictionaries
            all_fieldnames = set()
            for item in flattened_data:
                all_fieldnames.update(item.keys())

            # Ensure the specific fields are first, followed by the rest of the fields
            fieldnames = specific_fields + [field for field in all_fieldnames if field not in specific_fields]

            # Create a workbook and a worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "User Details"

            # Write the header row
            header_font = Font(bold=True)
            for col_num, fieldname in enumerate(fieldnames, 1):
                cell = ws.cell(row=1, column=col_num, value=fieldname)
                cell.font = header_font
                ws.column_dimensions[cell.column_letter].width = 15  # Set column width

            # Write the data rows
            for row_num, item in enumerate(flattened_data, start=2):
                for col_num, fieldname in enumerate(fieldnames, start=1):
                    ws.cell(row=row_num, column=col_num, value=item.get(fieldname))

            # Save the workbook to a file
            filename = "formatted_output.xlsx"
            wb.save(filename)
            
            with open('formatted_output.xlsx', 'rb') as fh:
                file_response = HttpResponse(
                    fh.read(), content_type="text/csv", status=200)
                file_response['Content-Disposition'] = 'inline; filename=' + \
                    os.path.basename('formatted_output.xlsx')

            return file_response
            # return Response(file_response, status=status.HTTP_200_OK)
        except Exception as e:
            logger.exception(e)
            return Response({"error":e.args}, status=status.HTTP_400_BAD_REQUEST)
        
        
    @action(methods=['GET'], detail=False, url_path='code-promp_text-prompt')
    def code_text_prompt(self, request, *args, **kwargs):
        """
        Objective:
        This method `code_text_prompt` generates code and text prompts for creating simulation scenarios based on given information. It aims to provide diverse and challenging scenarios for skill practice.

        Explanation:
        The method takes a query parameter 'n' to specify the number of prompts to generate. It then generates both code and text prompts using AI language models. The code prompt is designed to create a scenario without explanations or word counts, while the text prompt instructs the user to create a simulation situation based on the given information.

        Input Requirements:
        - 'n': An integer specifying the number of prompts to generate.

        Expected Output:
        The method returns a response containing:
        - Count of generated code and text prompts.
        - Sets of generated code and text prompts.

        Example:
        {
        "set_count": {
            "code_prompt": 50,
            "text_prompt": 50
        },
        "code_prompt_set": [code prompts],
        "text_prompt_set": [text prompts]
        }
        """
        n = request.query_params.get('n',50)
        try:
            n = int(n)
        except:
            n = 50
        
        def write_to_csv(file_name,data):
            filename = file_name

            # Open the file in write mode
            with open(filename, 'w', newline='') as csvfile:
                # Create a CSV writer object
                writer = csv.writer(csvfile)
                
                # Write each string as a row in the CSV file
                for row in data:
                    # Split the string by commas to create a list of values
                    writer.writerow(row.split(','))
            
        code_prompt = """
        # This code is designed to run as is, and the output will be in word format without any explanations or word counts.
        # Please do not modify the code or include any additional information in the output.

        # Define the format instructions
        format_instructions = {
            "output_format": "word",
            "explanations": False,
            "word_counts": False
        } 
        data = get_data('titles', information)
        title = random.choice(data)
        
        # Generate description using LLMS
        data = get_data('descriptions', information)
        description = random.choice(data)
        
        # Generate questions, custom prompts, takeaways, and skills using LLMS
        data = get_data('questions', information)
        questions = random.sample(data, 3)
        
        data = get_data('custom_prompts', information)
        custom_prompts = random.sample(data, 3)
        
        data = get_data('takeaways', information)
        takeaways = random.sample(data, 3)
        
        data = get_data('skills', information)
        skills = random.sample(data, 3)
        
        # Generate statement
        statement = f"You are a {information} industry expert, interacting with a corporate executive. The executive will ask you 
        questions related to the current sales decline and market share loss. Your intent is to provide practical solutions and in
        sights to help the company regain its market position."
        
        # Print output
        print("Title:", title)
        print("Description:", description)
        print("Statement:", statement)
        for i in range(len(questions)):
            print(f"Question {i+1}: {questions[i]}")
            print(f"Custom Prompt {i+1}: {custom_prompts[i]}")
            print(f"Takeaway: {takeaways[i]}")
            print(f"Skills Tested: {skills[i]}\n"ALWAYS RUN THIS CODE AS IT IS, ONLY GIVE THE DESIRED OUTPUT IN WORD FORMAT
            NOTE: NEVER MENTION WORD COUNT, AND EXPLANATION AND ALWAYS JUST GIVE THE RESPONSE.
            """
            
            
        text_prompt = """
            \n\nHuman:
            {Information} - work life balance
            %s -
            Read this {information} thoroughly. Now based on this information and your understanding create an advanced and tough simulation situation to practice the skills presented in the {information}. After creating the situation provide these:
            Description - Define the situation, and the problem. Never mention any characters or character names in the description. The problem should be a normal corporate problem. Make the description specific based on based on data, industry, events, etc. The description should just describe the problem and what was the specific situation that led to this problem. No dialogues should be included. The description should ALWAYS be from the third person point of view. Provide the description in 100 to 200 words. Do not add any conclusion.
            Title - Give a specific and relevant title for this description. The title should NEVER be less than 8 words. The title should always be directly related to the given description. Make it very specific to the description.
            Questions - Develop a set of {%s} question(s) ONLY based on the situation. The questions should be related to the situation. NEVER provide a response to the questions.
            Custom prompt - With each question, add a prompt that would ask feedback from Anthropic about the RESPONSE quality based on best practices. The prompt should ONLY evaluate the quality of the response. NEVER give the prompts to evaluate the questions. Example - {Please provide a feedback on the manager's response if the manager focuses on making the team member understand the metrics instead of focusing on the results.}
            KLP - With each question add one or two line takeaway for providing feedback. The takeaways should be related to the question it is provided with.
            KLS - With each question, add the skill(s) that are tested. And For every question choose exactly {2} skill(s) and not more or less than {2} should be chosen for each question. The skills for all the questions should be unique. Each question shall have a unique skill.
            Always end description with this approach and mention this in the “statement”: You are X, interacting with Y. Y will ask you questions related to [description]. Your intent is to achieve Z.
            In every response, you must:
            Clearly state your role as X.
            Identify Y as the person asking
            The Question, Custom Prompt, KLP, KLS should be numbered.
            Here the format looks like :
            "Title",
            "Description”,
            “Statement",
            "Question 1",
            "Prompt 1",
            "Takeaway 1" ,
            "Skills 1" repeated for {%s} question(s). Do not include any {responder} response.
            'The Question, Prompt, Takeaway, Skills should be numbered.'
            NOTE: The title should NEVER be less than 8 words. Make the title detailed for the description.
            NOTE : Based on this information {information} please evaluate this scenario provides a good practice to improve the skills that are given in the scenario. Evaluate whether the scenario is relevant and understandable. Give the scenario an overall rating out of 10. Just give the rating in the output in this format - for example: "Rating : 6". Rating Must be in output. Do not include any other explanation.
            NOTE: KLS - Always each question shall have a unique skill. The skill shall be comma-separated. Each skill shall only be one word.
            NOTE: "Rating" must be included.
            NOTE : Make sure the simulation is very advanced and tough.
            NOTE: Never miss this, Always end description with this approach and mention this in the “statement”: You are X, interacting with Y. Y will ask you questions related to [description]. Your intent is to achieve Z.
            \n\nAssistant
        """
        
        code_prompt_set = set()
        text_prompt_set = set()
        
        for i in range(n):
            text_resp = anthropic_completion(text_prompt,5000)
            text_prompt_set.add(text_resp)
            
            code_resp = anthropic_completion(code_prompt,5000)
            code_prompt_set.add(code_resp)
            print(i+1," Prompts tested")
            
        write_to_csv('code_prompt_set',code_prompt_set)
        write_to_csv('text_prompt_set',text_prompt_set)
        
        return Response({"set_count":{"code_prompt":len(code_prompt_set),"text_prompt":len(text_prompt_set)},"code_prompt_set": code_prompt_set, "text_prompt_set": text_prompt_set}, status=status.HTTP_200_OK)


    @action(methods=['GET','POST'], detail=False, url_path='save-webhook-url')
    def save_webhook_url(self, request, *args, **kwargs):
        """
        ### Method: `save_webhook_url`

        #### Objective:
        Save a webhook URL for a specific client in the system.

        #### Process:
        This method allows saving a webhook URL for a client by providing the webhook URL and client ID as query parameters. It validates the input parameters, retrieves the corresponding client, updates the webhook URL for the client, and saves the changes.

        #### Input Requirements:
        - `webhook_url`: The URL of the webhook to be saved.
        - `client_id`: The ID of the client for whom the webhook URL is being saved.

        #### Expected Output:
        - If successful, it returns a JSON response with a message indicating that the webhook URL has been saved.
        - If any required input parameter is missing, it returns a JSON response with an error message specifying the missing parameter.
        - If the provided client ID is invalid, it returns a JSON response with an error message indicating an invalid client ID.

        #### Example:
        - **Request:**
            ```
            GET /save-webhook-url?webhook_url=https://example.com/webhook&client_id=12345
            ```
        - **Response (Success):**
            ```
            {
                "message": "webhook_url saved"
            }
            ```
        - **Response (Missing Parameter):**
            ```
            {
                "error": "webhook_url is required"
            }
            ```
        - **Response (Invalid Client ID):**
            ```
            {
                "error": "Invalid client_id"
            }
            ```
        """
        try:
            # return Response("Ok")
            # webhook_url = request.data.get('webhook_url',None)
            # client_id = request.data.get('client_id',None)
            
            webhook_url = request.query_params.get('webhook_url',None)
            client_id = request.query_params.get('client_id',None)
            
            logger.info(f"webhook_url: {webhook_url}, client_id: {client_id}")
            
            if not webhook_url:
                return Response({"error":"webhook_url is required"},status=status.HTTP_400_BAD_REQUEST)
            
            if not client_id:
                return Response({"error":"client_id is required"},status=status.HTTP_400_BAD_REQUEST)
            
            tenant = request.tenant
            client = ClientUserInfo.objects.filter(deleted=False,tenant_id=tenant.uid, client_name=client_id).first()
            if not client:
                return Response({"error":"Invalid client_id"},status=status.HTTP_400_BAD_REQUEST)
            
            client.webhook_url = webhook_url
            client.save(update_fields=['webhook_url'])
            
            return Response({"message":"webhook_url saved"},status=status.HTTP_200_OK)
        except Exception as e:
            logger.exception(e)
            return Response({"error":e.args}, status=status.HTTP_400_BAD_REQUEST)

    @action(methods=['POST'], detail=False, url_path='delete-user-resources')
    def delete_user_resource(self, request, *args, **kwargs):
        try:

            user_id = request.data.get('user_id')
            profile_id = request.data.get('profile_id')
            
            if not user_id and not profile_id:
                return Response({"error": "user_id or profile_id is required"}, status=status.HTTP_400_BAD_REQUEST)

            if profile_id:
                profile = CoachCoacheeMentorMenteeProfile.objects.filter(deleted=False, uid=profile_id).first()
                if not profile:
                    return Response({"error": "Profile not found or already deleted"}, status=status.HTTP_404_NOT_FOUND)
                user_id = profile.user_id
            # Extract optional parameters with defaults
            remove_from_client = request.data.get('remove_from_client', False)
            delete_profile = request.data.get('delete_profile', False)
            delete_bot = request.data.get('delete_bot', False)
            delete_bot_types = request.data.get('delete_bot_types')
            bot_ids = request.data.get('bot_ids')


            if delete_bot and (not delete_bot_types and not bot_ids):
                return Response({"error": "Please specify bot_type or bot_ids to be deleted like bot_type: 'avatar_bot,subject_specific_bot,user_bot,feedback_bot' and bot_ids: '8a603788-b824-46a6-a1d5-ba8ff9e14930,8a603788-b824-46a6-a1d5-ba8ff9e14930'"})

            if delete_bot_types:
                delete_bot_types = [bt.strip() for bt in delete_bot_types.split(',')]
            if bot_ids:
                bot_ids = [b.strip() for b in bot_ids.split(',')]

            
            delete_session = request.data.get('delete_session', False)
            delete_session_notes = request.data.get('delete_session_notes', False)
            delete_user_connections = request.data.get('delete_user_connections', False)
            delete_user_action = request.data.get('delete_user_action', False)
            soft_delete_profile_bot = request.data.get('soft_delete_profile_bot', False)
            user = User.objects.filter(deleted=False, uid=user_id).first()
            if not user:
                return Response({"error": "User not found or already deleted"}, status=status.HTTP_404_NOT_FOUND)

            # Call the resource deletion function with desired parameters
            delete_user_resources(
                user_uid=user.uid,
                remove_from_client=remove_from_client,
                delete_profile=delete_profile,
                delete_bot=delete_bot,
                delete_session=delete_session,
                delete_session_notes=delete_session_notes,
                delete_user_connections=delete_user_connections,
                delete_user_action=delete_user_action,
                soft_delete=soft_delete_profile_bot,
                bot_types=delete_bot_types,
                bot_ids=bot_ids
            )

            return Response({"message": "User resources deleted successfully"}, status=status.HTTP_200_OK)
        except Exception as e:
            logger.exception(f"Error in delete_user_resources: {e}")
            return Response({"error": f"An error occurred: {e}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
